import os
import io
import csv
import unicodedata
from datetime import datetime, time, timezone
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

SAO_PAULO = ZoneInfo("America/Sao_Paulo")

def local_today():
    return datetime.now(SAO_PAULO).date()

def local_datetime(value):
    if not value:
        return None
    parsed = value if isinstance(value, datetime) else datetime.fromisoformat(str(value))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(SAO_PAULO)

def datetime_iso(value):
    parsed = local_datetime(value)
    return parsed.isoformat() if parsed else ""

def cents(value: str) -> int:
    try:
        normalized = (value or "0").strip().replace(".", "").replace(",", ".")
        return int(Decimal(normalized).quantize(Decimal("0.01")) * 100)
    except (InvalidOperation, ValueError):
        raise ValueError("Valor monetário inválido.")

def normalize_cpf(value):
    cpf = "".join(character for character in (value or "") if character.isdigit())
    if cpf and len(cpf) != 11:
        raise ValueError("O CPF deve possuir 11 números.")
    return cpf

def normalized_header(value):
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    return value.strip().lower().replace("-", "").replace("_", "").replace(" ", "")

def spreadsheet_rows(upload):
    extension = os.path.splitext(upload.filename or "")[1].lower()
    raw = upload.read()
    if extension == ".csv":
        text = raw.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        return list(csv.reader(io.StringIO(text), dialect))
    if extension == ".xlsx":
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    raise ValueError("Formato inválido. Envie uma planilha .xlsx ou .csv.")

def money(value):
    value = int(value or 0)
    return f"R$ {value / 100:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def brdate(value):
    try:
        return local_datetime(value).strftime("%d/%m/%Y %H:%M")
    except (ValueError, TypeError):
        return value

def cpfmask(value):
    return f"***.***.***-{value[-2:]}" if value else "—"

def month_bounds(month=None):
    month = month or local_today().strftime("%Y-%m")
    try:
        start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError:
        start = local_today().replace(day=1)
        month = start.strftime("%Y-%m")
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)
    start_utc = datetime.combine(start, time.min, SAO_PAULO).astimezone(timezone.utc).replace(tzinfo=None)
    end_utc = datetime.combine(end, time.min, SAO_PAULO).astimezone(timezone.utc).replace(tzinfo=None)
    return month, start_utc.isoformat(sep=" "), end_utc.isoformat(sep=" ")

def add_months(month, count):
    current = datetime.strptime(month, "%Y-%m").date()
    result = []
    for _ in range(count):
        result.append(current.strftime("%Y-%m"))
        current = current.replace(year=current.year + 1, month=1) if current.month == 12 else current.replace(month=current.month + 1)
    return result
