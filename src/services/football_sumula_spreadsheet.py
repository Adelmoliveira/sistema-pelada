"""Importação simples de participantes de súmulas históricas."""

from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PARTICIPANT_HEADERS = (
    "Partida", "Posição", "ID do peladeiro", "Nome do peladeiro", "Status", "Observação",
)
STATUS_MAP = {
    "CONFIRMADO": "CONFIRMADO", "AUSENTE": "AUSENTE",
    "DESISTENTE": "DESISTENTE", "RESERVA": "RESERVA",
}
SLOT_LIMITS = {"G": 2, "D": 8, "M": 6, "A": 6}
SLOT_START = {"G": 1, "D": 3, "M": 11, "A": 17}
PREFERRED_POSITION = {
    "G": "GOLEIRO", "D": "DEFENSOR", "M": "MEIO_CAMPO", "A": "ATACANTE",
}


def _key(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = "".join(char for char in value if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", value.strip()).upper()


def _text(value, limit=500):
    return str(value or "").strip()[:limit]


def _integer(value, label, row, errors, minimum, maximum):
    try:
        number = int(value)
        if isinstance(value, float) and value != number:
            raise ValueError
    except (TypeError, ValueError):
        errors.append(f"{label}, linha {row}: informe um número inteiro.")
        return None
    if number < minimum or number > maximum:
        errors.append(f"{label}, linha {row}: valor deve estar entre {minimum} e {maximum}.")
        return None
    return number


def _draw_order(match_number, slot, row, errors):
    slot_key = _key(slot).replace(" ", "")
    match = re.fullmatch(r"([GDMA])(\d+)", slot_key)
    if not match:
        errors.append(f"Participantes, linha {row}: posição inválida. Use G1, D1, M1 ou A1.")
        return None, None
    group, raw_number = match.groups()
    number = int(raw_number)
    if number < 1 or number > SLOT_LIMITS[group]:
        errors.append(
            f"Participantes, linha {row}: {group} aceita números de 1 a {SLOT_LIMITS[group]}."
        )
        return None, None
    local_order = SLOT_START[group] + number - 1
    return ((match_number - 1) * 22) + local_order, PREFERRED_POSITION[group]


def _player_resolver(players, errors):
    player_by_id = {int(player["id"]): player for player in players}
    player_by_name = {}
    for player in players:
        for name in (player["war_name"], player["name"]):
            if name:
                player_by_name.setdefault(_key(name), {})[int(player["id"])] = player

    def resolve(player_id, name, row):
        if player_id not in (None, ""):
            parsed_id = _integer(player_id, "Participantes", row, errors, 1, 999999999)
            player = player_by_id.get(parsed_id) if parsed_id else None
            if parsed_id and not player:
                errors.append(f"Participantes, linha {row}: ID {parsed_id} não encontrado.")
            return parsed_id if player else None
        normalized_name = _key(name)
        if not normalized_name:
            errors.append(f"Participantes, linha {row}: informe o ID ou o nome do peladeiro.")
            return None
        matches = player_by_name.get(normalized_name, {})
        if len(matches) == 1:
            return next(iter(matches))
        if not matches:
            errors.append(f"Participantes, linha {row}: peladeiro “{_text(name)}” não encontrado.")
        else:
            errors.append(f"Participantes, linha {row}: nome “{_text(name)}” é ambíguo; informe o ID.")
        return None

    return resolve


def _validate_participants(raw_rows, players, errors):
    resolve = _player_resolver(players, errors)
    participants, player_ids, orders = [], set(), set()
    for row, values in raw_rows:
        match_number = _integer(values[0], "Participantes", row, errors, 1, 3)
        draw_order, position = _draw_order(match_number, values[1], row, errors) if match_number else (None, None)
        player_id = resolve(values[2], values[3], row)
        status = STATUS_MAP.get(_key(values[4]) or "CONFIRMADO")
        if not status:
            errors.append(f"Participantes, linha {row}: status inválido.")
        if player_id in player_ids:
            errors.append(f"Participantes, linha {row}: peladeiro repetido.")
        if draw_order in orders:
            errors.append(f"Participantes, linha {row}: posição {_text(values[1])} repetida na partida {match_number}.")
        if player_id:
            player_ids.add(player_id)
        if draw_order:
            orders.add(draw_order)
        participants.append({
            "match": match_number, "slot": _key(values[1]).replace(" ", ""),
            "player_id": player_id, "status": status, "position": position,
            "order": draw_order, "observation": _text(values[5]),
        })
    if errors:
        raise ValueError("\n".join(errors[:30]) + ("\nHá outros erros; corrija os primeiros e tente novamente." if len(errors) > 30 else ""))
    if not participants:
        raise ValueError("Nenhum participante foi informado.")
    return {"participants": participants, "match_count": max(item["match"] for item in participants)}


def build_template(players):
    """Gera um Excel com apenas a aba necessária para importar participantes."""
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instruções"
    instructions.sheet_view.showGridLines = False
    instructions["A1"] = "Importar participantes da súmula"
    instructions["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="0B5E8E")
    instructions.merge_cells("A1:F1")
    notes = (
        "Preencha somente a aba Participantes. As demais informações da súmula podem ser registradas depois no sistema.",
        "Posições aceitas: G1–G2, D1–D8, M1–M6 e A1–A6.",
        "Use preferencialmente o ID disponível na aba Peladeiros. Se usar o nome, ele deve coincidir com o cadastro.",
        "As partidas 2 e 3 serão criadas automaticamente quando houver participantes nelas.",
        "Se qualquer linha estiver incorreta, nenhum participante será importado.",
    )
    for index, note in enumerate(notes, start=3):
        instructions.cell(index, 1, note)
        instructions.cell(index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        instructions.row_dimensions[index].height = 30
    instructions.column_dimensions["A"].width = 105

    sheet = workbook.create_sheet("Participantes")
    sheet.sheet_view.showGridLines = False
    sheet.append(PARTICIPANT_HEADERS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F301"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B5E8E")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 34
    for column, width in zip("ABCDEF", (12, 13, 18, 32, 16, 36)):
        sheet.column_dimensions[column].width = width
    for column, choices in (("A", "1,2,3"), ("B", "G1,G2,D1,D2,D3,D4,D5,D6,D7,D8,M1,M2,M3,M4,M5,M6,A1,A2,A3,A4,A5,A6"), ("E", "CONFIRMADO,AUSENTE,DESISTENTE,RESERVA")):
        validation = DataValidation(type="list", formula1=f'"{choices}"', allow_blank=False)
        validation.error = "Escolha uma das opções da lista."
        validation.errorTitle = "Valor inválido"
        validation.showErrorMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"{column}2:{column}301")

    lookup = workbook.create_sheet("Peladeiros")
    lookup.sheet_view.showGridLines = False
    lookup.append(("ID", "Nome de guerra", "Nome completo", "Posição"))
    for player in players:
        lookup.append((player["id"], player["war_name"] or "", player["name"], player["football_position"] or ""))
    for cell in lookup[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B5E8E")
    lookup.freeze_panes = "A2"
    lookup.auto_filter.ref = f"A1:D{max(2, lookup.max_row)}"
    for column, width in zip("ABCD", (12, 25, 38, 18)):
        lookup.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_import(file_stream, players):
    try:
        workbook = load_workbook(file_stream, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("O arquivo não é um Excel .xlsx válido.") from exc
    if "Participantes" not in workbook.sheetnames:
        raise ValueError("A aba obrigatória “Participantes” não foi encontrada.")
    sheet = workbook["Participantes"]
    actual = tuple(_text(cell.value) for cell in sheet[1][:len(PARTICIPANT_HEADERS)])
    if actual != PARTICIPANT_HEADERS:
        raise ValueError("Os títulos da aba “Participantes” foram alterados. Baixe um novo modelo.")
    rows = [
        (number, values)
        for number, values in enumerate(
            sheet.iter_rows(min_row=2, max_row=301, max_col=len(PARTICIPANT_HEADERS), values_only=True), start=2,
        )
        if any(value not in (None, "") for value in values)
    ]
    return _validate_participants(rows, players, [])


def parse_participant_text(raw_text, players):
    """Lê blocos no formato `1ª PARTIDA` e `D1: NOME`."""
    errors, rows, current_match = [], [], None
    for line_number, raw_line in enumerate((raw_text or "").splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        heading = re.fullmatch(r"(\d+)\s*[ªA]?\s*PARTIDA", _key(line))
        if heading:
            current_match = _integer(heading.group(1), "Texto", line_number, errors, 1, 3)
            continue
        participant = re.fullmatch(r"([GDMA]\s*\d+)\s*:\s*(.+)", line, flags=re.IGNORECASE)
        if not participant:
            errors.append(f"Texto, linha {line_number}: use o formato D1: NOME.")
            continue
        if not current_match:
            errors.append(f"Texto, linha {line_number}: informe antes o título da partida.")
            continue
        slot, name = participant.groups()
        rows.append((line_number, (current_match, slot, None, name, "CONFIRMADO", "Importado por texto")))
    return _validate_participants(rows, players, errors)


def import_into_sumula(db, sumula_id, data, user_id):
    occupied = int(db.execute("SELECT COUNT(*) FROM football_participants WHERE sumula_id=?", (sumula_id,)).fetchone()[0] or 0)
    matches = db.execute("SELECT * FROM football_matches WHERE sumula_id=? ORDER BY number", (sumula_id,)).fetchall()
    default_matches = len(matches) == 1 and int(matches[0]["number"]) == 1
    if occupied or not default_matches:
        raise ValueError("A importação só pode ser feita antes de adicionar participantes ou novas partidas manualmente.")
    for match_number in range(2, int(data["match_count"]) + 1):
        db.execute("INSERT INTO football_matches(sumula_id,number) VALUES(?,?)", (sumula_id, match_number))
    for item in data["participants"]:
        db.execute(
            "INSERT INTO football_participants(sumula_id,player_id,status,preferred_position,draw_order,observation) VALUES(?,?,?,?,?,?)",
            (sumula_id, item["player_id"], item["status"], item["position"], item["order"], item["observation"]),
        )
    return {"participants": len(data["participants"]), "matches": int(data["match_count"])}
