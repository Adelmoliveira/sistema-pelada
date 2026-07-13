from __future__ import annotations

import os
import sqlite3
import csv
import base64
import io
import unicodedata
from functools import wraps
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from flask import Flask, flash, g, jsonify, redirect, render_template, request, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash


app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao"),
    DATABASE=os.path.join(app.root_path, "bar.db"),
    MAX_CONTENT_LENGTH=5 * 1024 * 1024,
    PIX_KEY="adelmoliveira@gmail.com",
    PIX_MERCHANT_NAME="BAR PELADEIROS GPCTA",
    PIX_MERCHANT_CITY="SAO PAULO",
)


SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL UNIQUE COLLATE NOCASE,
    name TEXT NOT NULL,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN ('manager','staff','client')),
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS players (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    war_name TEXT DEFAULT '',
    cpf TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    emergency_phone TEXT DEFAULT '',
    email TEXT DEFAULT '',
    membership_type TEXT NOT NULL DEFAULT 'regular',
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    category TEXT NOT NULL,
    package_type TEXT NOT NULL DEFAULT '',
    units_per_case INTEGER NOT NULL DEFAULT 0 CHECK(units_per_case >= 0),
    price_cents INTEGER NOT NULL CHECK(price_cents >= 0),
    cost_cents INTEGER NOT NULL DEFAULT 0 CHECK(cost_cents >= 0),
    stock INTEGER NOT NULL DEFAULT 0 CHECK(stock >= 0),
    min_stock INTEGER NOT NULL DEFAULT 5 CHECK(min_stock >= 0),
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS sales (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id INTEGER NOT NULL REFERENCES players(id),
    payment_method TEXT NOT NULL CHECK(payment_method IN ('Pix','Dinheiro','Débito','Cortesia')),
    total_cents INTEGER NOT NULL,
    paid INTEGER NOT NULL DEFAULT 1,
    notes TEXT DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS sale_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sale_id INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
    product_id INTEGER NOT NULL REFERENCES products(id),
    quantity INTEGER NOT NULL CHECK(quantity > 0),
    unit_price_cents INTEGER NOT NULL,
    unit_cost_cents INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS restocks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    product_id INTEGER NOT NULL REFERENCES products(id),
    quantity INTEGER NOT NULL CHECK(quantity > 0),
    unit_cost_cents INTEGER NOT NULL DEFAULT 0,
    notes TEXT DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS membership_payments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    player_id INTEGER NOT NULL REFERENCES players(id),
    amount_cents INTEGER NOT NULL CHECK(amount_cents > 0),
    months_count INTEGER NOT NULL CHECK(months_count BETWEEN 1 AND 12),
    start_month TEXT NOT NULL,
    payment_method TEXT NOT NULL CHECK(payment_method IN ('Pix','Dinheiro','Débito')),
    notes TEXT DEFAULT '',
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS membership_months (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    payment_id INTEGER NOT NULL REFERENCES membership_payments(id) ON DELETE CASCADE,
    player_id INTEGER NOT NULL REFERENCES players(id),
    month TEXT NOT NULL,
    UNIQUE(player_id, month)
);
CREATE INDEX IF NOT EXISTS idx_sales_created ON sales(created_at);
CREATE INDEX IF NOT EXISTS idx_items_sale ON sale_items(sale_id);
"""


def db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_error=None):
    connection = g.pop("db", None)
    if connection is not None:
        connection.close()


@app.before_request
def load_user_and_protect_routes():
    g.user = None
    user_id = session.get("user_id")
    if user_id:
        g.user = db().execute("SELECT * FROM users WHERE id=? AND active=1", (user_id,)).fetchone()
        if not g.user:
            session.clear()
    public_endpoints = {"login", "setup", "static"}
    if request.endpoint in public_endpoints or request.endpoint is None:
        return None
    if not db().execute("SELECT 1 FROM users LIMIT 1").fetchone():
        return redirect(url_for("setup"))
    if not g.user:
        return redirect(url_for("login", next=request.path))


@app.context_processor
def inject_user():
    return {"current_user": g.get("user")}


def roles_allowed(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not g.user or g.user["role"] not in roles:
                flash("Seu usuário não possui acesso a essa funcionalidade.", "danger")
                return redirect(url_for("sale") if g.user and g.user["role"] == "client" else url_for("dashboard"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


@app.route("/setup", methods=["GET", "POST"])
def setup():
    if db().execute("SELECT 1 FROM users LIMIT 1").fetchone():
        return redirect(url_for("login"))
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        if len(username) < 3 or len(password) < 8:
            flash("Use um usuário com ao menos 3 caracteres e senha com ao menos 8.", "danger")
        elif password != request.form.get("password_confirm"):
            flash("As senhas não coincidem.", "danger")
        else:
            db().execute("INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'manager')",
                         (username, request.form["name"].strip(), generate_password_hash(password)))
            db().commit()
            flash("Gerente criado. Entre com seu usuário e senha.", "success")
            return redirect(url_for("login"))
    return render_template("setup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.user:
        return redirect(url_for("sale") if g.user["role"] == "client" else url_for("dashboard"))
    if request.method == "POST":
        user = db().execute("SELECT * FROM users WHERE username=? AND active=1",
                            (request.form["username"].strip(),)).fetchone()
        if user and check_password_hash(user["password_hash"], request.form["password"]):
            session.clear()
            session["user_id"] = user["id"]
            return redirect(url_for("sale") if user["role"] == "client" else url_for("dashboard"))
        flash("Usuário ou senha inválidos.", "danger")
    return render_template("login.html")


@app.post("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/users", methods=["GET", "POST"])
@roles_allowed("manager")
def users():
    conn = db()
    if request.method == "POST":
        try:
            username, password = request.form["username"].strip(), request.form["password"]
            if len(username) < 3 or len(password) < 8:
                raise ValueError("Usuário deve ter 3 caracteres e senha ao menos 8.")
            conn.execute("INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,?)", (
                username, request.form["name"].strip(), generate_password_hash(password), request.form["role"]))
            conn.commit()
            flash("Usuário criado.", "success")
        except (ValueError, sqlite3.IntegrityError) as exc:
            flash(f"Não foi possível criar o usuário: {exc}", "danger")
        return redirect(url_for("users"))
    rows = conn.execute("SELECT * FROM users ORDER BY active DESC,name").fetchall()
    return render_template("users.html", users=rows)


@app.post("/users/<int:user_id>/toggle")
@roles_allowed("manager")
def toggle_user(user_id):
    if user_id == g.user["id"]:
        flash("Você não pode desativar o próprio usuário.", "danger")
    else:
        db().execute("UPDATE users SET active=1-active WHERE id=?", (user_id,))
        db().commit()
        flash("Acesso do usuário atualizado.", "success")
    return redirect(url_for("users"))


def init_db():
    connection = sqlite3.connect(app.config["DATABASE"])
    migrate_payment_method(connection)
    connection.executescript(SCHEMA)
    columns = {row[1] for row in connection.execute("PRAGMA table_info(players)")}
    if "email" not in columns:
        connection.execute("ALTER TABLE players ADD COLUMN email TEXT DEFAULT ''")
        connection.commit()
    if "membership_type" not in columns:
        connection.execute("ALTER TABLE players ADD COLUMN membership_type TEXT NOT NULL DEFAULT 'regular'")
        connection.commit()
    if "war_name" not in columns:
        connection.execute("ALTER TABLE players ADD COLUMN war_name TEXT DEFAULT ''")
    if "emergency_phone" not in columns:
        connection.execute("ALTER TABLE players ADD COLUMN emergency_phone TEXT DEFAULT ''")
    if "cpf" not in columns:
        connection.execute("ALTER TABLE players ADD COLUMN cpf TEXT DEFAULT ''")
    connection.commit()
    connection.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_players_cpf ON players(cpf) WHERE cpf<>''")
    connection.commit()
    product_columns = {row[1] for row in connection.execute("PRAGMA table_info(products)")}
    if "package_type" not in product_columns:
        connection.execute("ALTER TABLE products ADD COLUMN package_type TEXT NOT NULL DEFAULT ''")
    if "units_per_case" not in product_columns:
        connection.execute("ALTER TABLE products ADD COLUMN units_per_case INTEGER NOT NULL DEFAULT 0")
    connection.commit()
    migrate_product_categories(connection)
    connection.close()


def migrate_payment_method(connection):
    """Troca a restrição antiga de Fiado por Débito preservando os dados."""
    row = connection.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='sales'"
    ).fetchone()
    if not row or "Fiado" not in (row[0] or ""):
        return
    connection.execute("PRAGMA foreign_keys = OFF")
    connection.executescript("""
        BEGIN;
        ALTER TABLE sale_items RENAME TO sale_items_old;
        ALTER TABLE sales RENAME TO sales_old;
        CREATE TABLE sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER NOT NULL REFERENCES players(id),
            payment_method TEXT NOT NULL CHECK(payment_method IN ('Pix','Dinheiro','Débito','Cortesia')),
            total_cents INTEGER NOT NULL,
            paid INTEGER NOT NULL DEFAULT 1,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
            product_id INTEGER NOT NULL REFERENCES products(id),
            quantity INTEGER NOT NULL CHECK(quantity > 0),
            unit_price_cents INTEGER NOT NULL,
            unit_cost_cents INTEGER NOT NULL DEFAULT 0
        );
        INSERT INTO sales(id,player_id,payment_method,total_cents,paid,notes,created_at)
        SELECT id,player_id,CASE WHEN payment_method='Fiado' THEN 'Débito' ELSE payment_method END,
               total_cents,1,notes,created_at FROM sales_old;
        INSERT INTO sale_items SELECT * FROM sale_items_old;
        DROP TABLE sale_items_old;
        DROP TABLE sales_old;
        CREATE INDEX IF NOT EXISTS idx_sales_created ON sales(created_at);
        CREATE INDEX IF NOT EXISTS idx_items_sale ON sale_items(sale_id);
        COMMIT;
    """)
    connection.execute("PRAGMA foreign_keys = ON")


def migrate_product_categories(connection):
    """Remove a lista fechada de categorias criada na primeira versão."""
    row = connection.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='products'"
    ).fetchone()
    if not row or "CHECK(category IN" not in (row[0] or ""):
        return
    connection.execute("PRAGMA foreign_keys = OFF")
    connection.executescript("""
        BEGIN;
        ALTER TABLE sale_items RENAME TO sale_items_category_old;
        ALTER TABLE restocks RENAME TO restocks_category_old;
        ALTER TABLE products RENAME TO products_category_old;
        CREATE TABLE products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            category TEXT NOT NULL,
            package_type TEXT NOT NULL DEFAULT '',
            units_per_case INTEGER NOT NULL DEFAULT 0 CHECK(units_per_case >= 0),
            price_cents INTEGER NOT NULL CHECK(price_cents >= 0),
            cost_cents INTEGER NOT NULL DEFAULT 0 CHECK(cost_cents >= 0),
            stock INTEGER NOT NULL DEFAULT 0 CHECK(stock >= 0),
            min_stock INTEGER NOT NULL DEFAULT 5 CHECK(min_stock >= 0),
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL REFERENCES sales(id) ON DELETE CASCADE,
            product_id INTEGER NOT NULL REFERENCES products(id),
            quantity INTEGER NOT NULL CHECK(quantity > 0),
            unit_price_cents INTEGER NOT NULL,
            unit_cost_cents INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE restocks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL REFERENCES products(id),
            quantity INTEGER NOT NULL CHECK(quantity > 0),
            unit_cost_cents INTEGER NOT NULL DEFAULT 0,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        INSERT INTO products(id,name,category,package_type,units_per_case,price_cents,cost_cents,stock,min_stock,active,created_at)
        SELECT id,name,category,package_type,units_per_case,price_cents,cost_cents,stock,min_stock,active,created_at
        FROM products_category_old;
        INSERT INTO sale_items SELECT * FROM sale_items_category_old;
        INSERT INTO restocks SELECT * FROM restocks_category_old;
        DROP TABLE sale_items_category_old;
        DROP TABLE restocks_category_old;
        DROP TABLE products_category_old;
        CREATE INDEX IF NOT EXISTS idx_items_sale ON sale_items(sale_id);
        COMMIT;
    """)
    connection.execute("PRAGMA foreign_keys = ON")


def cents(value: str) -> int:
    try:
        normalized = (value or "0").strip().replace(".", "").replace(",", ".")
        return int(Decimal(normalized).quantize(Decimal("0.01")) * 100)
    except (InvalidOperation, ValueError):
        raise ValueError("Valor monetário inválido.")


def normalize_cpf(value):
    cpf = "".join(character for character in (value or "") if character.isdigit())
    if cpf and len(cpf) != 11:
        raise ValueError("O CPF deve possuir 11 números.")
    return cpf


def normalized_header(value):
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode()
    return value.strip().lower().replace("-", "").replace("_", "").replace(" ", "")


def pix_text(value, limit):
    value = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode()
    return " ".join(value.upper().split())[:limit]


def pix_tlv(identifier, value):
    return f"{identifier}{len(value):02d}{value}"


def pix_crc16(payload):
    crc = 0xFFFF
    for byte in payload.encode("utf-8"):
        crc ^= byte << 8
        for _ in range(8):
            crc = ((crc << 1) ^ 0x1021) & 0xFFFF if crc & 0x8000 else (crc << 1) & 0xFFFF
    return f"{crc:04X}"


def pix_payload(amount_cents):
    merchant_account = pix_tlv("00", "br.gov.bcb.pix") + pix_tlv("01", app.config["PIX_KEY"])
    payload = "".join((
        pix_tlv("00", "01"),
        pix_tlv("26", merchant_account),
        pix_tlv("52", "0000"),
        pix_tlv("53", "986"),
        pix_tlv("54", f"{amount_cents / 100:.2f}"),
        pix_tlv("58", "BR"),
        pix_tlv("59", pix_text(app.config["PIX_MERCHANT_NAME"], 25)),
        pix_tlv("60", pix_text(app.config["PIX_MERCHANT_CITY"], 15)),
        pix_tlv("62", pix_tlv("05", "***")),
    )) + "6304"
    return payload + pix_crc16(payload)


def spreadsheet_rows(upload):
    extension = os.path.splitext(upload.filename or "")[1].lower()
    raw = upload.read()
    if extension == ".csv":
        text = raw.decode("utf-8-sig", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        return list(csv.reader(io.StringIO(text), dialect))
    if extension == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    raise ValueError("Formato inválido. Envie uma planilha .xlsx ou .csv.")


@app.template_filter("money")
def money(value):
    value = int(value or 0)
    return f"R$ {value / 100:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


@app.template_filter("brdate")
def brdate(value):
    try:
        return datetime.fromisoformat(value).strftime("%d/%m/%Y %H:%M")
    except (ValueError, TypeError):
        return value


@app.template_filter("cpfmask")
def cpfmask(value):
    return f"***.***.***-{value[-2:]}" if value else "—"


def month_bounds(month=None):
    month = month or date.today().strftime("%Y-%m")
    try:
        start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError:
        start = date.today().replace(day=1)
        month = start.strftime("%Y-%m")
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)
    return month, start.isoformat(), end.isoformat()


def add_months(month, count):
    current = datetime.strptime(month, "%Y-%m").date()
    result = []
    for _ in range(count):
        result.append(current.strftime("%Y-%m"))
        current = current.replace(year=current.year + 1, month=1) if current.month == 12 else current.replace(month=current.month + 1)
    return result


@app.route("/")
@roles_allowed("manager", "staff")
def dashboard():
    conn = db()
    today = date.today().isoformat()
    month, start, end = month_bounds()
    metrics = conn.execute("""
        SELECT
          COALESCE(SUM(CASE WHEN date(created_at)=? AND payment_method!='Cortesia' THEN total_cents END),0) day_total,
          COALESCE(SUM(CASE WHEN created_at>=? AND created_at<? AND payment_method!='Cortesia' THEN total_cents END),0) month_total,
          COUNT(CASE WHEN created_at>=? AND created_at<? THEN 1 END) month_sales,
          COALESCE(SUM(CASE WHEN created_at>=? AND created_at<? AND payment_method='Débito' THEN total_cents END),0) debit_total
        FROM sales
    """, (today, start, end, start, end, start, end)).fetchone()
    low = conn.execute("SELECT * FROM products WHERE active=1 AND stock<=min_stock ORDER BY stock, name").fetchall()
    recent = conn.execute("""SELECT s.*, p.name player_name FROM sales s JOIN players p ON p.id=s.player_id
                            ORDER BY s.id DESC LIMIT 8""").fetchall()
    return render_template("dashboard.html", metrics=metrics, low=low, recent=recent, month=month)


@app.route("/products", methods=["GET", "POST"])
@roles_allowed("manager", "staff")
def products():
    conn = db()
    if request.method == "POST":
        try:
            units_per_case = int(request.form.get("units_per_case") or 0)
            loose_units = int(request.form.get("stock") or 0)
            cases = int(request.form.get("initial_cases") or 0)
            if min(units_per_case, loose_units, cases) < 0:
                raise ValueError("As quantidades não podem ser negativas.")
            if cases and not units_per_case:
                raise ValueError("Informe quantas unidades vêm em cada caixa.")
            initial_stock = loose_units + cases * units_per_case
            conn.execute("""INSERT INTO products(name,category,package_type,units_per_case,price_cents,cost_cents,stock,min_stock)
                            VALUES(?,?,?,?,?,?,?,?)""", (
                request.form["name"].strip(), request.form["category"], request.form.get("package_type", ""),
                units_per_case, cents(request.form["price"]), cents(request.form.get("cost", "0")),
                initial_stock, int(request.form.get("min_stock", 5))))
            conn.commit()
            flash("Produto cadastrado.", "success")
        except (sqlite3.IntegrityError, ValueError) as exc:
            flash(f"Não foi possível cadastrar: {exc}", "danger")
        return redirect(url_for("products"))
    items = conn.execute("SELECT * FROM products ORDER BY active DESC, category, name").fetchall()
    return render_template("products.html", products=items)


@app.post("/products/<int:product_id>/toggle")
@roles_allowed("manager", "staff")
def toggle_product(product_id):
    product = db().execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    if not product:
        flash("Produto não encontrado.", "warning")
    else:
        db().execute("UPDATE products SET active=1-active WHERE id=?", (product_id,))
        db().commit()
        flash("Produto excluído dos cadastros ativos; o histórico foi preservado." if product["active"]
              else "Produto restaurado.", "success")
    return redirect(url_for("products"))


@app.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
@roles_allowed("manager", "staff")
def edit_product(product_id):
    conn = db()
    product = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    if not product:
        flash("Produto não encontrado.", "warning")
        return redirect(url_for("products"))
    if request.method == "POST":
        try:
            units_per_case = int(request.form.get("units_per_case") or 0)
            min_stock = int(request.form.get("min_stock") or 0)
            if units_per_case < 0 or min_stock < 0:
                raise ValueError("As quantidades não podem ser negativas.")
            conn.execute("""UPDATE products SET name=?,category=?,package_type=?,units_per_case=?,
                price_cents=?,cost_cents=?,min_stock=? WHERE id=?""", (
                request.form["name"].strip(), request.form["category"], request.form.get("package_type", ""),
                units_per_case, cents(request.form["price"]), cents(request.form.get("cost", "0")),
                min_stock, product_id))
            conn.commit()
            flash("Produto atualizado.", "success")
            return redirect(url_for("products"))
        except (ValueError, sqlite3.IntegrityError) as exc:
            flash("Já existe outro produto com esse nome." if isinstance(exc, sqlite3.IntegrityError) else str(exc), "danger")
        product = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    return render_template("edit_product.html", product=product)


@app.route("/players", methods=["GET", "POST"])
@roles_allowed("manager")
def players():
    conn = db()
    if request.method == "POST":
        try:
            membership_type = request.form.get("membership_type", "regular")
            if membership_type not in ("regular", "goalkeeper", "board"):
                raise ValueError("Classificação financeira inválida.")
            conn.execute("""INSERT INTO players
                (name,war_name,cpf,phone,emergency_phone,email,membership_type) VALUES(?,?,?,?,?,?,?)""", (
                request.form["name"].strip(), request.form.get("war_name", "").strip(), normalize_cpf(request.form.get("cpf")),
                request.form.get("phone", "").strip(), request.form.get("emergency_phone", "").strip(),
                request.form.get("email", "").strip().lower(), membership_type))
            conn.commit()
            flash("Peladeiro cadastrado.", "success")
        except (sqlite3.IntegrityError, ValueError) as exc:
            flash("Já existe um peladeiro com esse nome ou CPF." if isinstance(exc, sqlite3.IntegrityError) else str(exc), "danger")
        return redirect(url_for("players"))
    player_filter = request.args.get("filter", "active")
    filters = {
        "active": ("active=1", ()),
        "regular": ("active=1 AND membership_type=?", ("regular",)),
        "board": ("active=1 AND membership_type=?", ("board",)),
        "goalkeeper": ("active=1 AND membership_type=?", ("goalkeeper",)),
        "inactive": ("active=0", ()),
        "all": ("1=1", ()),
    }
    if player_filter not in filters:
        player_filter = "active"
    where, params = filters[player_filter]
    items = conn.execute(f"SELECT * FROM players WHERE {where} ORDER BY active DESC,name", params).fetchall()
    return render_template("players.html", players=items, player_filter=player_filter)


@app.post("/players/<int:player_id>/membership-type")
@roles_allowed("manager")
def player_membership_type(player_id):
    membership_type = request.form.get("membership_type")
    if membership_type not in ("regular", "goalkeeper", "board"):
        flash("Classificação inválida.", "danger")
    else:
        db().execute("UPDATE players SET membership_type=? WHERE id=?", (membership_type, player_id))
        db().commit()
        flash("Classificação financeira atualizada.", "success")
    return redirect(url_for("players"))


@app.route("/players/<int:player_id>/edit", methods=["GET", "POST"])
@roles_allowed("manager")
def edit_player(player_id):
    conn = db()
    player = conn.execute("SELECT * FROM players WHERE id=?", (player_id,)).fetchone()
    if not player:
        flash("Peladeiro não encontrado.", "warning")
        return redirect(url_for("players"))
    if request.method == "POST":
        membership_type = request.form.get("membership_type", "regular")
        if membership_type not in ("regular", "goalkeeper", "board"):
            flash("Classificação financeira inválida.", "danger")
        else:
            try:
                conn.execute("""UPDATE players SET name=?,war_name=?,cpf=?,email=?,phone=?,emergency_phone=?,membership_type=?
                                WHERE id=?""", (request.form["name"].strip(), request.form.get("war_name", "").strip(),
                    normalize_cpf(request.form.get("cpf")),
                    request.form.get("email", "").strip().lower(), request.form.get("phone", "").strip(),
                    request.form.get("emergency_phone", "").strip(), membership_type, player_id))
                conn.commit()
                flash("Cadastro do peladeiro atualizado.", "success")
                return redirect(url_for("players"))
            except sqlite3.IntegrityError:
                flash("Já existe outro peladeiro com esse nome ou CPF.", "danger")
            except ValueError as exc:
                flash(str(exc), "danger")
        player = conn.execute("SELECT * FROM players WHERE id=?", (player_id,)).fetchone()
    return render_template("edit_player.html", player=player)


@app.post("/players/<int:player_id>/toggle-active")
@roles_allowed("manager")
def toggle_player_active(player_id):
    player = db().execute("SELECT * FROM players WHERE id=?", (player_id,)).fetchone()
    if not player:
        flash("Peladeiro não encontrado.", "warning")
    else:
        new_status = 0 if player["active"] else 1
        db().execute("UPDATE players SET active=? WHERE id=?", (new_status, player_id))
        db().commit()
        flash("Peladeiro excluído dos cadastros ativos; o histórico foi preservado." if not new_status
              else "Peladeiro restaurado.", "success")
    return redirect(url_for("players"))


@app.post("/players/import")
@roles_allowed("manager")
def import_players():
    upload = request.files.get("spreadsheet")
    if not upload or not upload.filename:
        flash("Escolha uma planilha para importar.", "danger")
        return redirect(url_for("players"))
    try:
        rows = spreadsheet_rows(upload)
        if not rows:
            raise ValueError("A planilha está vazia.")
        headers = {normalized_header(value): index for index, value in enumerate(rows[0])}
        name_index = next((headers[key] for key in ("nome", "name", "peladeiro") if key in headers), None)
        email_index = next((headers[key] for key in ("email", "emailaddress") if key in headers), None)
        if name_index is None or email_index is None:
            raise ValueError("A primeira linha precisa ter as colunas Nome e E-mail.")
        imported = updated = skipped = 0
        conn = db()
        with conn:
            for row in rows[1:]:
                name = str(row[name_index] or "").strip() if name_index < len(row) else ""
                email = str(row[email_index] or "").strip().lower() if email_index < len(row) else ""
                if not name or not email or "@" not in email:
                    skipped += 1
                    continue
                existing = conn.execute(
                    "SELECT * FROM players WHERE lower(name)=lower(?) OR lower(email)=lower(?) LIMIT 1",
                    (name, email),
                ).fetchone()
                if existing:
                    if not existing["email"] and existing["name"].lower() == name.lower():
                        conn.execute("UPDATE players SET email=? WHERE id=?", (email, existing["id"]))
                        updated += 1
                    else:
                        skipped += 1
                    continue
                conn.execute("INSERT INTO players(name,email) VALUES(?,?)", (name, email))
                imported += 1
        flash(f"Importação concluída: {imported} novos, {updated} atualizados e {skipped} ignorados.", "success")
    except (ValueError, csv.Error, sqlite3.Error) as exc:
        flash(f"Não foi possível importar: {exc}", "danger")
    return redirect(url_for("players"))


@app.route("/sale", methods=["GET", "POST"])
@roles_allowed("manager", "staff", "client")
def sale():
    conn = db()
    if request.method == "POST":
        product_ids = request.form.getlist("product_id")
        quantities = request.form.getlist("quantity")
        requested = {}
        try:
            player_id = int(request.form["player_id"])
            for raw_id, raw_qty in zip(product_ids, quantities):
                qty = int(raw_qty or 0)
                if qty > 0:
                    requested[int(raw_id)] = requested.get(int(raw_id), 0) + qty
            if not requested:
                raise ValueError("Escolha ao menos um produto.")
            placeholders = ",".join("?" for _ in requested)
            products_by_id = {r["id"]: r for r in conn.execute(f"SELECT * FROM products WHERE active=1 AND id IN ({placeholders})", tuple(requested))}
            if len(products_by_id) != len(requested):
                raise ValueError("Produto inválido ou inativo.")
            for pid, qty in requested.items():
                if products_by_id[pid]["stock"] < qty:
                    raise ValueError(f"Estoque insuficiente de {products_by_id[pid]['name']}.")
            total = sum(products_by_id[pid]["price_cents"] * qty for pid, qty in requested.items())
            method = request.form["payment_method"]
            if g.user["role"] == "client" and method not in ("Pix", "Dinheiro"):
                raise ValueError("Clientes podem registrar pagamentos somente em Pix ou Dinheiro.")
            paid = 1
            with conn:
                cur = conn.execute("INSERT INTO sales(player_id,payment_method,total_cents,paid,notes) VALUES(?,?,?,?,?)",
                                   (player_id, method, total, paid, request.form.get("notes", "").strip()))
                for pid, qty in requested.items():
                    product = products_by_id[pid]
                    conn.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                                 (cur.lastrowid, pid, qty, product["price_cents"], product["cost_cents"]))
                    updated = conn.execute("UPDATE products SET stock=stock-? WHERE id=? AND stock>=?", (qty, pid, qty))
                    if updated.rowcount != 1:
                        raise ValueError("O estoque mudou durante a venda. Tente novamente.")
            flash(f"Venda #{cur.lastrowid} registrada: {money(total)}.", "success")
            return redirect(url_for("sale"))
        except (ValueError, sqlite3.IntegrityError) as exc:
            flash(str(exc), "danger")
    player_rows = conn.execute("SELECT * FROM players WHERE active=1 ORDER BY name").fetchall()
    product_rows = conn.execute("SELECT * FROM products WHERE active=1 AND stock>0 ORDER BY category,name").fetchall()
    return render_template("sale.html", players=player_rows, products=product_rows)


@app.route("/stock", methods=["GET", "POST"])
@roles_allowed("manager", "staff")
def stock():
    conn = db()
    if request.method == "POST":
        try:
            pid = int(request.form["product_id"])
            product = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
            if not product:
                raise ValueError("Produto inválido.")
            loose_units = int(request.form.get("quantity") or 0)
            cases = int(request.form.get("cases") or 0)
            if min(loose_units, cases) < 0:
                raise ValueError("As quantidades não podem ser negativas.")
            if cases and not product["units_per_case"]:
                raise ValueError("Este produto não possui unidades por caixa cadastradas.")
            qty = loose_units + cases * product["units_per_case"]
            if qty <= 0:
                raise ValueError("Informe unidades avulsas ou quantidade de caixas.")
            cost = cents(request.form.get("unit_cost", "0"))
            with conn:
                conn.execute("INSERT INTO restocks(product_id,quantity,unit_cost_cents,notes) VALUES(?,?,?,?)",
                             (pid, qty, cost, (f"{cases} caixa(s). " if cases else "") + request.form.get("notes", "").strip()))
                conn.execute("UPDATE products SET stock=stock+?, cost_cents=CASE WHEN ?>0 THEN ? ELSE cost_cents END WHERE id=?",
                             (qty, cost, cost, pid))
            flash("Reposição registrada e estoque atualizado.", "success")
        except (ValueError, sqlite3.IntegrityError) as exc:
            flash(str(exc), "danger")
        return redirect(url_for("stock"))
    product_rows = conn.execute("SELECT * FROM products WHERE active=1 ORDER BY stock,name").fetchall()
    history = conn.execute("""SELECT r.*,p.name product_name FROM restocks r JOIN products p ON p.id=r.product_id
                              ORDER BY r.id DESC LIMIT 30""").fetchall()
    return render_template("stock.html", products=product_rows, history=history)


@app.route("/reports")
@roles_allowed("manager")
def reports():
    conn = db()
    month, start, end = month_bounds(request.args.get("month"))
    summary = conn.execute("""SELECT COALESCE(SUM(CASE WHEN payment_method!='Cortesia' THEN total_cents END),0) revenue,
        COUNT(*) sales_count, COALESCE(SUM(CASE WHEN payment_method='Pix' THEN total_cents END),0) pix,
        COALESCE(SUM(CASE WHEN payment_method='Dinheiro' THEN total_cents END),0) cash,
        COALESCE(SUM(CASE WHEN payment_method='Débito' THEN total_cents END),0) debit,
        COALESCE(SUM(CASE WHEN payment_method='Cortesia' THEN total_cents END),0) courtesy
        FROM sales WHERE created_at>=? AND created_at<?""", (start, end)).fetchone()
    by_product = conn.execute("""SELECT p.name, SUM(i.quantity) quantity,
        SUM(i.quantity*i.unit_price_cents) total, SUM(i.quantity*(i.unit_price_cents-i.unit_cost_cents)) profit
        FROM sale_items i JOIN sales s ON s.id=i.sale_id JOIN products p ON p.id=i.product_id
        WHERE s.created_at>=? AND s.created_at<? GROUP BY p.id ORDER BY quantity DESC""", (start, end)).fetchall()
    by_player = conn.execute("""SELECT p.name, COUNT(s.id) purchases, SUM(s.total_cents) total
        FROM sales s JOIN players p ON p.id=s.player_id WHERE s.created_at>=? AND s.created_at<?
        GROUP BY p.id ORDER BY total DESC""", (start, end)).fetchall()
    sales_rows = conn.execute("""SELECT s.*,p.name player_name FROM sales s JOIN players p ON p.id=s.player_id
        WHERE s.created_at>=? AND s.created_at<? ORDER BY s.id DESC""", (start, end)).fetchall()
    profit = sum(r["profit"] for r in by_product)
    report_year, due_month = int(month[:4]), int(month[5:7])
    contributors = conn.execute("SELECT id FROM players WHERE active=1 AND membership_type='regular'").fetchall()
    debts = []
    for player in contributors:
        paid = conn.execute("""SELECT COUNT(*) FROM membership_months
            WHERE player_id=? AND month>=? AND month<=?""",
            (player["id"], f"{report_year}-01", f"{report_year}-{due_month:02d}")).fetchone()[0]
        debts.append(max(0, due_month - paid))
    membership = {
        "up_to_date": sum(debt == 0 for debt in debts),
        "owing": sum(debt > 0 for debt in debts),
        "over_2": sum(debt > 2 for debt in debts),
        "over_4": sum(debt > 4 for debt in debts),
        "over_6": sum(debt > 6 for debt in debts),
        "active": conn.execute("SELECT COUNT(*) FROM players WHERE active=1").fetchone()[0],
        "inactive": conn.execute("SELECT COUNT(*) FROM players WHERE active=0").fetchone()[0],
        "exempt": conn.execute("SELECT COUNT(*) FROM players WHERE active=1 AND membership_type IN ('goalkeeper','board')").fetchone()[0],
    }
    return render_template("reports.html", month=month, summary=summary, by_product=by_product,
                           by_player=by_player, sales=sales_rows, profit=profit, membership=membership)


@app.route("/finance", methods=["GET", "POST"])
@roles_allowed("manager")
def finance():
    conn = db()
    monthly_fee = 1500
    if request.method == "POST":
        try:
            player_id = int(request.form["player_id"])
            eligible = conn.execute("SELECT 1 FROM players WHERE id=? AND active=1 AND membership_type='regular'", (player_id,)).fetchone()
            if not eligible:
                raise ValueError("Este peladeiro é isento ou está inativo.")
            start_month = request.form["start_month"]
            months_count = int(request.form["months_count"])
            covered_months = add_months(start_month, months_count)
            amount = monthly_fee * months_count
            with conn:
                cur = conn.execute("""INSERT INTO membership_payments
                    (player_id,amount_cents,months_count,start_month,payment_method,notes)
                    VALUES(?,?,?,?,?,?)""", (player_id, amount, months_count, start_month,
                    request.form["payment_method"], request.form.get("notes", "").strip()))
                for covered_month in covered_months:
                    conn.execute("INSERT INTO membership_months(payment_id,player_id,month) VALUES(?,?,?)",
                                 (cur.lastrowid, player_id, covered_month))
            flash(f"Mensalidade registrada: {months_count} mês(es), total de {money(amount)}.", "success")
        except (ValueError, sqlite3.IntegrityError) as exc:
            message = "Um ou mais meses selecionados já foram pagos por este peladeiro." if isinstance(exc, sqlite3.IntegrityError) else str(exc)
            flash(f"Não foi possível registrar: {message}", "danger")
        return redirect(url_for("finance", year=request.args.get("year", date.today().year)))
    try:
        year = int(request.args.get("year", date.today().year))
    except ValueError:
        year = date.today().year
    players_rows = conn.execute("SELECT * FROM players WHERE active=1 AND membership_type='regular' ORDER BY name").fetchall()
    exempt_count = conn.execute("SELECT COUNT(*) FROM players WHERE active=1 AND membership_type IN ('goalkeeper','board')").fetchone()[0]
    paid_rows = conn.execute("SELECT player_id,month FROM membership_months WHERE month LIKE ?", (f"{year}-%",)).fetchall()
    paid_by_player = {}
    for row in paid_rows:
        paid_by_player.setdefault(row["player_id"], set()).add(int(row["month"][-2:]))
    all_status_rows = [{"player": player, "months": paid_by_player.get(player["id"], set())} for player in players_rows]
    try:
        members_page = max(1, int(request.args.get("members_page", 1)))
    except ValueError:
        members_page = 1
    members_per_page = 10
    members_pages = max(1, (len(all_status_rows) + members_per_page - 1) // members_per_page)
    members_page = min(members_page, members_pages)
    status_rows = all_status_rows[(members_page - 1) * members_per_page:members_page * members_per_page]
    try:
        history_page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        history_page = 1
    per_page = 10
    history_total = conn.execute("SELECT COUNT(*) FROM membership_payments").fetchone()[0]
    history_pages = max(1, (history_total + per_page - 1) // per_page)
    history_page = min(history_page, history_pages)
    history = conn.execute("""SELECT mp.*,p.name player_name FROM membership_payments mp
        JOIN players p ON p.id=mp.player_id ORDER BY mp.id DESC LIMIT ? OFFSET ?""",
        (per_page, (history_page - 1) * per_page)).fetchall()
    collected = conn.execute("SELECT COALESCE(SUM(amount_cents),0) FROM membership_payments WHERE created_at>=? AND created_at<?",
                             (f"{year}-01-01", f"{year + 1}-01-01")).fetchone()[0]
    due_month = 12 if year < date.today().year else (date.today().month if year == date.today().year else 0)
    expected_to_date = len(players_rows) * due_month * monthly_fee
    covered_to_date = sum(sum(1 for month in row["months"] if month <= due_month) for row in all_status_rows) * monthly_fee
    return render_template("finance.html", players=players_rows, statuses=status_rows, history=history,
                           year=year, monthly_fee=monthly_fee, collected=collected,
                           expected=expected_to_date, outstanding=max(0, expected_to_date-covered_to_date),
                           current_month=date.today().strftime("%Y-%m"), history_page=history_page,
                           history_pages=history_pages, history_total=history_total,
                           members_page=members_page, members_pages=members_pages,
                           members_total=len(all_status_rows), exempt_count=exempt_count)


@app.post("/finance/<int:payment_id>/delete")
@roles_allowed("manager")
def delete_membership_payment(payment_id):
    conn = db()
    with conn:
        deleted = conn.execute("DELETE FROM membership_payments WHERE id=?", (payment_id,))
    flash("Recebimento apagado." if deleted.rowcount else "Recebimento não encontrado.",
          "success" if deleted.rowcount else "warning")
    return redirect(request.referrer or url_for("finance"))


@app.post("/sales/<int:sale_id>/delete")
@roles_allowed("manager", "staff")
def delete_sale(sale_id):
    conn = db()
    sale_row = conn.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
    if not sale_row:
        flash("Venda não encontrada ou já apagada.", "warning")
        return redirect(request.referrer or url_for("reports"))
    items = conn.execute(
        "SELECT product_id, quantity FROM sale_items WHERE sale_id=?", (sale_id,)
    ).fetchall()
    with conn:
        for item in items:
            conn.execute(
                "UPDATE products SET stock=stock+? WHERE id=?",
                (item["quantity"], item["product_id"]),
            )
        conn.execute("DELETE FROM sales WHERE id=?", (sale_id,))
    flash(f"Venda #{sale_id} apagada e itens devolvidos ao estoque.", "success")
    return redirect(request.referrer or url_for("reports"))


@app.route("/pix")
@roles_allowed("manager", "staff")
def pix():
    conn = db()
    day = request.args.get("day", date.today().isoformat())
    rows = conn.execute("""SELECT s.*,p.name player_name FROM sales s JOIN players p ON p.id=s.player_id
        WHERE date(s.created_at)=? AND s.payment_method='Pix' ORDER BY s.id DESC""", (day,)).fetchall()
    total = sum(r["total_cents"] for r in rows)
    return render_template("pix.html", rows=rows, total=total, day=day)


@app.get("/pix/qrcode")
@roles_allowed("manager", "staff", "client")
def pix_qrcode():
    try:
        amount_cents = int(request.args.get("amount_cents", 0))
        if amount_cents <= 0 or amount_cents > 100_000_000:
            raise ValueError
    except ValueError:
        return jsonify(error="Selecione produtos para gerar um Pix com valor válido."), 400
    import qrcode
    payload = pix_payload(amount_cents)
    qr = qrcode.QRCode(version=None, box_size=8, border=3)
    qr.add_data(payload)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    output = io.BytesIO()
    image.save(output, format="PNG")
    encoded = base64.b64encode(output.getvalue()).decode("ascii")
    return jsonify(
        payload=payload,
        image=f"data:image/png;base64,{encoded}",
        key=app.config["PIX_KEY"],
        amount=money(amount_cents),
    )


init_db()

if __name__ == "__main__":
    app.run(debug=True)
