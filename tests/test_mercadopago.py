import base64
import hashlib
import hmac
import json
import os
import re
import sqlite3
import sys
import tempfile
import time
import unittest
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from types import ModuleType
from unittest.mock import patch

from PIL import Image
from openpyxl import load_workbook

# Tests must never inherit the developer's production Supabase connection from
# .env.local.  Every test configures an isolated SQLite database in setUp.
os.environ["DATABASE_URL"] = ""
os.environ["SUPABASE_DB_URL"] = ""

from app import app
from flask import has_request_context
from src.db import (
    DbWrapper,
    connect_db,
    get_db,
    init_postgres,
    init_sqlite,
    initialize_sqlite_database,
    read_user_from_session,
    run_postgres_migrations,
)
from src.environment import environment_config
from src.routes.auth import make_password_hash
from src.routes.football import _sumula
from src.routes.sales import pix_access_token
from src.services.mercadopago import validate_webhook_signature
from src.services.mercadopago import MercadoPagoError
from src.services.mercadopago import create_pix_order
from src.services.email_reminders import dispatch_reminders, get_reminder_settings, outstanding_players, send_gmail
from src.services.cash_register import get_session, session_summary
from src.services.monthly_sales_report import monthly_sales_data
from src.services.stock_alerts import notify_low_stock
from src.services.push_notifications import send_player_push
from src.services.material_photos import MAX_UPLOAD_BYTES, process_material_photo
from src.utils import alphabetical_key, brdate, local_today, month_bounds
from werkzeug.datastructures import FileStorage
from werkzeug.security import check_password_hash


class MercadoPagoFlowTest(unittest.TestCase):
    SCHEMA_SQL = re.compile(
        r"^\s*(?:CREATE\s+(?:OR\s+REPLACE\s+)?(?:TABLE|FUNCTION|INDEX)|ALTER\s+TABLE|"
        r"DROP\s+(?:TABLE|INDEX|CONSTRAINT)|DO\s+\$\$)",
        re.IGNORECASE,
    )

    def test_postgres_schema_omits_sqlite_seed_syntax(self):
        class Recorder:
            def __init__(self):
                self.statements = []

            def execute(self, statement, params=()):
                self.statements.append(statement)
                return self

            def commit(self):
                return None

        recorder = Recorder()
        init_postgres(recorder)
        self.assertFalse(
            any("INSERT OR IGNORE" in statement.upper() for statement in recorder.statements)
        )
        schedule_seeds = [
            statement
            for statement in recorder.statements
            if statement.upper().startswith("INSERT INTO TRIBUTE_SCHEDULES")
        ]
        self.assertTrue(schedule_seeds)
        self.assertTrue(all("RETURNING WEEKDAY" in statement.upper() for statement in schedule_seeds))
        self.assertTrue(any("ADD COLUMN IF NOT EXISTS CLUB_QR_DATA" in statement.upper() for statement in recorder.statements))
        self.assertTrue(any("IDX_PLAYERS_CLUB_QR_TOKEN" in statement.upper() for statement in recorder.statements))
        self.assertTrue(any(
            "LOAD_ENTRIES_AREA_CODE_CHECK" in statement.upper() and "'INT'" in statement.upper()
            for statement in recorder.statements
        ))

    def test_postgres_connect_does_not_run_schema_setup(self):
        """Opening a production connection may set the timezone, but never run DDL."""
        statements = []

        class Cursor:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return None

            def execute(self, statement, _params=()):
                statements.append(statement)

        class Connection:
            def cursor(self):
                return Cursor()

        config = type("AppConfig", (), {"config": {"DATABASE_URL": "postgresql://example.invalid/db"}})()
        with patch.dict(os.environ, {"DATABASE_URL": ""}), \
             patch("psycopg2.connect", return_value=Connection()), \
             patch("src.db.init_postgres") as schema_setup:
            connection = connect_db(config)

        self.assertIsInstance(connection, DbWrapper)
        schema_setup.assert_not_called()
        self.assertEqual(statements, ["SET TIME ZONE 'UTC'"])
        self.assertFalse(any(self.SCHEMA_SQL.match(statement) for statement in statements))

    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        app.config.update(
            TESTING=True,
            DATABASE=str(Path(self.tempdir.name) / "test.db"),
            DATABASE_URL=None,
            WTF_CSRF_ENABLED=False,
            SECRET_KEY="test-secret",
            MERCADOPAGO_ACCESS_TOKEN="APP_USR-test",
            MERCADOPAGO_POS_ID="CAIXA_TESTE",
            MERCADOPAGO_WEBHOOK_SECRET="webhook-secret",
            GMAIL_SMTP_USER="diretoriagpcta@gmail.com",
            GMAIL_APP_PASSWORD="app-password-test",
            CRON_SECRET="cron-secret-test",
        )
        initialize_sqlite_database(app.config["DATABASE"])
        self.create_sports_schema()
        with app.app_context():
            db = get_db()
            db.execute("INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'manager')", ("teste", "Teste", "hash"))
            db.execute("INSERT INTO players(name,email) VALUES(?,?)", ("Peladeiro", "peladeiro@example.com"))
            db.execute(
                "INSERT INTO products(name,category,price_cents,cost_cents,stock) VALUES(?,?,?,?,?)",
                ("Água", "Bebida", 300, 100, 5),
            )
            db.commit()
            user = db.execute("SELECT * FROM users WHERE username='teste'").fetchone()
            self.user_id = user["id"]
            self.token = pix_access_token(user)
            self.player_id = db.execute("SELECT id FROM players WHERE name='Peladeiro'").fetchone()["id"]
            self.product_id = db.execute("SELECT id FROM products WHERE name='Água'").fetchone()["id"]
        self.client = app.test_client()

    def tearDown(self):
        # Windows can briefly retain a just-closed SQLite file handle. Force a
        # garbage collection to run so any pending sqlite finalizers close
        # file handles before attempting to remove temporary directories.
        import gc
        gc.collect()
        # Only retry fixture cleanup; application connections are still closed
        # by Flask's teardown on every request/app context.
        for attempt in range(3):
            try:
                self.tempdir.cleanup()
                break
            except OSError:
                if attempt == 2:
                    raise
                time.sleep(0.05 * (attempt + 1))

    def headers(self):
        return {"Accept": "application/json", "X-Pix-Token": self.token}

    def create_sports_schema(self):
        with app.app_context():
            db = get_db()
            db.conn.executescript("""
                CREATE TABLE sports_material_types (
                    id INTEGER PRIMARY KEY, code TEXT NOT NULL UNIQUE, name TEXT NOT NULL,
                    active INTEGER NOT NULL DEFAULT 1, sort_order INTEGER NOT NULL DEFAULT 0
                );
                CREATE TABLE sports_product_config (
                    product_id INTEGER PRIMARY KEY, type_id INTEGER NOT NULL,
                    allow_custom_name INTEGER NOT NULL DEFAULT 0,
                    allow_custom_number INTEGER NOT NULL DEFAULT 0,
                    allow_backorder INTEGER NOT NULL DEFAULT 0,
                    ready_sale_enabled INTEGER NOT NULL DEFAULT 1
                );
                CREATE TABLE sports_product_variants (
                    id INTEGER PRIMARY KEY, product_id INTEGER NOT NULL, size TEXT NOT NULL,
                    stock INTEGER NOT NULL DEFAULT 0, min_stock INTEGER NOT NULL DEFAULT 0,
                    active INTEGER NOT NULL DEFAULT 1, updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE sports_sale_item_details (
                    sale_item_id INTEGER PRIMARY KEY, variant_id INTEGER NOT NULL,
                    variant_size TEXT NOT NULL, custom_name TEXT NOT NULL DEFAULT '',
                    custom_number TEXT NOT NULL DEFAULT '', order_mode TEXT NOT NULL,
                    fulfillment_status TEXT NOT NULL, delivered_at TEXT, delivered_by INTEGER,
                    canceled_at TEXT, canceled_by INTEGER, cancellation_reason TEXT DEFAULT '',
                    cancellation_resolution TEXT NOT NULL DEFAULT 'none',
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE sports_stock_reservations (
                    id INTEGER PRIMARY KEY, sale_item_id INTEGER NOT NULL UNIQUE,
                    variant_id INTEGER NOT NULL, quantity INTEGER NOT NULL,
                    status TEXT NOT NULL DEFAULT 'reserved', expires_at TEXT,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE sports_order_status_history (
                    id INTEGER PRIMARY KEY, sale_item_id INTEGER NOT NULL,
                    from_status TEXT NOT NULL, to_status TEXT NOT NULL,
                    changed_by INTEGER, notes TEXT NOT NULL DEFAULT '',
                    changed_at TEXT DEFAULT CURRENT_TIMESTAMP
                );
            """)
            db.execute("INSERT INTO sports_material_types(id,code,name) VALUES(1,'shirt','Camisa avulsa')")
            db.execute("""INSERT INTO sports_material_types(id,code,name)
                          VALUES(2,'commemorative_coin','Moeda comemorativa')""")
            db.commit()

    def create_sports_product(self, name, type_id=1, size="M", stock=5,
                              active=1, allow_name=0, allow_number=0, allow_backorder=0):
        with app.app_context():
            db = get_db()
            product = db.execute(
                """INSERT INTO products(name,category,price_cents,cost_cents,stock,active)
                   VALUES(?,'Material Esportivo',2000,1000,0,?)""",
                (name, active),
            )
            db.execute(
                """INSERT INTO sports_product_config
                   (product_id,type_id,allow_custom_name,allow_custom_number,allow_backorder)
                   VALUES(?,?,?,?,?)""",
                (product.lastrowid, type_id, allow_name, allow_number, allow_backorder),
            )
            variant = db.execute(
                """INSERT INTO sports_product_variants(product_id,size,stock,min_stock,active)
                   VALUES(?,?,?,1,1)""",
                (product.lastrowid, size, stock),
            )
            db.commit()
            return product.lastrowid, variant.lastrowid

    def login_manager(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

    def sports_sale_form(self, product_id, variant_id, **overrides):
        form = {
            "department": "sports", "sale_type": "player",
            "player_id": str(self.player_id), "payment_method": "Dinheiro",
            "product_id": [str(product_id)], "variant_id": [str(variant_id)],
            "quantity": ["1"], "custom_name": [""], "custom_number": [""],
            "order_mode": ["ready"],
        }
        form.update(overrides)
        return form

    def test_sports_catalog_restores_sized_coin_and_bar_products(self):
        shirt_id, _ = self.create_sports_product("Camisa Teste", size="M", stock=5)
        coin_id, _ = self.create_sports_product("Moeda Teste", type_id=2, size="Único", stock=5)
        self.login_manager()

        response = self.client.get("/sale?catalog=sports")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Camisa Teste", html)
        self.assertIn("Moeda Teste", html)
        self.assertIn('"single_variant": true', html)
        self.assertIn('"size": "\\u00danico"', html)
        self.assertIn('data-group="Material Esportivo"', html)
        self.assertIn("Água", html)
        self.assertIn('data-group="Bebidas"', html)
        self.assertNotEqual(shirt_id, coin_id)

    def test_sports_ready_sale_debits_only_variant_stock(self):
        product_id, variant_id = self.create_sports_product("Camisa Estoque", stock=5)
        self.login_manager()

        response = self.client.post("/sale", data=self.sports_sale_form(product_id, variant_id))
        self.assertEqual(response.status_code, 303)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)).fetchone()["stock"], 4)
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (product_id,)).fetchone()["stock"], 0)
            detail = db.execute("SELECT variant_id,order_mode,fulfillment_status FROM sports_sale_item_details").fetchone()
            self.assertEqual((detail["variant_id"], detail["order_mode"], detail["fulfillment_status"]),
                             (variant_id, "ready", "reserved"))

    def test_sports_rejects_inactive_variant_and_zero_ready_stock(self):
        inactive_product, inactive_variant = self.create_sports_product("Camisa Inativa", stock=5)
        zero_product, zero_variant = self.create_sports_product("Camisa Zerada", stock=0)
        with app.app_context():
            db = get_db()
            db.execute("UPDATE sports_product_variants SET active=0 WHERE id=?", (inactive_variant,))
            db.commit()
        self.login_manager()

        self.client.post("/sale", data=self.sports_sale_form(inactive_product, inactive_variant))
        self.client.post("/sale", data=self.sports_sale_form(zero_product, zero_variant))
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM sports_sale_item_details").fetchone()["total"], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM sales").fetchone()["total"], 0)

    def test_sports_backorder_and_personalization_follow_configuration(self):
        blocked_id, blocked_variant = self.create_sports_product("Sem Encomenda", stock=0)
        allowed_id, allowed_variant = self.create_sports_product(
            "Com Encomenda", stock=0, allow_name=1, allow_number=1, allow_backorder=1)
        self.login_manager()

        self.client.post("/sale", data=self.sports_sale_form(
            blocked_id, blocked_variant, order_mode=["backorder"], custom_name=["Nome"]));
        self.client.post("/sale", data=self.sports_sale_form(
            allowed_id, allowed_variant, order_mode=["backorder"],
            custom_name=["  Nome   Teste  "], custom_number=["10"]));
        with app.app_context():
            db = get_db()
            details = db.execute("SELECT * FROM sports_sale_item_details").fetchall()
            self.assertEqual(len(details), 1)
            self.assertEqual(details[0]["custom_name"], "Nome Teste")
            self.assertEqual(details[0]["custom_number"], "10")
            self.assertEqual(details[0]["fulfillment_status"], "requested")
            self.assertEqual(db.execute("SELECT stock FROM sports_product_variants WHERE id=?", (allowed_variant,)).fetchone()["stock"], 0)

    def test_sports_pix_ready_uses_variant_stock_and_is_idempotent(self):
        from src.routes.sales import apply_mercadopago_status

        product_id, variant_id = self.create_sports_product("Camisa Pix", stock=5)
        order = {
            "id": "ORD-SPORTS-READY", "status": "action_required",
            "transactions": {"payments": [{
                "id": "PAY-SPORTS-READY",
                "payment_method": {"qr_code": "000201SPORTS"},
            }]},
        }
        with patch("src.routes.sales.create_pix_order", return_value=order) as create_order:
            response = self.client.post(
                "/pix/mercadopago/orders", headers=self.headers(),
                json={"department": "sports", "player_id": self.player_id,
                      "use_bar_credit": True,
                      "items": [{"department": "sports", "product_id": product_id,
                                 "variant_id": variant_id, "quantity": 1,
                                 "custom_name": "", "custom_number": "",
                                 "order_mode": "ready"}]},
            )
        self.assertEqual(response.status_code, 201, response.get_json())
        sale_id = response.get_json()["sale_id"]
        self.assertEqual(create_order.call_args.args[2], 2000)
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            detail = db.execute(
                """SELECT d.variant_id,d.fulfillment_status,r.status reservation_status
                   FROM sports_sale_item_details d
                   JOIN sale_items si ON si.id=d.sale_item_id
                   JOIN sports_stock_reservations r ON r.sale_item_id=si.id
                   WHERE si.sale_id=?""", (sale_id,),
            ).fetchone()
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)
            ).fetchone()[0], 4)
            self.assertEqual(db.execute(
                "SELECT stock FROM products WHERE id=?", (product_id,)
            ).fetchone()[0], 0)
            self.assertEqual((detail["variant_id"], detail["fulfillment_status"],
                              detail["reservation_status"]), (variant_id, "reserved", "reserved"))
            approved = {"status": "processed", "status_detail": "accredited",
                        "total_paid_amount": "20.00",
                        "transactions": {"payments": [{"id": "PAY-SPORTS-APPROVED"}]}}
            self.assertEqual(apply_mercadopago_status(db, sale, approved), "approved")
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertEqual(apply_mercadopago_status(db, sale, approved), "approved")
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)
            ).fetchone()[0], 4)
            self.assertEqual(db.execute(
                "SELECT status FROM sports_stock_reservations r JOIN sale_items si ON si.id=r.sale_item_id WHERE si.sale_id=?",
                (sale_id,),
            ).fetchone()[0], "consumed")
            self.assertEqual((sale["paid"], sale["payment_status"]), (1, "approved"))
        self.login_manager()
        queue = self.client.get("/material-esportivo/vendas")
        self.assertIn("Camisa Pix", queue.get_data(as_text=True))

    def test_sports_pix_terminal_restores_ready_stock_and_backorder_never_debits(self):
        from src.routes.sales import apply_mercadopago_status

        ready_product, ready_variant = self.create_sports_product("Camisa Cancelada", stock=2)
        backorder_product, backorder_variant = self.create_sports_product(
            "Camisa Encomenda Pix", stock=0, allow_backorder=1)

        def create_sports_order(product_id, variant_id, mode, order_id):
            order = {"id": order_id, "transactions": {"payments": [{
                "id": f"PAY-{order_id}", "payment_method": {"qr_code": "000201SPORTS"},
            }]}}
            with patch("src.routes.sales.create_pix_order", return_value=order):
                response = self.client.post(
                    "/pix/mercadopago/orders", headers=self.headers(),
                    json={"department": "sports", "player_id": self.player_id,
                          "items": [{"department": "sports", "product_id": product_id,
                                     "variant_id": variant_id, "quantity": 1,
                                     "custom_name": "", "custom_number": "",
                                     "order_mode": mode}]},
                )
            self.assertEqual(response.status_code, 201, response.get_json())
            return response.get_json()["sale_id"]

        ready_sale = create_sports_order(ready_product, ready_variant, "ready", "ORD-SPORTS-CANCEL")
        self.login_manager()
        requested = self.client.post("/sale", data=self.sports_sale_form(
            backorder_product, backorder_variant, order_mode=["backorder"]
        ))
        self.assertEqual(requested.status_code, 303)
        with app.app_context():
            db = get_db()
            backorder_sale = db.execute(
                "SELECT MAX(id) id FROM sales WHERE payment_status='requested'"
            ).fetchone()["id"]
            ready = db.execute("SELECT * FROM sales WHERE id=?", (ready_sale,)).fetchone()
            self.assertEqual(apply_mercadopago_status(
                db, ready, {"status": "canceled", "transactions": {"payments": []}}
            ), "canceled")
            ready = db.execute("SELECT * FROM sales WHERE id=?", (ready_sale,)).fetchone()
            self.assertEqual(apply_mercadopago_status(
                db, ready, {"status": "canceled", "transactions": {"payments": []}}
            ), "canceled")
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (ready_variant,)
            ).fetchone()[0], 2)
            self.assertEqual(db.execute(
                "SELECT status FROM sports_stock_reservations r JOIN sale_items si ON si.id=r.sale_item_id WHERE si.sale_id=?",
                (ready_sale,),
            ).fetchone()[0], "released")
            detail = db.execute(
                """SELECT d.fulfillment_status FROM sports_sale_item_details d
                   JOIN sale_items si ON si.id=d.sale_item_id WHERE si.sale_id=?""",
                (backorder_sale,),
            ).fetchone()
            self.assertEqual(detail["fulfillment_status"], "requested")
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (backorder_variant,)
            ).fetchone()[0], 0)
            self.assertIsNone(db.execute(
                """SELECT r.id FROM sports_stock_reservations r
                   JOIN sale_items si ON si.id=r.sale_item_id WHERE si.sale_id=?""",
                (backorder_sale,),
            ).fetchone())

    def test_environment_config_wires_homologation_safeguards_and_banner(self):
        homologation = environment_config("homologation")
        self.assertEqual(homologation, {
            "APP_ENV": "homologation",
            "IS_HOMOLOGATION": True,
            "EXTERNAL_PAYMENTS_ENABLED": False,
            "CRON_ENABLED": False,
        })
        self.login_manager()
        with patch.dict(app.config, homologation):
            page = self.client.get("/sale")
            self.assertEqual(page.status_code, 200)
            page_text = page.get_data(as_text=True)
            self.assertIn("AMBIENTE DE HOMOLOGAÇÃO", page_text)
            self.assertIn("Pix indisponível na homologação", page_text)

    @patch("src.routes.sales.create_pix_order")
    def test_homologation_blocks_external_sales_pix(self, create_order_mock):
        with patch.dict(app.config, environment_config("homologation")):
            response = self.client.post(
                "/pix/mercadopago/orders",
                headers=self.headers(),
                json={"player_id": self.player_id,
                      "items": [{"product_id": self.product_id, "quantity": 1}]},
            )
        self.assertEqual(response.status_code, 403)
        create_order_mock.assert_not_called()

    def test_homologation_blocks_external_credit_topup(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE users SET role='client',player_id=? WHERE id=?", (self.player_id, self.user_id))
            db.commit()
        self.login_manager()
        with patch.dict(app.config, environment_config("homologation")):
            response = self.client.post("/creditos/comprar", json={"amount_cents": 1000})
        self.assertEqual(response.status_code, 403)

    def test_production_environment_enables_external_payments_and_hides_banner(self):
        production = environment_config("production")
        self.assertEqual(production, {
            "APP_ENV": "production",
            "IS_HOMOLOGATION": False,
            "EXTERNAL_PAYMENTS_ENABLED": True,
            "CRON_ENABLED": True,
        })
        self.login_manager()
        with patch.dict(app.config, production):
            page = self.client.get("/sale")
            self.assertEqual(page.status_code, 200)
            page_text = page.get_data(as_text=True)
            self.assertNotIn("AMBIENTE DE HOMOLOGAÇÃO", page_text)
            self.assertNotIn("Pix indisponível na homologação", page_text)
            self.assertIn("<option>Pix</option>", page_text)
            from src.routes.sales import mercadopago_enabled
            with app.test_request_context():
                self.assertTrue(mercadopago_enabled())

    def test_bar_sale_still_uses_products_stock(self):
        self.login_manager()
        response = self.client.post("/sale", data={
            "department": "bar", "sale_type": "player", "player_id": self.player_id,
            "payment_method": "Dinheiro", "product_id": [self.product_id], "quantity": ["1"],
        })
        self.assertEqual(response.status_code, 303)
        with app.app_context():
            self.assertEqual(get_db().execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 4)

    def test_sports_orders_menu_respects_manager_staff_and_client_roles(self):
        self.login_manager()
        manager_html = self.client.get("/sale").get_data(as_text=True)
        self.assertIn("Cadastro de Material Esportivo", manager_html)
        self.assertIn("Pedidos de Material Esportivo", manager_html)
        self.assertIn("Relatório de Material Esportivo", manager_html)

        with app.app_context():
            db = get_db()
            staff_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES('staff-sports','Staff','hash','staff')"
            ).lastrowid
            client_id = db.execute(
                """INSERT INTO users(username,name,password_hash,role,player_id)
                   VALUES('client-sports','Cliente','hash','client',?)""",
                (self.player_id,),
            ).lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = staff_id
        staff_html = self.client.get("/sale").get_data(as_text=True)
        self.assertIn("Pedidos de Material Esportivo", staff_html)
        self.assertIn("Cadastro de Material Esportivo", staff_html)
        self.assertNotIn("Relatório de Material Esportivo", staff_html)
        self.assertEqual(self.client.get("/material-esportivo/vendas").status_code, 200)

        with self.client.session_transaction() as session:
            session["user_id"] = client_id
        client_html = self.client.get("/sale").get_data(as_text=True)
        self.assertNotIn("Pedidos de Material Esportivo", client_html)
        self.assertNotIn("Cadastro de Material Esportivo", client_html)
        self.assertNotIn("Relatório de Material Esportivo", client_html)
        self.assertIn("Compra rápida", client_html)
        self.assertIn("Material Esportivo", client_html)
        self.assertNotEqual(self.client.get("/material-esportivo/vendas").status_code, 200)

    def test_sports_orders_staff_defaults_reserved_manager_sees_all_and_excludes_bar(self):
        reserved_product, reserved_variant = self.create_sports_product("Camisa Reservada", stock=2)
        requested_product, requested_variant = self.create_sports_product(
            "Camisa Solicitada", stock=0, allow_backorder=1)
        self.login_manager()
        self.client.post("/sale", data=self.sports_sale_form(reserved_product, reserved_variant))
        self.client.post("/sale", data=self.sports_sale_form(
            requested_product, requested_variant, order_mode=["backorder"]))
        self.client.post("/sale", data={
            "department": "bar", "sale_type": "player", "player_id": self.player_id,
            "payment_method": "Dinheiro", "product_id": [self.product_id], "quantity": ["1"],
        })

        manager_page = self.client.get("/material-esportivo/vendas")
        manager_html = manager_page.get_data(as_text=True)
        self.assertEqual(manager_page.status_code, 200)
        self.assertIn("Camisa Reservada", manager_html)
        self.assertIn("Camisa Solicitada", manager_html)
        self.assertNotIn("Água", manager_html)

        with app.app_context():
            db = get_db()
            staff_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES('staff-filter','Staff','hash','staff')"
            ).lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = staff_id
        staff_html = self.client.get("/material-esportivo/vendas").get_data(as_text=True)
        self.assertIn("Camisa Reservada", staff_html)
        self.assertNotIn("Camisa Solicitada", staff_html)
        self.assertIn('value="reserved" selected', staff_html)

    def test_sports_fulfillment_updates_only_selected_item_and_records_history(self):
        first_product, first_variant = self.create_sports_product("Camisa Um", stock=2)
        second_product, second_variant = self.create_sports_product("Camisa Dois", stock=2)
        self.login_manager()
        self.client.post("/sale", data=self.sports_sale_form(first_product, first_variant))
        self.client.post("/sale", data=self.sports_sale_form(second_product, second_variant))
        with app.app_context():
            db = get_db()
            details = db.execute(
                """SELECT d.sale_item_id,si.product_id,si.sale_id FROM sports_sale_item_details d
                   JOIN sale_items si ON si.id=d.sale_item_id ORDER BY d.sale_item_id"""
            ).fetchall()
            target_id = next(row["sale_item_id"] for row in details if row["product_id"] == first_product)
            target_sale_id = next(row["sale_id"] for row in details if row["product_id"] == first_product)
            other_id = next(row["sale_item_id"] for row in details if row["product_id"] == second_product)

        payment = self.client.post(
            f"/orders/{target_sale_id}/confirm-payment",
            json={"amount_received_cents": 2000},
        )
        self.assertEqual(payment.status_code, 200, payment.get_json())
        response = self.client.post(
            f"/material-esportivo/vendas/{target_id}/status",
            json={"to_status": "delivered"},
        )
        self.assertEqual(response.status_code, 200, response.get_json())
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT fulfillment_status FROM sports_sale_item_details WHERE sale_item_id=?", (target_id,)
            ).fetchone()["fulfillment_status"], "delivered")
            self.assertEqual(db.execute(
                "SELECT fulfillment_status FROM sports_sale_item_details WHERE sale_item_id=?", (other_id,)
            ).fetchone()["fulfillment_status"], "reserved")
            history = db.execute(
                "SELECT from_status,to_status FROM sports_order_status_history WHERE sale_item_id=?", (target_id,)
            ).fetchone()
            self.assertEqual((history["from_status"], history["to_status"]), ("reserved", "delivered"))

    def test_sports_backorder_supplier_pdf_receive_excess_and_notification_are_idempotent(self):
        product_id, variant_id = self.create_sports_product(
            "Camisa Fornecedor", stock=0, allow_name=1, allow_number=1, allow_backorder=1
        )
        self.login_manager()
        for index in range(4):
            response = self.client.post("/sale", data=self.sports_sale_form(
                product_id, variant_id, order_mode=["backorder"],
                custom_name=[f"Nome {index}"], custom_number=[str(index + 1)],
            ))
            self.assertEqual(response.status_code, 303)
        with app.app_context():
            db = get_db()
            item_ids = [row["sale_item_id"] for row in db.execute(
                "SELECT sale_item_id FROM sports_sale_item_details ORDER BY sale_item_id"
            ).fetchall()]

        pdf = self.client.get("/material-esportivo/vendas/fornecedor.pdf")
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf.mimetype, "application/pdf")
        with app.app_context():
            self.assertEqual(get_db().execute(
                "SELECT COUNT(*) total FROM sports_sale_item_details WHERE fulfillment_status='requested'"
            ).fetchone()["total"], 4)

        sent = self.client.post(
            "/material-esportivo/vendas/confirmar-envio", json={"sale_item_ids": item_ids}
        )
        self.assertEqual(sent.status_code, 200, sent.get_json())
        received = self.client.post("/material-esportivo/vendas/receber", json={
            "sale_item_ids": item_ids, "variant_id": variant_id, "received_quantity": 10,
        })
        self.assertEqual(received.status_code, 200, received.get_json())
        self.assertEqual(received.get_json()["stock_excess"], 6)
        duplicate = self.client.post("/material-esportivo/vendas/receber", json={
            "sale_item_ids": item_ids, "variant_id": variant_id, "received_quantity": 10,
        })
        self.assertEqual(duplicate.status_code, 409)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)
            ).fetchone()["stock"], 6)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM sports_sale_item_details WHERE fulfillment_status='available'"
            ).fetchone()["total"], 4)
            events = db.execute(
                "SELECT event_key,payload FROM notification_outbox WHERE event_type='sports_order_available_push'"
            ).fetchall()
            self.assertEqual(len(events), 4)
            self.assertTrue(all("Seu produto chegou" in row["payload"] for row in events))

    def test_sports_backorder_partial_receive_payment_gate_and_cancellation_stock_once(self):
        product_id, variant_id = self.create_sports_product(
            "Camisa Parcial", stock=0, allow_backorder=1
        )
        self.login_manager()
        for _ in range(3):
            self.client.post("/sale", data=self.sports_sale_form(
                product_id, variant_id, order_mode=["backorder"]
            ))
        with app.app_context():
            db = get_db()
            rows = db.execute(
                """SELECT d.sale_item_id,si.sale_id FROM sports_sale_item_details d
                   JOIN sale_items si ON si.id=d.sale_item_id ORDER BY d.sale_item_id"""
            ).fetchall()
            item_ids = [row["sale_item_id"] for row in rows]
        blocked = self.client.post(
            f"/material-esportivo/vendas/{item_ids[0]}/pagamento",
            json={"payment_method": "Dinheiro"},
        )
        self.assertEqual(blocked.status_code, 409)
        self.assertEqual(self.client.post(
            "/material-esportivo/vendas/confirmar-envio", json={"sale_item_ids": item_ids}
        ).status_code, 200)
        partial = self.client.post("/material-esportivo/vendas/receber", json={
            "sale_item_ids": item_ids[:2], "variant_id": variant_id, "received_quantity": 2,
        })
        self.assertEqual(partial.status_code, 200, partial.get_json())
        with app.app_context():
            db = get_db()
            statuses = [row["fulfillment_status"] for row in db.execute(
                "SELECT fulfillment_status FROM sports_sale_item_details ORDER BY sale_item_id"
            ).fetchall()]
            self.assertEqual(statuses, ["available", "available", "in_production"])
            before_history = db.execute(
                "SELECT COUNT(*) total FROM sports_order_status_history"
            ).fetchone()["total"]
            before_outbox = db.execute(
                "SELECT COUNT(*) total FROM notification_outbox"
            ).fetchone()["total"]
            before_stock = db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)
            ).fetchone()["stock"]
        pending_page = self.client.get("/material-esportivo/vendas?status=available").get_data(as_text=True)
        self.assertIn("Pagamento pendente", pending_page)
        self.assertNotIn(
            f'data-item="{item_ids[0]}" data-status="delivered"', pending_page
        )
        direct_delivery = self.client.post(
            f"/material-esportivo/vendas/{item_ids[0]}/status",
            json={"to_status": "delivered"},
        )
        self.assertEqual(direct_delivery.status_code, 409)
        self.assertIn("ainda não está pago", direct_delivery.get_json()["error"])
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT fulfillment_status FROM sports_sale_item_details WHERE sale_item_id=?",
                (item_ids[0],),
            ).fetchone()["fulfillment_status"], "available")
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM sports_order_status_history"
            ).fetchone()["total"], before_history)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM notification_outbox"
            ).fetchone()["total"], before_outbox)
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)
            ).fetchone()["stock"], before_stock)
        payment = self.client.post(
            f"/material-esportivo/vendas/{item_ids[0]}/pagamento",
            json={"payment_method": "Dinheiro"},
        )
        self.assertEqual(payment.status_code, 200, payment.get_json())
        confirmed = self.client.post(
            f"/orders/{rows[0]['sale_id']}/confirm-payment",
            json={"amount_received_cents": 2000},
        )
        self.assertEqual(confirmed.status_code, 200, confirmed.get_json())
        paid_page = self.client.get("/material-esportivo/vendas?status=available").get_data(as_text=True)
        self.assertIn(f'data-item="{item_ids[0]}" data-status="delivered"', paid_page)
        delivered = self.client.post(
            f"/material-esportivo/vendas/{item_ids[0]}/status",
            json={"to_status": "delivered"},
        )
        self.assertEqual(delivered.status_code, 200, delivered.get_json())
        duplicate_delivery = self.client.post(
            f"/material-esportivo/vendas/{item_ids[0]}/status",
            json={"to_status": "delivered"},
        )
        self.assertEqual(duplicate_delivery.status_code, 200, duplicate_delivery.get_json())
        self.assertTrue(duplicate_delivery.get_json()["already_delivered"])
        self.assertEqual(self.client.post(
            f"/material-esportivo/vendas/{item_ids[0]}/cancelar", json={"reason": "Teste"}
        ).status_code, 409)
        canceled = self.client.post(
            f"/material-esportivo/vendas/{item_ids[1]}/cancelar", json={"reason": "Desistência"}
        )
        self.assertEqual(canceled.status_code, 200, canceled.get_json())
        repeated = self.client.post(
            f"/material-esportivo/vendas/{item_ids[1]}/cancelar", json={"reason": "Desistência"}
        )
        self.assertEqual(repeated.status_code, 200, repeated.get_json())
        with app.app_context():
            self.assertEqual(get_db().execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variant_id,)
            ).fetchone()["stock"], 1)

    def test_sports_backorder_receives_multiple_dynamic_variants_atomically(self):
        product_id, p_variant = self.create_sports_product(
            "Camisa Multi", size="P", stock=0, allow_name=1, allow_number=1, allow_backorder=1
        )
        with app.app_context():
            db = get_db()
            variants = {"P": p_variant}
            for size in ("M", "G", "GG", "Infantil 12"):
                variants[size] = db.execute(
                    "INSERT INTO sports_product_variants(product_id,size,stock,min_stock,active) VALUES(?,?,0,1,1)",
                    (product_id, size),
                ).lastrowid
            db.commit()
        self.login_manager()

        requests = [("P", 2, "Nome P", "9"), ("M", 2, "", ""),
                    ("G", 1, "", ""), ("GG", 1, "", ""),
                    ("Infantil 12", 1, "", "")]
        for size, quantity, custom_name, custom_number in requests:
            response = self.client.post("/sale", data=self.sports_sale_form(
                product_id, variants[size], order_mode=["backorder"], quantity=[str(quantity)],
                custom_name=[custom_name], custom_number=[custom_number],
            ))
            self.assertEqual(response.status_code, 303)
        with app.app_context():
            db = get_db()
            rows = db.execute(
                """SELECT d.sale_item_id,d.variant_size,d.custom_name,d.custom_number
                   FROM sports_sale_item_details d ORDER BY d.sale_item_id"""
            ).fetchall()
            by_size = {row["variant_size"]: row for row in rows}
        item_ids = [row["sale_item_id"] for row in rows]
        self.assertEqual(self.client.post(
            "/material-esportivo/vendas/confirmar-envio", json={"sale_item_ids": item_ids}
        ).status_code, 200)

        groups = [
            {"product_id": product_id, "variant_id": variants["P"], "received_quantity": 3,
             "sale_item_ids": [by_size["P"]["sale_item_id"]]},
            {"product_id": product_id, "variant_id": variants["M"], "received_quantity": 2,
             "sale_item_ids": [by_size["M"]["sale_item_id"]]},
            {"product_id": product_id, "variant_id": variants["G"], "received_quantity": 4,
             "sale_item_ids": [by_size["G"]["sale_item_id"]]},
            {"product_id": product_id, "variant_id": variants["GG"], "received_quantity": 1,
             "sale_item_ids": [by_size["GG"]["sale_item_id"]]},
            {"product_id": product_id, "variant_id": variants["Infantil 12"], "received_quantity": "",
             "sale_item_ids": [by_size["Infantil 12"]["sale_item_id"]]},
        ]
        response = self.client.post("/material-esportivo/vendas/receber", json={"groups": groups})
        self.assertEqual(response.status_code, 200, response.get_json())
        self.assertEqual(response.get_json()["groups"], 4)
        with app.app_context():
            db = get_db()
            stocks = {row["size"]: row["stock"] for row in db.execute(
                "SELECT size,stock FROM sports_product_variants WHERE product_id=?", (product_id,)
            ).fetchall()}
            self.assertEqual(stocks, {"P": 1, "M": 0, "G": 3, "GG": 0, "Infantil 12": 0})
            self.assertEqual(db.execute(
                "SELECT fulfillment_status FROM sports_sale_item_details WHERE sale_item_id=?",
                (by_size["Infantil 12"]["sale_item_id"],),
            ).fetchone()["fulfillment_status"], "in_production")
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM notification_outbox WHERE event_type='sports_order_available_push'"
            ).fetchone()["total"], 4)
            preserved = db.execute(
                "SELECT custom_name,custom_number FROM sports_sale_item_details WHERE sale_item_id=?",
                (by_size["P"]["sale_item_id"],),
            ).fetchone()
            self.assertEqual((preserved["custom_name"], preserved["custom_number"]), ("Nome P", "9"))
        repeated = self.client.post("/material-esportivo/vendas/receber", json={"groups": groups[:4]})
        self.assertEqual(repeated.status_code, 409)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT stock FROM sports_product_variants WHERE id=?", (variants["P"],)
            ).fetchone()["stock"], 1)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM notification_outbox WHERE event_type='sports_order_available_push'"
            ).fetchone()["total"], 4)

    def test_sports_available_pix_and_full_credit_unlock_delivery(self):
        from src.routes.sales import apply_mercadopago_status

        self.login_manager()

        def make_available(name):
            product_id, variant_id = self.create_sports_product(
                name, stock=0, allow_backorder=1
            )
            self.client.post("/sale", data=self.sports_sale_form(
                product_id, variant_id, order_mode=["backorder"]
            ))
            with app.app_context():
                row = get_db().execute(
                    """SELECT d.sale_item_id,si.sale_id FROM sports_sale_item_details d
                       JOIN sale_items si ON si.id=d.sale_item_id
                       WHERE si.product_id=? ORDER BY d.sale_item_id DESC LIMIT 1""",
                    (product_id,),
                ).fetchone()
            self.client.post(
                "/material-esportivo/vendas/confirmar-envio",
                json={"sale_item_ids": [row["sale_item_id"]]},
            )
            received = self.client.post("/material-esportivo/vendas/receber", json={
                "sale_item_ids": [row["sale_item_id"]], "variant_id": variant_id,
                "received_quantity": 1,
            })
            self.assertEqual(received.status_code, 200, received.get_json())
            return row["sale_item_id"], row["sale_id"]

        pix_item, pix_sale = make_available("Encomenda Pix Paga")
        pix_order = {"id": "ORDER-BACKORDER-PAID", "transactions": {"payments": [{
            "id": "PAY-BACKORDER-PAID", "payment_method": {"qr_code": "000201SPORTS"},
        }]}}
        with patch("src.routes.sales.create_pix_order", return_value=pix_order):
            started = self.client.post(
                f"/material-esportivo/vendas/{pix_item}/pagamento",
                json={"payment_method": "Pix"},
            )
        self.assertEqual(started.status_code, 200, started.get_json())
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales WHERE id=?", (pix_sale,)).fetchone()
            apply_mercadopago_status(db, sale, {
                "status": "processed", "status_detail": "accredited",
                "total_paid_amount": "20.00", "transactions": {"payments": [{"id": "PAY-BACKORDER-PAID"}]},
            })
        pix_delivery = self.client.post(
            f"/material-esportivo/vendas/{pix_item}/status", json={"to_status": "delivered"}
        )
        self.assertEqual(pix_delivery.status_code, 200, pix_delivery.get_json())

        credit_item, _ = make_available("Encomenda Crédito Paga")
        with app.app_context():
            db = get_db()
            db.execute(
                """INSERT INTO bar_credit_accounts(player_id,balance_cents)
                   VALUES(?,5000) ON CONFLICT(player_id) DO UPDATE SET balance_cents=5000""",
                (self.player_id,),
            )
            db.commit()
        credit_payment = self.client.post(
            f"/material-esportivo/vendas/{credit_item}/pagamento",
            json={"payment_method": "Créditos"},
        )
        self.assertEqual(credit_payment.status_code, 200, credit_payment.get_json())
        credit_delivery = self.client.post(
            f"/material-esportivo/vendas/{credit_item}/status", json={"to_status": "delivered"}
        )
        self.assertEqual(credit_delivery.status_code, 200, credit_delivery.get_json())

    def test_sports_multivariant_invalid_group_rolls_back_all_groups(self):
        product_a, variant_a = self.create_sports_product("Camisa A G", size="G", stock=0, allow_backorder=1)
        product_b, variant_b = self.create_sports_product("Short B G", size="G", stock=0, allow_backorder=1)
        self.login_manager()
        for product_id, variant_id in ((product_a, variant_a), (product_b, variant_b)):
            self.client.post("/sale", data=self.sports_sale_form(
                product_id, variant_id, order_mode=["backorder"]
            ))
        with app.app_context():
            rows = get_db().execute(
                """SELECT d.sale_item_id,si.product_id FROM sports_sale_item_details d
                   JOIN sale_items si ON si.id=d.sale_item_id ORDER BY d.sale_item_id"""
            ).fetchall()
        ids = {row["product_id"]: row["sale_item_id"] for row in rows}
        self.client.post("/material-esportivo/vendas/confirmar-envio", json={"sale_item_ids": list(ids.values())})
        response = self.client.post("/material-esportivo/vendas/receber", json={"groups": [
            {"product_id": product_a, "variant_id": variant_a, "received_quantity": 2,
             "sale_item_ids": [ids[product_a]]},
            {"product_id": product_b, "variant_id": variant_a, "received_quantity": 1,
             "sale_item_ids": [ids[product_b]]},
        ]})
        self.assertIn(response.status_code, (400, 409))
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM sports_sale_item_details WHERE fulfillment_status='in_production'"
            ).fetchone()["total"], 2)
            self.assertEqual(db.execute(
                "SELECT SUM(stock) total FROM sports_product_variants WHERE id IN (?,?)", (variant_a, variant_b)
            ).fetchone()["total"], 0)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) total FROM notification_outbox WHERE event_type='sports_order_available_push'"
            ).fetchone()["total"], 0)

    def create_order(self, order_id, quantity):
        response_data = {
            "id": order_id,
            "status": "action_required",
            "transactions": {"payments": [{
                "id": f"PAY-{order_id}",
                "payment_method": {
                    "id": "pix",
                    "type": "bank_transfer",
                    "qr_code": "000201010212TESTE6304ABCD",
                },
            }]},
        }
        with patch("src.routes.sales.create_pix_order", return_value=response_data):
            response = self.client.post(
                "/pix/mercadopago/orders",
                headers=self.headers(),
                json={
                    "player_id": self.player_id,
                    "items": [{"product_id": self.product_id, "quantity": quantity}],
                },
            )
        self.assertEqual(response.status_code, 201, response.get_json())
        return response.get_json()["sale_id"]

    def test_orders_template_uses_sale_item_id_from_feed_and_restores_button_on_error(self):
        template_path = Path(__file__).resolve().parents[1] / "templates" / "orders.html"
        source = template_path.read_text(encoding="utf-8")
        self.assertIn('data-item-id="${item.id}"', source)
        self.assertNotIn('data-item-id="${index}"', source)
        self.assertNotIn('data-item-id="${item.id+1}"', source)
        self.assertIn("button.disabled=false;button.textContent=original;", source)

    def test_orders_feed_keeps_sale_item_id_distinct_from_product_id(self):
        sale_item_id = 45
        with app.app_context():
            db = get_db()
            sale_id = db.execute(
                """INSERT INTO sales(
                       player_id,payment_method,total_cents,paid,payment_status,
                       ready_for_delivery,paid_at
                   ) VALUES(?, 'Dinheiro', 300, 0, 'pending_cash', 1, CURRENT_TIMESTAMP)""",
                (self.player_id,),
            ).lastrowid
            self.assertNotEqual(sale_item_id, self.product_id)
            db.execute(
                """INSERT INTO sale_items(
                       id,sale_id,product_id,quantity,unit_price_cents,unit_cost_cents
                   ) VALUES(?,?,?,?,?,?)""",
                (sale_item_id, sale_id, self.product_id, 1, 300, 100),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        feed = self.client.get("/orders/feed", headers={"Accept": "application/json"})

        self.assertEqual(feed.status_code, 200)
        order = next(item for item in feed.get_json()["pending"] if item["id"] == sale_id)
        self.assertEqual(order["items"][0]["id"], sale_item_id)

        route_source = (Path(__file__).resolve().parents[1] / "src" / "routes" / "sales.py").read_text(encoding="utf-8")
        self.assertIn("si.id AS sale_item_id", route_source)
        self.assertIn("p.id AS product_id", route_source)
        self.assertIn('"id": row["sale_item_id"]', route_source)

    def test_payment_approval_and_expiration_are_idempotent(self):
        sale_id = self.create_order("ORD-APPROVED", 2)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 3)
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertEqual((sale["paid"], sale["payment_status"]), (0, "pending"))

        approved = {
            "id": "ORD-APPROVED",
            "status": "processed",
            "status_detail": "accredited",
            "total_paid_amount": "6.00",
            "transactions": {"payments": [{"id": "PAY-APPROVED"}]},
        }
        with patch("src.routes.sales.get_order", return_value=approved):
            first = self.client.get(f"/pix/mercadopago/orders/{sale_id}/status", headers=self.headers())
            second = self.client.get(f"/pix/mercadopago/orders/{sale_id}/status", headers=self.headers())
        self.assertTrue(first.get_json()["paid"])
        self.assertTrue(second.get_json()["paid"])
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 3)

        expired_sale_id = self.create_order("ORD-EXPIRED", 1)
        expired = {"id": "ORD-EXPIRED", "status": "expired", "status_detail": "expired", "total_amount": "3.00"}
        with patch("src.routes.sales.get_order", return_value=expired):
            self.client.get(f"/pix/mercadopago/orders/{expired_sale_id}/status", headers=self.headers())
            self.client.get(f"/pix/mercadopago/orders/{expired_sale_id}/status", headers=self.headers())
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 3)
            sale = db.execute("SELECT * FROM sales WHERE id=?", (expired_sale_id,)).fetchone()
            self.assertEqual(sale["payment_status"], "expired")

    def test_delivery_creates_notification_outbox_events_for_delivery_operation(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 300, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                (sale_id, self.product_id, 2, 300, 100),
            )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        response = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["ok"])
        with app.app_context():
            rows = get_db().execute(
                "SELECT event_key,delivery_id,event_type,status FROM notification_outbox WHERE sale_id=? ORDER BY id",
                (sale_id,),
            ).fetchall()
            op_ids = {row["delivery_id"] for row in rows}
        self.assertEqual({row["event_type"] for row in rows}, {"delivery_push", "delivery_update_email", "purchase_receipt_email"})
        self.assertEqual(len(rows), 3)
        self.assertEqual(len(op_ids), 1)
        self.assertTrue(all(row["event_key"].startswith(f"delivery:{next(iter(op_ids))}:") for row in rows))
        self.assertTrue(all(row["status"] == "pending" for row in rows))

    def test_partial_deliveries_generate_distinct_operations_and_event_keys(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 300, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            item_id = db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                (sale_id, self.product_id, 2, 300, 100),
            ).lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        first = self.client.post(f"/orders/{sale_id}/deliver", json={"sale_item_id": item_id, "quantity": 1}, headers={"Accept": "application/json"})
        second = self.client.post(f"/orders/{sale_id}/deliver", json={"sale_item_id": item_id, "quantity": 1}, headers={"Accept": "application/json"})
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        with app.app_context():
            rows = get_db().execute(
                "SELECT delivery_id,event_key FROM notification_outbox WHERE sale_id=? ORDER BY delivery_id,id",
                (sale_id,),
            ).fetchall()
            delivery_ids = {row["delivery_id"] for row in rows}
            event_keys = [row["event_key"] for row in rows]
        self.assertEqual(len(delivery_ids), 2)
        self.assertEqual(len(event_keys), 6)
        self.assertEqual(len(set(event_keys)), 6)

    def test_retry_does_not_duplicate_events_for_same_operation(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 200, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                (sale_id, self.product_id, 1, 200, 100),
            )
            operation_id = db.execute(
                "INSERT INTO sale_delivery_operations(sale_id,delivered_by,delivered_at) VALUES(?,?,CURRENT_TIMESTAMP)",
                (sale_id, self.user_id),
            ).lastrowid
            db.commit()
            payload = {
                "delivery_push": {"player_id": self.player_id, "kind": "pedido_retirada", "period": "retry"},
                "delivery_update_email": {"sale_id": sale_id, "delivered_items": [], "remaining_items": []},
                "purchase_receipt_email": {"sale_id": sale_id},
            }
            first_count = __import__("src.services.notification_outbox", fromlist=["enqueue_delivery_events"]).enqueue_delivery_events(db, sale_id, operation_id, payload)
            second_count = __import__("src.services.notification_outbox", fromlist=["enqueue_delivery_events"]).enqueue_delivery_events(db, sale_id, operation_id, payload)
            event_count = db.execute("SELECT COUNT(*) FROM notification_outbox WHERE sale_id=? AND delivery_id=?", (sale_id, operation_id)).fetchone()[0]
        self.assertEqual(first_count, 3)
        self.assertEqual(second_count, 0)
        self.assertEqual(event_count, 3)

    def test_total_delivery_with_multiple_items_uses_single_operation(self):
        with app.app_context():
            db = get_db()
            product_b = db.execute("INSERT INTO products(name,category,price_cents,cost_cents,stock) VALUES(?,?,?,?,?)", ("Cerveja", "Bebida", 450, 150, 5)).lastrowid
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 750, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)", (sale_id, self.product_id, 1, 300, 100))
            db.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)", (sale_id, product_b, 1, 450, 150))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        response = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(response.status_code, 200)
        with app.app_context():
            operation_count = get_db().execute("SELECT COUNT(*) FROM sale_delivery_operations WHERE sale_id=?", (sale_id,)).fetchone()[0]
            delivery_ids = get_db().execute("SELECT DISTINCT delivery_id FROM notification_outbox WHERE sale_id=? ORDER BY delivery_id", (sale_id,)).fetchall()
        self.assertEqual(operation_count, 1)
        self.assertEqual(len(delivery_ids), 1)

    def test_after_delivery_outbox_failure_rolls_back_sale(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 300, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                (sale_id, self.product_id, 1, 300, 100),
            )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with patch("src.services.notification_outbox.enqueue_delivery_events", side_effect=RuntimeError("boom")):
            response = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(response.status_code, 500)
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT paid,delivered_at FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertIsNone(sale["delivered_at"])
            self.assertEqual(db.execute("SELECT COUNT(*) FROM sale_item_deliveries WHERE sale_item_id IN (SELECT id FROM sale_items WHERE sale_id=?)", (sale_id,)).fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM sale_delivery_operations WHERE sale_id=?", (sale_id,)).fetchone()[0], 0)

    def test_deliver_order_does_not_call_direct_notification_helpers(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 200, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                (sale_id, self.product_id, 1, 200, 100),
            )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with patch("src.services.push_notifications.send_player_push_once", side_effect=AssertionError("direct push")), \
             patch("src.services.purchase_receipts.send_delivery_update", side_effect=AssertionError("direct email")), \
             patch("src.services.purchase_receipts.send_purchase_receipt", side_effect=AssertionError("direct receipt")):
            response = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["receipt_status"], "queued")

    def test_worker_processes_pending_outbox_events_and_marks_sent(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 200, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO notification_outbox(event_key,event_type,sale_id,delivery_id,payload,status,attempts,available_at) VALUES(?,?,?,?,?,'pending',0,CURRENT_TIMESTAMP)",
                (f"delivery:{sale_id}:delivery_push", "delivery_push", sale_id, sale_id, '{"player_id": 1, "kind": "pedido_retirada", "period": "99", "title": "Retirada confirmada", "body": "Teste", "url": "/minha-conta"}'),
            )
            db.commit()
        with patch("src.services.notification_outbox.send_player_push_once", return_value={"sent": 1, "skipped": 0}) as push_mock:
            response = self.client.get(
                "/cron/process-notification-outbox",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sent"], 1)
        self.assertTrue(push_mock.called)

    def test_homologation_outbox_worker_runs_without_external_dispatch(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 200, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            event_key = f"delivery:{sale_id}:homologation_delivery_push"
            db.execute(
                "INSERT INTO notification_outbox(event_key,event_type,sale_id,delivery_id,payload,status,attempts,available_at) VALUES(?,?,?,?,?,'pending',0,CURRENT_TIMESTAMP)",
                (event_key, "delivery_push", sale_id, sale_id, '{}'),
            )
            db.commit()
        homologation = environment_config("homologation")
        with patch.dict(app.config, homologation), \
             patch.dict(os.environ, {"APP_ENV": "homologation"}), \
             patch("src.services.notification_outbox.send_player_push_once") as push_mock:
            response = self.client.get(
                "/cron/process-notification-outbox",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        push_mock.assert_not_called()
        with app.app_context():
            row = get_db().execute(
                "SELECT status,last_error FROM notification_outbox WHERE event_key=?", (event_key,)
            ).fetchone()
        self.assertEqual((row["status"], row["last_error"]), ("sent", "APP_ENV=homologation"))

    def test_worker_retry_and_failed_after_limit(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 200, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO notification_outbox(event_key,event_type,sale_id,delivery_id,payload,status,attempts,available_at) VALUES(?,?,?,?,?,'pending',0,CURRENT_TIMESTAMP)",
                (f"delivery:{sale_id}:retry_delivery_push", "delivery_push", sale_id, sale_id, '{"player_id": 1, "kind": "pedido_retirada", "period": "1", "title": "Retirada", "body": "Teste", "url": "/minha-conta"}'),
            )
            db.commit()
        with patch("src.services.notification_outbox.send_player_push_once", side_effect=RuntimeError("temporary")):
            response = self.client.get(
                "/cron/process-notification-outbox",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        retry_event_key = f"delivery:{sale_id}:retry_delivery_push"
        with app.app_context():
            row = get_db().execute("SELECT status,attempts FROM notification_outbox WHERE event_key=?", (retry_event_key,)).fetchone()
        self.assertEqual(row["status"], "pending")
        self.assertEqual(row["attempts"], 1)

        with app.app_context():
            db = get_db()
            db.execute("UPDATE notification_outbox SET attempts=4, available_at=CURRENT_TIMESTAMP WHERE event_key=?", (retry_event_key,))
            db.commit()
        with patch("src.services.notification_outbox.send_player_push_once", side_effect=RuntimeError("temporary")):
            response = self.client.get(
                "/cron/process-notification-outbox",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        with app.app_context():
            row = get_db().execute("SELECT status,attempts FROM notification_outbox WHERE event_key=?", (retry_event_key,)).fetchone()
        self.assertEqual(row["status"], "failed")
        self.assertEqual(row["attempts"], 5)

    def test_unknown_event_type_fails_gracefully(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (self.player_id, "Dinheiro", 200, 1, "approved"),
            )
            sale_id = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()["id"]
            db.execute(
                "INSERT INTO notification_outbox(event_key,event_type,sale_id,delivery_id,payload,status,attempts,available_at) VALUES(?,?,?,?,?,'pending',0,CURRENT_TIMESTAMP)",
                ("delivery:unknown:test", "delivery_push", sale_id, sale_id, '{}'),
            )
            db.commit()
        response = self.client.get(
            "/cron/process-notification-outbox",
            headers={"Authorization": "Bearer cron-secret-test"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("processed", response.get_json())

    def test_notification_outbox_cron_requires_authentication(self):
        response = self.client.get("/cron/process-notification-outbox")
        self.assertEqual(response.status_code, 401)

    def test_notification_outbox_cron_uses_service_default_batch_size(self):
        result = {"processed": 0, "sent": 0, "retried": 0, "failed": 0}
        with patch("src.services.notification_outbox.process_notification_outbox", return_value=result) as worker:
            response = self.client.get(
                "/cron/process-notification-outbox",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        worker.assert_called_once()
        self.assertEqual(worker.call_args.kwargs, {})

    def test_homologation_blocks_payment_reminders_but_not_outbox_route(self):
        with patch.dict(app.config, environment_config("homologation")), \
             patch("src.routes.finance.dispatch_reminders") as dispatch_mock:
            response = self.client.get(
                "/cron/payment-reminders",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 403)
        dispatch_mock.assert_not_called()

    def test_sqlite_bootstrap_includes_notification_outbox(self):
        with app.app_context():
            row = get_db().execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='notification_outbox'"
            ).fetchone()
        self.assertIsNotNone(row)

    def test_sqlite_migrates_legacy_sale_item_deliveries_without_backfill(self):
        legacy_path = str(Path(self.tempdir.name) / "legacy_delivery.db")
        with sqlite3.connect(legacy_path) as conn:
            conn.execute("CREATE TABLE users(id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT NOT NULL UNIQUE, name TEXT NOT NULL, password_hash TEXT NOT NULL, password_required INTEGER NOT NULL DEFAULT 1, role TEXT NOT NULL DEFAULT 'manager', active INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)")
            conn.execute("CREATE TABLE players(id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, email TEXT DEFAULT '', active INTEGER NOT NULL DEFAULT 1)")
            conn.execute("CREATE TABLE products(id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, category TEXT NOT NULL, price_cents INTEGER NOT NULL, cost_cents INTEGER NOT NULL, stock INTEGER NOT NULL, active INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)")
            conn.execute("CREATE TABLE sales(id INTEGER PRIMARY KEY AUTOINCREMENT, player_id INTEGER NOT NULL REFERENCES players(id), payment_method TEXT NOT NULL, total_cents INTEGER NOT NULL, paid INTEGER NOT NULL DEFAULT 0, payment_status TEXT NOT NULL DEFAULT 'approved', ready_for_delivery INTEGER NOT NULL DEFAULT 0, delivered_at TEXT, delivered_by INTEGER, notes TEXT DEFAULT '', created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)")
            conn.execute("CREATE TABLE sale_items(id INTEGER PRIMARY KEY AUTOINCREMENT, sale_id INTEGER NOT NULL REFERENCES sales(id), product_id INTEGER NOT NULL REFERENCES products(id), quantity INTEGER NOT NULL, unit_price_cents INTEGER NOT NULL, unit_cost_cents INTEGER NOT NULL DEFAULT 0)")
            conn.execute("CREATE TABLE sale_item_deliveries(id INTEGER PRIMARY KEY AUTOINCREMENT, sale_item_id INTEGER NOT NULL REFERENCES sale_items(id), quantity INTEGER NOT NULL, delivered_by INTEGER, delivered_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP)")
            conn.execute("INSERT INTO users(username,name,password_hash,password_required,role,active,created_at) VALUES(?,?,?,?,?,?,CURRENT_TIMESTAMP)", ("legacy", "Legacy", "hash", 1, "manager", 1))
            conn.execute("INSERT INTO players(name,email,active) VALUES(?, ?, 1)", ("Jogador", "jogador@example.com"))
            conn.execute("INSERT INTO products(name,category,price_cents,cost_cents,stock,active,created_at) VALUES(?,?,?,?,?,1,CURRENT_TIMESTAMP)", ("Cerveja", "Bebida", 500, 200, 10))
            conn.execute("INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery,created_at) VALUES(?,?,?,?,?,1,CURRENT_TIMESTAMP)", (1, "Dinheiro", 500, 1, "approved"))
            conn.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)", (1, 1, 2, 500, 200))
            conn.execute("INSERT INTO sale_item_deliveries(sale_item_id,quantity,delivered_by,delivered_at) VALUES(?,?,?,?)", (1, 1, None, "2024-01-01T12:00:00"))
            conn.commit()
        initialize_sqlite_database(legacy_path)
        with sqlite3.connect(legacy_path) as conn:
            columns = {row[1] for row in conn.execute("PRAGMA table_info(sale_item_deliveries)").fetchall()}
            self.assertIn("delivery_operation_id", columns)
            legacy_value = conn.execute("SELECT delivery_operation_id FROM sale_item_deliveries WHERE id=1").fetchone()[0]
            self.assertIsNone(legacy_value)
            sale_id = conn.execute("SELECT sale_id FROM sale_items WHERE id=1").fetchone()[0]
            conn.execute("INSERT INTO sale_delivery_operations(sale_id,delivered_by,delivered_at) VALUES(?,?,CURRENT_TIMESTAMP)", (sale_id, None))
            operation_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.execute(
                "INSERT INTO sale_item_deliveries(sale_item_id,quantity,delivery_operation_id,delivered_by,delivered_at) VALUES(?,?,?,?,CURRENT_TIMESTAMP)",
                (1, 1, operation_id, None),
            )
            row = conn.execute(
                "SELECT delivery_operation_id FROM sale_item_deliveries WHERE id=?",
                (2,),
            ).fetchone()
            self.assertEqual(row[0], operation_id)

    def test_webhook_signature(self):
        data_id = "ORDABC123"
        request_id = "request-123"
        timestamp = "1742505638683"
        template = f"id:{data_id.lower()};request-id:{request_id};ts:{timestamp};"
        signature = hmac.new(b"webhook-secret", template.encode(), hashlib.sha256).hexdigest()
        header = f"ts={timestamp},v1={signature}"
        self.assertTrue(validate_webhook_signature(header, request_id, data_id, "webhook-secret"))
        self.assertFalse(validate_webhook_signature(header, request_id, data_id, "wrong-secret"))

    @patch("src.services.mercadopago._request")
    def test_pix_order_uses_interoperable_bank_transfer(self, request_mock):
        request_mock.return_value = {"id": "ORD-PIX"}
        create_pix_order("token", "pelada_ref", 300, "key", "peladeiro@example.com")
        method, path, token, payload, idempotency_key = request_mock.call_args.args
        payment = payload["transactions"]["payments"][0]
        self.assertEqual((method, path, token, idempotency_key), ("POST", "/v1/orders", "token", "key"))
        self.assertEqual(payload["type"], "online")
        self.assertEqual(payload["processing_mode"], "automatic")
        self.assertEqual(payment["payment_method"], {"id": "pix", "type": "bank_transfer"})
        self.assertEqual(payload["payer"]["email"], "peladeiro@example.com")

    def test_pix_requires_player_email_before_reserving_stock(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET email='' WHERE id=?", (self.player_id,))
            db.commit()
        response = self.client.post(
            "/pix/mercadopago/orders",
            headers=self.headers(),
            json={
                "player_id": self.player_id,
                "items": [{"product_id": self.product_id, "quantity": 1}],
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("e-mail", response.get_json()["error"])
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 5)
            self.assertEqual(db.execute("SELECT COUNT(*) AS total FROM sales").fetchone()["total"], 0)

    def test_dates_use_sao_paulo_business_timezone(self):
        self.assertEqual(brdate(datetime(2026, 7, 14, 0, 30)), "13/07/2026 21:30")
        month, start, end = month_bounds("2026-07")
        self.assertEqual((month, start, end), ("2026-07", "2026-07-01 03:00:00", "2026-08-01 03:00:00"))
        with app.app_context():
            local_day = get_db().execute("SELECT date(?)", ("2026-07-14 00:30:00",)).fetchone()[0]
        self.assertEqual(local_day, "2026-07-13")

    def test_player_first_access_creates_password_and_identifies_sale(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Craque' WHERE id=?", (self.player_id,))
            db.commit()

        first_access = self.client.post("/login", data={"username": "Craque"})
        self.assertEqual(first_access.status_code, 200)
        self.assertIn("Primeiro acesso", first_access.get_data(as_text=True))
        self.assertIn('action="/cliente/senha"', first_access.get_data(as_text=True))
        configured = self.client.post(
            "/cliente/senha",
            data={"password": "senha-segura", "password_confirm": "senha-segura"},
        )
        self.assertEqual(configured.status_code, 303)
        self.assertEqual(configured.headers["Location"], "/minha-conta")
        with self.client.session_transaction() as session:
            client_user_id = session["user_id"]
        with app.app_context():
            db = get_db()
            client_user = db.execute("SELECT * FROM users WHERE id=?", (client_user_id,)).fetchone()
            self.assertEqual((client_user["role"], client_user["player_id"], client_user["username"]), ("client", self.player_id, "Craque"))

        page = self.client.get("/sale").get_data(as_text=True)
        self.assertIn("Craque", page)
        self.assertIn("<strong class=\"d-block fs-5\">Peladeiro</strong>", page)
        self.assertIn("<small>Peladeiro</small>", page)
        self.assertNotIn("Quem está comprando?", page)
        self.assertIn("Novo chamado", self.client.get("/sale").get_data(as_text=True))

        sale = self.client.post(
            "/sale",
            data={"product_id": self.product_id, "quantity": 1, "payment_method": "Dinheiro", "notes": ""},
        )
        self.assertEqual(sale.status_code, 303)
        with app.app_context():
            db = get_db()
            created = db.execute("SELECT player_id,payment_status FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            self.assertEqual((created["player_id"], created["payment_status"]), (self.player_id, "pending_cash"))

    def test_login_without_username_returns_form_instead_of_bad_request(self):
        response = self.client.post("/login", data={"password": "qualquer"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("Informe seu nome de usuário", response.get_data(as_text=True))

    def test_credit_balance_get_is_read_only_and_returns_minimal_payload(self):
        with app.app_context():
            db = get_db()
            client_user_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("saldo.polling", "Saldo Polling", "hash", self.player_id),
            ).lastrowid
            db.execute("DELETE FROM bar_credit_accounts WHERE player_id=?", (self.player_id,))
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id

        connection = connect_db(app)
        statements = []

        class ReadRecorder:
            def execute(self, statement, params=()):
                statements.append(" ".join(statement.split()).upper())
                return connection.execute(statement, params)

        try:
            with patch("src.routes.credits.get_db", return_value=ReadRecorder()):
                missing = self.client.get("/creditos/saldo")
            self.assertEqual(missing.status_code, 200)
            self.assertEqual(missing.get_json(), {"balance_cents": 0})
            self.assertEqual(len(statements), 1)
            self.assertTrue(statements[0].startswith("SELECT BALANCE_CENTS"))
            self.assertFalse(any("INSERT" in statement for statement in statements))

            connection.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,?)",
                (self.player_id, 4321),
            )
            connection.commit()
            statements.clear()
            with patch("src.routes.credits.get_db", return_value=ReadRecorder()):
                existing = self.client.get("/creditos/saldo")
            self.assertEqual(existing.status_code, 200)
            self.assertEqual(existing.get_json(), {"balance_cents": 4321})
            self.assertEqual(len(statements), 1)
            self.assertFalse(any("INSERT" in statement for statement in statements))
        finally:
            connection.close()

        pending_credits = self.client.get("/creditos/pendentes")
        pending_delivery = self.client.get("/minhas-compras/pending-count")
        self.assertEqual(set(pending_credits.get_json()), {"count"})
        self.assertEqual(set(pending_delivery.get_json()), {"count"})

    def test_stage8_credit_and_delivery_polling_contracts(self):
        credit_script = Path("static/credit-pending.js").read_text(encoding="utf-8")
        pwa_script = Path("static/pwa.js").read_text(encoding="utf-8")
        sale_template = Path("templates/sale.html").read_text(encoding="utf-8")
        orders_template = Path("templates/orders.html").read_text(encoding="utf-8")
        history_template = Path("templates/order_history.html").read_text(encoding="utf-8")
        credits_template = Path("templates/credits.html").read_text(encoding="utf-8")

        self.assertIn("CREDIT_PENDING_POLL_INTERVAL_MS = 60000", credit_script)
        self.assertIn("document.visibilityState !== 'visible'", credit_script)
        self.assertIn("state.requestInFlight", credit_script)
        self.assertIn("window.clearInterval(state.timer)", credit_script)
        self.assertIn("visibilitychange", credit_script)
        self.assertIn("refresh();\n    startPolling();", credit_script)

        self.assertIn("DELIVERY_POLL_INTERVAL_MS = 60000", pwa_script)
        self.assertIn('document.visibilityState !== "visible"', pwa_script)
        self.assertIn("deliveryPollingState.requestInFlight", pwa_script)
        self.assertIn("window.clearInterval(deliveryPollingState.timer)", pwa_script)
        self.assertIn("!hasPendingDeliveryIndicator()", pwa_script)
        self.assertIn("visibilitychange", pwa_script)
        self.assertIn("syncPendingDelivery();\n    startDeliveryPolling();", pwa_script)

        self.assertIn("CREDIT_BALANCE_POLL_INTERVAL_MS=60000", sale_template)
        self.assertIn("document.visibilityState!=='visible'", sale_template)
        self.assertIn("creditBalanceRequestInProgress", sale_template)
        self.assertIn("clearInterval(creditRefreshTimer)", sale_template)
        self.assertIn("visibilitychange", sale_template)
        self.assertIn("refreshCreditBalance();startCreditBalancePolling()", sale_template)

        # Os pollings temporários de confirmação Pix continuam rápidos e
        # encerram quando a operação chega a um estado final.
        self.assertIn("pollTimer=setTimeout(()=>pollPayment(statusUrl),5000)", sale_template)
        self.assertIn("pollTimer=setTimeout(()=>pollPayment(statusUrl),7000)", sale_template)
        self.assertIn("setTimeout(async()=>", credits_template)
        self.assertIn("},5000)", credits_template)

        # Operações que mudam entregas continuam atualizando a tela sem
        # aguardar o próximo ciclo de sessenta segundos.
        self.assertIn("await refresh()", orders_template)
        self.assertIn("await refresh()", history_template)

    def test_manager_changes_player_password_from_player_record(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Craque' WHERE id=?", (self.player_id,))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        response = self.client.post(
            f"/players/{self.player_id}/password",
            data={"new_password": "senha-nova-123"},
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            db = get_db()
            user = db.execute("SELECT * FROM users WHERE player_id=?", (self.player_id,)).fetchone()
            self.assertIsNotNone(user)
            self.assertTrue(check_password_hash(user["password_hash"], "senha-nova-123"))

    def test_player_photo_is_saved_and_sent_to_delivery_queue(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        image = BytesIO()
        Image.new("RGB", (24, 24), (20, 100, 180)).save(image, format="JPEG")
        image.seek(0)
        created = self.client.post(
            "/players",
            data={"name": "Peladeiro com foto", "war_name": "Foto", "membership_type": "regular",
                  "photo": (image, "foto.jpg")},
            content_type="multipart/form-data",
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            player = db.execute("SELECT * FROM players WHERE war_name='Foto'").fetchone()
            self.assertTrue(player["thumbnail_data"].startswith("data:image/jpeg;base64,"))
            sale = db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,ready_for_delivery) VALUES(?,?,?,?,?,1)",
                (player["id"], "Dinheiro", 300, 1, "approved"),
            )
            db.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                       (sale.lastrowid, self.product_id, 1, 300, 100))
            db.commit()
        players_page = self.client.get("/players").get_data(as_text=True)
        self.assertIn('alt="Foto de Foto"', players_page)
        feed = self.client.get("/orders/feed")
        self.assertEqual(feed.status_code, 200)
        self.assertTrue(feed.get_json()["pending"][-1]["player_photo"].startswith("data:image/jpeg;base64,"))

    def test_client_can_update_own_photo_from_my_account(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Craque' WHERE id=?", (self.player_id,))
            db.commit()
        self.client.post("/login", data={"username": "Craque"})
        self.client.post("/cliente/senha", data={"password": "senha-segura", "password_confirm": "senha-segura"})
        image = BytesIO()
        Image.new("RGB", (24, 24), (180, 80, 20)).save(image, format="JPEG")
        image.seek(0)
        response = self.client.post("/minha-conta", data={
            "birth_date": "1990-07-17", "phone": "(12) 99999-1111", "emergency_phone": "(12) 98888-2222",
            "postal_code": "12245000", "photo": (image, "perfil.jpg")
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Foto atualizada com sucesso", response.get_data(as_text=True))
        with app.app_context():
            player = get_db().execute("SELECT thumbnail_data FROM players WHERE id=?", (self.player_id,)).fetchone()
            self.assertTrue(player["thumbnail_data"].startswith("data:image/jpeg;base64,"))

    def test_manager_edit_preserves_and_can_replace_player_photo(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        first_photo = BytesIO()
        Image.new("RGB", (800, 600), (30, 90, 160)).save(first_photo, format="JPEG")
        first_photo.seek(0)
        created = self.client.post(
            "/players",
            data={
                "name": "Foto preservada",
                "war_name": "Preserva",
                "gender": "male",
                "membership_type": "regular",
                "photo": (first_photo, "original.jpg"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            player = db.execute(
                "SELECT id,photo_data,thumbnail_data FROM players WHERE war_name=?", ("Preserva",)
            ).fetchone()
            player_id = player["id"]
            original = (player["photo_data"], player["thumbnail_data"])

        preserved = self.client.post(
            f"/players/{player_id}/edit",
            data={
                "name": "Foto preservada",
                "war_name": "Preserva",
                "gender": "male",
                "membership_type": "regular",
                "football_position": "",
                "football_join_date": "",
            },
        )
        self.assertEqual(preserved.status_code, 302)
        with app.app_context():
            player = get_db().execute(
                "SELECT photo_data,thumbnail_data FROM players WHERE id=?", (player_id,)
            ).fetchone()
            self.assertEqual((player["photo_data"], player["thumbnail_data"]), original)

        replacement = BytesIO()
        Image.new("RGB", (800, 600), (180, 70, 35)).save(replacement, format="JPEG")
        replacement.seek(0)
        replaced = self.client.post(
            f"/players/{player_id}/edit",
            data={
                "name": "Foto preservada",
                "war_name": "Preserva",
                "gender": "male",
                "membership_type": "regular",
                "football_position": "",
                "football_join_date": "",
                "photo": (replacement, "nova.jpg"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(replaced.status_code, 302)
        with app.app_context():
            player = get_db().execute(
                "SELECT photo_data,thumbnail_data FROM players WHERE id=?", (player_id,)
            ).fetchone()
            self.assertTrue(player["photo_data"].startswith("data:image/jpeg;base64,"))
            self.assertTrue(player["thumbnail_data"].startswith("data:image/jpeg;base64,"))
            self.assertNotEqual((player["photo_data"], player["thumbnail_data"]), original)

    def test_client_can_use_and_revoke_club_qr_card_without_login(self):
        with app.app_context():
            db = get_db()
            client_user_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("qr-clube", "QR Clube", "hash", self.player_id),
            ).lastrowid
            db.execute("UPDATE players SET war_name='Craque QR' WHERE id=?", (self.player_id,))
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id
        image = BytesIO()
        Image.new("RGB", (96, 96), "white").save(image, format="PNG")
        expected_data = base64.b64encode(image.getvalue()).decode("ascii")
        image.seek(0)
        uploaded = self.client.post(
            "/minha-conta/carteirinha",
            data={"action": "upload", "club_qr": (image, "entrada-gpcta.png")},
            content_type="multipart/form-data",
        )
        self.assertEqual(uploaded.status_code, 302)

        with app.app_context():
            player = get_db().execute(
                "SELECT club_qr_data,club_qr_token FROM players WHERE id=?",
                (self.player_id,),
            ).fetchone()
            first_token = player["club_qr_token"]
            self.assertGreaterEqual(len(first_token), 40)
            self.assertEqual(player["club_qr_data"], f"data:image/png;base64,{expected_data}")

        with self.client.session_transaction() as session:
            session.clear()
        public_card = self.client.get(f"/carteirinha/{first_token}")
        self.assertEqual(public_card.status_code, 200)
        self.assertIn("Craque QR", public_card.get_data(as_text=True))
        self.assertIn("QR Code de entrada no GPCTA", public_card.get_data(as_text=True))
        self.assertIn("QR Code fornecido pelo DCTA", public_card.get_data(as_text=True))
        self.assertNotIn("QR Code fornecido pelo clube", public_card.get_data(as_text=True))
        self.assertIn('class="card-logo-button"', public_card.get_data(as_text=True))
        self.assertIn('aria-label="Voltar à tela de login"', public_card.get_data(as_text=True))
        self.assertIn('action="/logout"', public_card.get_data(as_text=True))
        self.assertIn("no-store", public_card.headers["Cache-Control"])
        self.assertEqual(public_card.headers["X-Robots-Tag"], "noindex, nofollow")
        manifest = self.client.get(f"/carteirinha/{first_token}/manifest.webmanifest")
        self.assertEqual(manifest.status_code, 200)
        self.assertEqual(manifest.get_json()["start_url"], f"/carteirinha/{first_token}")
        self.assertEqual(self.client.get("/carteirinha/token-invalido").status_code, 404)
        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id
        account_html = self.client.get("/minha-conta").get_data(as_text=True)
        self.assertIn('id="toggle-club-card-device"', account_html)
        self.assertIn(f'data-token="{first_token}"', account_html)
        self.assertIn("gpcta-club-card-token", account_html)
        self.assertIn("Ativar neste dispositivo", account_html)
        self.assertIn("Desativar neste dispositivo", account_html)
        self.assertIn("window.addEventListener('pageshow',render)", account_html)
        self.assertNotIn('id="activate-club-card-device"', account_html)
        self.assertNotIn('id="deactivate-club-card-device"', account_html)

        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id
        regenerated = self.client.post(
            "/minha-conta/carteirinha",
            data={"action": "regenerate"},
        )
        self.assertEqual(regenerated.status_code, 302)
        with app.app_context():
            second_token = get_db().execute(
                "SELECT club_qr_token FROM players WHERE id=?", (self.player_id,)
            ).fetchone()["club_qr_token"]
        self.assertNotEqual(first_token, second_token)
        with self.client.session_transaction() as session:
            session.clear()
        self.assertEqual(self.client.get(f"/carteirinha/{first_token}").status_code, 404)
        self.assertEqual(self.client.get(f"/carteirinha/{second_token}").status_code, 200)

        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id
        removed = self.client.post("/minha-conta/carteirinha", data={"action": "remove"})
        self.assertEqual(removed.status_code, 302)
        with self.client.session_transaction() as session:
            session.clear()
        self.assertEqual(self.client.get(f"/carteirinha/{second_token}").status_code, 404)

    def test_login_logo_opens_the_club_card_saved_on_the_device(self):
        page = self.client.get("/login").get_data(as_text=True)
        self.assertIn('id="club-card-logo"', page)
        self.assertIn('aria-label="Abrir carteirinha GPCTA"', page)
        self.assertIn("gpcta-club-card-token", page)
        self.assertIn("fetch(target,{method:'HEAD',cache:'no-store'})", page)
        self.assertNotIn("Toque para abrir sua carteirinha", page)

    def test_manager_can_update_branding_logo(self):
        """The branding settings table is keyed by text, not a numeric id."""
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        image = BytesIO()
        Image.new("RGB", (48, 48), (12, 86, 140)).save(image, format="PNG")
        image.seek(0)
        response = self.client.post(
            "/minha-conta/logo",
            data={"logo": (image, "logo.png")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            setting = get_db().execute(
                "SELECT value FROM app_settings WHERE key=?", ("branding_logo_data",)
            ).fetchone()
            self.assertTrue(setting["value"].startswith("data:image/jpeg;base64,"))
        logo = self.client.get("/branding/logo")
        self.assertEqual(logo.status_code, 200)
        self.assertTrue(logo.data.startswith(b"\xff\xd8"))

    def test_product_photo_and_food_category_are_available_in_quick_sale(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        photo = BytesIO()
        Image.new("RGB", (900, 600), (220, 150, 40)).save(photo, format="PNG")
        photo.seek(0)
        response = self.client.post(
            "/products",
            data={
                "name": "Sanduíche natural",
                "category": "Alimentos",
                "package_type": "",
                "units_per_case": "0",
                "price": "12,50",
                "cost": "7,00",
                "stock": "8",
                "initial_cases": "0",
                "min_stock": "2",
                "supplier_email": "",
                "photo": (photo, "sanduiche.png"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            product = get_db().execute("SELECT * FROM products WHERE name=?", ("Sanduíche natural",)).fetchone()
            self.assertEqual(product["category"], "Alimentos")
            self.assertTrue(product["photo_data"].startswith("data:image/jpeg;base64,"))
            self.assertTrue(product["thumbnail_data"].startswith("data:image/jpeg;base64,"))
            product_id = product["id"]

        sale_page = self.client.get("/sale").get_data(as_text=True)
        self.assertIn("Sanduíche natural", sale_page)
        self.assertIn("Foto de Sanduíche natural", sale_page)
        # UI uses product cards with data-group rather than a select option for categories
        self.assertIn('data-group="Alimentos"', sale_page)

        removed = self.client.post(
            f"/products/{product_id}/edit",
            data={
                "name": "Sanduíche natural",
                "category": "Alimentos",
                "package_type": "",
                "units_per_case": "0",
                "price": "12,50",
                "cost": "7,00",
                "min_stock": "2",
                "stock": "8",
                "stock_reason": "",
                "supplier_email": "",
                "remove_photo": "1",
            },
        )
        self.assertEqual(removed.status_code, 302)
        with app.app_context():
            product = get_db().execute("SELECT photo_data,thumbnail_data FROM products WHERE id=?", (product_id,)).fetchone()
            self.assertEqual((product["photo_data"], product["thumbnail_data"]), ("", ""))

    def test_client_quick_sale_has_live_topbar_cart_indicator(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE users SET role='client',player_id=? WHERE id=?", (self.player_id, self.user_id))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/sale").get_data(as_text=True)
        self.assertIn('id="topbar-cart"', page)
        self.assertIn('id="topbar-cart-count"', page)
        self.assertIn('href="#sale-cart-panel"', page)
        self.assertIn("topbarCart.hidden=quantity===0", page)
        self.assertIn("panel.scrollIntoView({behavior:'smooth'", page)

    def test_client_can_update_profile_and_change_own_password(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Perfil' WHERE id=?", (self.player_id,))
            db.commit()
        self.client.post("/login", data={"username": "Perfil"})
        self.client.post("/cliente/senha", data={"password": "senha-segura", "password_confirm": "senha-segura"})
        response = self.client.post("/minha-conta", data={
            "birth_date": "1990-07-17", "phone": "(12) 99999-1111", "emergency_phone": "Maria (12) 98888-2222",
            "postal_code": "12245000", "address_street": "Rua Teste", "address_number": "50",
            "address_complement": "Casa", "address_neighborhood": "Centro", "address_city": "São José dos Campos", "address_state": "sp",
        })
        self.assertEqual(response.status_code, 200)
        with app.app_context():
            player = get_db().execute("SELECT * FROM players WHERE id=?", (self.player_id,)).fetchone()
            self.assertEqual((player["birth_date"], player["postal_code"], player["address_state"]), ("1990-07-17", "12245000", "SP"))
            self.assertEqual(player["emergency_phone"], "Maria (12) 98888-2222")
        changed = self.client.post("/minha-conta/senha", data={"password": "senha-nova-123", "password_confirm": "senha-nova-123"})
        self.assertEqual(changed.status_code, 302)
        self.client.post("/logout")
        login = self.client.post("/login", data={"username": "Perfil", "password": "senha-nova-123"})
        self.assertEqual(login.status_code, 303)

    def test_birthday_notice_is_shown_to_authenticated_users(self):
        with app.app_context():
            db = get_db()
            today = local_today()
            db.execute("UPDATE players SET war_name='Aniversariante', birth_date=? WHERE id=?",
                       (f"1990-{today.month:02d}-{today.day:02d}", self.player_id))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/finance").get_data(as_text=True)
        self.assertIn("Hoje é aniversário do peladeiro", page)
        self.assertIn("Aniversariante", page)

    def test_client_can_view_month_birthdays_from_sidebar(self):
        with app.app_context():
            db = get_db()
            today = local_today()
            db.execute("UPDATE players SET war_name='Aniversariante', birth_date=? WHERE id=?",
                       (f"1990-{today.month:02d}-{today.day:02d}", self.player_id))
            db.commit()
        self.client.post("/login", data={"username": "Aniversariante"})
        self.client.post("/cliente/senha", data={"password": "senha-segura", "password_confirm": "senha-segura"})
        page = self.client.get("/aniversariantes")
        self.assertEqual(page.status_code, 200)
        self.assertIn("Aniversariantes do mês", page.get_data(as_text=True))
        self.assertIn("Aniversariante", page.get_data(as_text=True))
        sidebar = self.client.get("/sale").get_data(as_text=True)
        self.assertIn("Parabéns, Peladeiro Aniversariante!", sidebar)
        self.assertIn("muitas resenhas e gols na nossa pelada", sidebar)
        self.assertIn("Aniversariantes do mês", sidebar)
        self.assertIn('<span>Compra rápida</span>', sidebar)
        self.assertNotIn('<span>Bar</span>', sidebar)

    def test_female_birthday_message_uses_name_without_peladeiro(self):
        with app.app_context():
            db = get_db()
            today = local_today()
            db.execute("UPDATE players SET war_name='Maria Eduarda', gender='female', birth_date=? WHERE id=?",
                       (f"1990-{today.month:02d}-{today.day:02d}", self.player_id))
            db.commit()
        self.client.post("/login", data={"username": "Maria Eduarda"})
        self.client.post("/cliente/senha", data={"password": "senha-segura", "password_confirm": "senha-segura"})
        page = self.client.get("/sale").get_data(as_text=True)
        self.assertIn("Parabéns, Maria Eduarda!", page)
        self.assertNotIn("Parabéns, Peladeiro Maria Eduarda!", page)

    def test_female_birthday_notice_uses_feminine_generic_message(self):
        with app.app_context():
            db = get_db()
            today = local_today()
            db.execute("UPDATE players SET war_name='Maria Eduarda', gender='female', birth_date=? WHERE id=?",
                       (f"1990-{today.month:02d}-{today.day:02d}", self.player_id))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/finance").get_data(as_text=True)
        self.assertIn("Hoje é aniversário da", page)
        self.assertIn("à aniversariante", page)

    def test_new_maintenance_request_prefills_logged_client_war_name(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Nome de Guerra' WHERE id=?", (self.player_id,))
            db.commit()
        self.client.post("/login", data={"username": "Nome de Guerra"})
        self.client.post("/cliente/senha", data={"password": "senha-segura", "password_confirm": "senha-segura"})
        page = self.client.get("/infra/maintenance/new").get_data(as_text=True)
        self.assertIn('value="Nome de Guerra" readonly', page)
        self.assertIn("Preenchido automaticamente pelo usuário logado", page)

    def test_pix_reconciliation_uses_payment_confirmation_date(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        today = local_today().isoformat()
        with app.app_context():
            db = get_db()
            included = db.execute(
                """INSERT INTO sales(player_id,payment_method,total_cents,paid,created_at,paid_at)
                VALUES(?,'Pix',700,1,'2026-01-01 12:00:00',?)""",
                (self.player_id, f"{today} 15:00:00"),
            ).lastrowid
            excluded = db.execute(
                """INSERT INTO sales(player_id,payment_method,total_cents,paid,created_at,paid_at)
                VALUES(?,'Pix',900,1,?,'2026-01-02 12:00:00')""",
                (self.player_id, f"{today} 15:01:00"),
            ).lastrowid
            db.commit()
        page = self.client.get(f"/pix?day={today}").get_data(as_text=True)
        self.assertIn(f"#{included}", page)
        self.assertNotIn(f"#{excluded}", page)
        self.assertIn("Pix confirmados", page)
        invalid = self.client.get("/pix?day=data-invalida").get_data(as_text=True)
        self.assertIn("data informada era inválida", invalid)

    def test_finance_dashboard_shows_monthly_and_annual_average_ticket(self):
        today = local_today()
        current_month = today.replace(day=1).isoformat()
        previous_month = (today.replace(day=1) - timedelta(days=1)).replace(day=1).isoformat()
        with app.app_context():
            db = get_db()
            for total, paid_at, method in (
                (1000, f"{current_month} 10:00:00", "Pix"),
                (3000, f"{current_month} 11:00:00", "Dinheiro"),
                (6000, f"{previous_month} 12:00:00", "Débito"),
                (9000, f"{current_month} 13:00:00", "Cortesia"),
            ):
                db.execute(
                    """INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,paid_at)
                       VALUES(?,?,?,?,?,?)""",
                    (self.player_id, method, total, 1, "approved", paid_at),
                )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        page = self.client.get("/").get_data(as_text=True)
        self.assertIn("Ticket médio do bar", page)
        self.assertIn("R$ 20,00", page)
        self.assertIn("R$ 33,33", page)
        self.assertIn("2 venda(s) paga(s)", page)
        self.assertIn("3 venda(s) paga(s)", page)
        self.assertIn('id="ticket-average-chart"', page)

    def test_football_dashboard_shows_monthly_annual_attendance_and_extremes(self):
        today = local_today()
        other_months = [month for month in range(1, 13) if month != today.month][:2]
        dates_and_counts = (
            (date(today.year, today.month, 1).isoformat(), 1),
            (date(today.year, other_months[0], 1).isoformat(), 3),
            (date(today.year, other_months[1], 1).isoformat(), 2),
        )
        with app.app_context():
            db = get_db()
            player_ids = [self.player_id]
            for index in range(2):
                player_ids.append(db.execute(
                    "INSERT INTO players(name,war_name) VALUES(?,?)",
                    (f"Participante {index}", f"P{index}"),
                ).lastrowid)
            for match_date, confirmed_count in dates_and_counts:
                sumula_id = db.execute(
                    """INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by)
                       VALUES(?,'SABADO','FINALIZADA',?)""",
                    (match_date, self.user_id),
                ).lastrowid
                for player_id in player_ids[:confirmed_count]:
                    db.execute(
                        "INSERT INTO football_participants(sumula_id,player_id,status) VALUES(?,?,'CONFIRMADO')",
                        (sumula_id, player_id),
                    )
                if confirmed_count < len(player_ids):
                    db.execute(
                        "INSERT INTO football_participants(sumula_id,player_id,status) VALUES(?,?,'AUSENTE')",
                        (sumula_id, player_ids[-1]),
                    )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        page = self.client.get("/futebol").get_data(as_text=True)
        highest_label = ("Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez")[other_months[0] - 1]
        current_label = ("Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez")[today.month - 1]
        self.assertIn("Participação nas peladas", page)
        self.assertIn("Média do mês atual", page)
        self.assertIn("Média anual", page)
        self.assertIn(f"Maior participação: {highest_label}", page)
        self.assertIn(f"Menor participação: {current_label}", page)
        self.assertIn('id="football-attendance-chart"', page)

    def test_football_dashboard_shows_monthly_and_annual_responsibility_leaders(self):
        today = local_today()
        current_dates = (today.replace(day=1), today.replace(day=2))
        earlier_month = 1 if today.month != 1 else 2
        earlier_date = date(today.year, earlier_month, 1)
        with app.app_context():
            db = get_db()
            helper_player_id = db.execute(
                "INSERT INTO players(name,war_name) VALUES(?,?)",
                ("Ajudante", "Apoio"),
            ).lastrowid
            sumula_ids = []
            for match_date in (*current_dates, earlier_date):
                sumula_ids.append(
                    db.execute(
                        "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','FINALIZADA',?)",
                        (match_date.isoformat(), self.user_id),
                    ).lastrowid
                )
            responsibilities = (
                (sumula_ids[0], self.player_id, "SORTEIO"),
                (sumula_ids[0], helper_player_id, "SUMULA"),
                (sumula_ids[0], helper_player_id, "QUADRO"),
                (sumula_ids[0], self.player_id, "ARBITRO_VOLUNTARIO"),
                (sumula_ids[1], self.player_id, "SORTEIO"),
                (sumula_ids[1], self.player_id, "SUMULA"),
                (sumula_ids[1], helper_player_id, "QUADRO"),
                (sumula_ids[1], helper_player_id, "ARBITRO_VOLUNTARIO"),
                (sumula_ids[2], helper_player_id, "SORTEIO"),
                (sumula_ids[2], helper_player_id, "SUMULA"),
                (sumula_ids[2], self.player_id, "QUADRO"),
                (sumula_ids[2], helper_player_id, "ARBITRO_VOLUNTARIO"),
            )
            for sumula_id, player_id, responsibility_type in responsibilities:
                db.execute(
                    "INSERT INTO football_responsibles(sumula_id,player_id,responsibility_type) VALUES(?,?,?)",
                    (sumula_id, player_id, responsibility_type),
                )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/futebol").get_data(as_text=True)
        self.assertIn("Destaques de apoio à gestão", page)
        self.assertRegex(page, r'data-responsibility-period="month" data-responsibility-type="SORTEIO"><strong>Peladeiro</strong><br><span class="text-muted small">2 vezes</span>')
        self.assertRegex(page, r'data-responsibility-period="month" data-responsibility-type="SUMULA"><strong>Apoio, Peladeiro</strong><br><span class="text-muted small">1 vez</span>')
        self.assertRegex(page, r'data-responsibility-period="month" data-responsibility-type="QUADRO"><strong>Apoio</strong><br><span class="text-muted small">2 vezes</span>')
        self.assertRegex(page, r'data-responsibility-period="month" data-responsibility-type="ARBITRO_VOLUNTARIO"><strong>Apoio, Peladeiro</strong><br><span class="text-muted small">1 vez</span>')
        self.assertRegex(page, r'data-responsibility-period="year" data-responsibility-type="SORTEIO"><strong>Peladeiro</strong><br><span class="text-muted small">2 vezes</span>')
        self.assertRegex(page, r'data-responsibility-period="year" data-responsibility-type="SUMULA"><strong>Apoio</strong><br><span class="text-muted small">2 vezes</span>')
        self.assertRegex(page, r'data-responsibility-period="year" data-responsibility-type="QUADRO"><strong>Apoio</strong><br><span class="text-muted small">2 vezes</span>')
        self.assertRegex(page, r'data-responsibility-period="year" data-responsibility-type="ARBITRO_VOLUNTARIO"><strong>Apoio</strong><br><span class="text-muted small">2 vezes</span>')

    def test_football_dashboard_shows_top_five_discipline_ranking(self):
        today = local_today()
        with app.app_context():
            db = get_db()
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','FINALIZADA',?)",
                (today.isoformat(), self.user_id),
            ).lastrowid
            canceled_sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','CANCELADA',?)",
                ((today - timedelta(days=1)).isoformat(), self.user_id),
            ).lastrowid
            ranked_players = []
            for index, name in enumerate(("Alfa", "Bravo", "Charlie", "Delta", "Eco", "Foxtrot"), start=1):
                player_id = db.execute(
                    "INSERT INTO players(name,war_name) VALUES(?,?)", (name, name.upper())
                ).lastrowid
                ranked_players.append(player_id)
                for occurrence in range(7 - index):
                    card = "VERMELHO" if occurrence == 0 and index == 1 else "AMARELO"
                    db.execute(
                        "INSERT INTO football_incidents(sumula_id,type,level,player_id,card,description,created_by) "
                        "VALUES(?,'DISCIPLINAR','INFORMATIVO',?,?,?,?)",
                        (sumula_id, player_id, card, f"Ocorrência {occurrence}", self.user_id),
                    )
            for occurrence in range(10):
                db.execute(
                    "INSERT INTO football_incidents(sumula_id,type,level,player_id,card,description,created_by) "
                    "VALUES(?,'DISCIPLINAR','INFORMATIVO',?,'VERMELHO',?,?)",
                    (canceled_sumula_id, ranked_players[-1], f"Cancelada {occurrence}", self.user_id),
                )
            db.execute(
                "INSERT INTO football_incidents(sumula_id,type,level,player_id,description,created_by) "
                "VALUES(?,'LESAO','INFORMATIVO',?,'Lesão não disciplinar',?)",
                (sumula_id, ranked_players[-1], self.user_id),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/futebol").get_data(as_text=True)
        self.assertIn("Top 5 · Peladeiros mais indisciplinados", page)
        for rank, name in enumerate(("ALFA", "BRAVO", "CHARLIE", "DELTA", "ECO"), start=1):
            self.assertRegex(page, rf'data-discipline-rank="{rank}"[^>]*>.*?<strong>{name}</strong>')
        self.assertNotRegex(page, r'data-discipline-rank="6"')
        self.assertIn("Vermelho: 1", page)
        self.assertIn("Amarelo: 5", page)

    def test_player_names_sort_ignoring_case_and_accents(self):
        names = ["Zeca", "áureo", "Ana", "Álvaro", "bruno"]
        self.assertEqual(
            sorted(names, key=alphabetical_key),
            ["Álvaro", "Ana", "áureo", "bruno", "Zeca"],
        )

    def test_players_page_sorts_by_displayed_name_after_import(self):
        with app.app_context():
            db = get_db()
            db.execute("INSERT INTO players(name,war_name,email) VALUES(?,?,?)", ("Zeca", "", "zeca@example.com"))
            db.execute("INSERT INTO players(name,war_name,email) VALUES(?,?,?)", ("Ana", "Bia", "bia@example.com"))
            db.execute("INSERT INTO players(name,war_name,email) VALUES(?,?,?)", ("Álvaro", "", "alvaro@example.com"))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/players").get_data(as_text=True)
        self.assertLess(page.index("<strong>Álvaro</strong>"), page.index("<strong>Bia</strong>"))
        self.assertLess(page.index("<strong>Bia</strong>"), page.index("<strong>Peladeiro</strong>"))
        self.assertLess(page.index("<strong>Peladeiro</strong>"), page.index("<strong>Zeca</strong>"))
        urgent_page = self.client.get("/urgent").get_data(as_text=True)
        self.assertLess(urgent_page.index("<td>Álvaro</td>"), urgent_page.index("<td>Ana</td>"))
        self.assertLess(urgent_page.index("<td>Ana</td>"), urgent_page.index("<td>Peladeiro</td>"))
        self.assertLess(urgent_page.index("<td>Peladeiro</td>"), urgent_page.index("<td>Zeca</td>"))

    def test_manager_can_list_complete_player_records_and_download_pdf(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Craque', birth_date='1990-07-17', phone='11999999999', emergency_phone='11888888888', postal_code='12245000', address_street='Rua Teste', address_number='50', address_city='São José dos Campos', address_state='SP' WHERE id=?", (self.player_id,))
            for index in range(11):
                db.execute("INSERT INTO players(name,war_name) VALUES(?,?)", (f"Peladeiro Extra {index}", f"Extra{index}"))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/players/report?q=Craque")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("Cadastro completo dos peladeiros", html)
        self.assertIn("Craque", html)
        self.assertIn("Rua Teste", html)
        self.assertIn(f"/players/report/{self.player_id}", html)
        self.assertIn('target="_blank"', html)
        detail = self.client.get(f"/players/report/{self.player_id}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn("Contato de emergência", detail.get_data(as_text=True))
        report = self.client.get("/players/report.pdf?q=Craque")
        self.assertEqual(report.status_code, 200)
        self.assertTrue(report.data.startswith(b"%PDF-"))
        self.assertIn("cadastro-completo-peladeiros.pdf", report.headers["Content-Disposition"])
        paged = self.client.get("/players/report?page=2")
        self.assertEqual(paged.status_code, 200)
        self.assertIn("Próxima", paged.get_data(as_text=True))

    def test_manager_sidebar_groups_modules_and_links(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/players").get_data(as_text=True)
        self.assertIn('id="app-sidebar"', page)
        modules = ["Bar", "Financeiro", "Infraestrutura", "Relatórios", "Urgente", "Administração"]
        positions = [page.index(f"<span>{label}</span>") for label in modules]
        self.assertEqual(positions, sorted(positions))
        for links in (
            ["Caixa", "Conferir Pix", "Estoque", "Produtos", "Pedidos", "Venda rápida"],
            ["Mensalidades", "Livro-caixa", "Lembretes"],
            ["Manutenção", "Materiais", "Relação de Carga"],
            ["Peladeiros", "Cadastro completo / PDF", "Usuários"],
        ):
            link_positions = [page.index(f">{label}</a>") for label in links]
            self.assertEqual(link_positions, sorted(link_positions))
        self.assertIn('data-bs-target="#sidebar-bar"', page)
        self.assertIn(">Bar e vendas</a>", page)
        self.assertIn(">Compras e estoque</a>", page)
        self.assertIn(">Consolidado</a>", page)
        self.assertIn('class="offcanvas-lg offcanvas-start app-sidebar"', page)
        self.assertIn('alt="Logo GPCTA"', page)
        self.assertNotIn('class="navbar ', page)
        self.assertNotIn('class="sidebar-user"', page)
        self.assertIn('class="topbar-account"', page)
        # The account block now wraps the manager name in a link so it is
        # reachable from the top bar as well as the sidebar.
        self.assertIn('<strong>Teste</strong>', page)
        self.assertIn('<small>Gerente</small>', page)
        self.assertEqual(page.count('<strong>Teste</strong>'), 1)
        self.assertEqual(page.count('>Relatórios</a>'), 0)
        finance_page = self.client.get("/finance").get_data(as_text=True)
        self.assertIn('<input type="month" name="start_month"', finance_page)
        self.assertIn(f'value="{local_today().strftime("%Y-%m")}"', finance_page)
        self.assertIn('action="/logout"', page)
        self.assertIn('class="logout-button"', page)
        self.assertIn('aria-label="Sair do sistema"', page)
        self.assertIn('id="pwa-install"', page)
        self.assertIn("Aniversariantes do mês", page)
        self.assertIn('href="/aniversariantes"', page)

    def test_display_profile_opens_live_orders_and_birthdays_panel(self):
        with app.app_context():
            db = get_db()
            cursor = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'display')",
                ("painel", "Painel da TV", "hash"),
            )
            display_id = cursor.lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = display_id
        panel = self.client.get("/painel")
        self.assertEqual(panel.status_code, 200)
        html = panel.get_data(as_text=True)
        self.assertIn("Pedidos em andamento", html)
        self.assertIn("Aniversariantes do mês", html)
        self.assertEqual(self.client.get("/painel/feed").status_code, 200)

    def test_display_user_is_created_and_logs_in_without_password(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        created = self.client.post("/users", data={
            "name": "Monitor", "username": "monitor-tv", "role": "display", "password": "",
        })
        self.assertEqual(created.status_code, 302)
        self.client.post("/logout")
        login = self.client.post("/login", data={"username": "monitor-tv", "password": ""})
        self.assertEqual(login.status_code, 303)
        self.assertTrue(login.headers["Location"].endswith("/painel"))

    def test_urgent_is_visible_and_accessible_to_every_user_role(self):
        with app.app_context():
            db = get_db()
            role_ids = {"manager": self.user_id}
            for role in ("staff", "client", "infra", "maintenance"):
                cursor = db.execute(
                    "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,?)",
                    (f"teste.{role}", f"Teste {role}", "hash", role),
                )
                role_ids[role] = cursor.lastrowid
            db.commit()

        for role, user_id in role_ids.items():
            with self.subTest(role=role):
                with self.client.session_transaction() as session:
                    session["user_id"] = user_id
                response = self.client.get("/urgent")
                self.assertEqual(response.status_code, 200)
                page = response.get_data(as_text=True)
                self.assertIn('class="sidebar-module sidebar-direct urgent active"', page)
                self.assertIn("<span>Urgente</span>", page)

    def test_passwordless_maintenance_user_only_opens_new_requests(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        created = self.client.post(
            "/users",
            data={"name": "Portaria", "username": "manutencao", "role": "maintenance", "password": ""},
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            user = get_db().execute("SELECT * FROM users WHERE username='manutencao'").fetchone()
            self.assertEqual((user["role"], user["password_required"]), ("maintenance", 0))
            maintenance_user_id = user["id"]

        self.client.post("/logout")
        login = self.client.post(
            "/login", data={"username": "manutencao", "password": "", "next": "/logout"}
        )
        self.assertEqual(login.status_code, 303)
        self.assertTrue(login.headers["Location"].endswith("/infra/maintenance/new"))

        form = self.client.get("/infra/maintenance/new")
        self.assertEqual(form.status_code, 200)
        page = form.get_data(as_text=True)
        self.assertIn("<span>Novo chamado</span>", page)
        self.assertIn("<span>Urgente</span>", page)
        self.assertNotIn("<span>Infraestrutura</span>", page)
        self.assertNotIn("Acompanhamento e resolução", page)
        self.assertIn('id="pwa-install"', page)
        self.assertNotIn("← Voltar", page)

        submitted = self.client.post(
            "/infra/maintenance/new",
            data={
                "title": "Lâmpada queimada",
                "area_code": "SAL",
                "location": "Entrada principal",
                "category": "electrical",
                "priority": "medium",
                "description": "A luminária da entrada não acende.",
                "occurred_on": "2026-07-14",
                "notes": "Verificar antes do evento.",
                "status": "completed",
                "responsible": "valor indevido",
                "cost": "999,99",
            },
        )
        self.assertEqual(submitted.status_code, 302)
        self.assertTrue(submitted.headers["Location"].endswith("/infra/maintenance/new"))
        with app.app_context():
            maintenance = get_db().execute(
                "SELECT * FROM maintenance_requests WHERE created_by=?", (maintenance_user_id,)
            ).fetchone()
            self.assertIsNotNone(maintenance)
            self.assertEqual(
                (maintenance["status"], maintenance["responsible"], maintenance["cost_cents"], maintenance["notes"]),
                ("open", "", 0, "Verificar antes do evento."),
            )

        for forbidden_path in ("/infra/maintenance", "/infra/materials", "/sale", "/users"):
            denied = self.client.get(forbidden_path)
            self.assertEqual(denied.status_code, 302)
            self.assertTrue(denied.headers["Location"].endswith("/infra/maintenance/new"))

        stale_post = self.client.post("/", headers={"Accept": "text/html"})
        self.assertEqual(stale_post.status_code, 303)
        self.assertTrue(stale_post.headers["Location"].endswith("/infra/maintenance/new"))

    def test_staff_sees_bar_and_can_only_open_new_maintenance_requests(self):
        with app.app_context():
            db = get_db()
            cursor = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'staff')",
                ("atendente", "Atendente", "hash"),
            )
            staff_id = cursor.lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = staff_id

        form = self.client.get("/infra/maintenance/new")
        self.assertEqual(form.status_code, 200)
        page = form.get_data(as_text=True)
        self.assertIn("<span>Bar</span>", page)
        self.assertIn("<span>Infraestrutura</span>", page)
        self.assertIn(">Novo chamado</a>", page)
        self.assertIn(">Relação de Carga</a>", page)
        self.assertIn(">Validar conferência</a>", page)
        self.assertIn("<span>Urgente</span>", page)
        self.assertNotIn(">Materiais</a>", page)
        self.assertNotIn(">Manutenção</a>", page)
        self.assertNotIn("Acompanhamento e resolução", page)

        relation = self.client.get("/infra/load-relation")
        self.assertEqual(relation.status_code, 200)
        self.assertNotIn("Edição em lote", relation.get_data(as_text=True))
        check = self.client.get("/infra/load-relation/check")
        self.assertEqual(check.status_code, 200)

        submitted = self.client.post(
            "/infra/maintenance/new",
            data={
                "title": "Torneira pingando",
                "area_code": "BAR",
                "location": "Pia do balcão",
                "category": "plumbing",
                "priority": "high",
                "description": "A torneira não fecha completamente.",
                "occurred_on": "2026-07-15",
                "status": "completed",
                "responsible": "valor indevido",
                "cost": "500,00",
            },
        )
        self.assertEqual(submitted.status_code, 302)
        self.assertTrue(submitted.headers["Location"].endswith("/infra/maintenance/new"))
        with app.app_context():
            maintenance = get_db().execute(
                "SELECT * FROM maintenance_requests WHERE created_by=?", (staff_id,)
            ).fetchone()
            self.assertEqual(
                (maintenance["status"], maintenance["responsible"], maintenance["cost_cents"]),
                ("open", "", 0),
            )

        for forbidden_path in ("/infra/maintenance", "/infra/materials", "/infra/load-relation/qr-codes"):
            denied = self.client.get(forbidden_path)
            self.assertEqual(denied.status_code, 302)
            self.assertEqual(denied.headers["Location"], "/orders")

    def test_client_can_create_maintenance_request_without_internal_access_error(self):
        """A abertura pelo peladeiro não deve redirecionar para detalhes internos."""
        with app.app_context():
            db = get_db()
            db.execute("UPDATE players SET war_name='Peladeiro' WHERE id=?", (self.player_id,))
            db.execute(
                "UPDATE users SET role='client', player_id=? WHERE id=?",
                (self.player_id, self.user_id),
            )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        response = self.client.post(
            "/infra/maintenance/new",
            data={
                "title": "Lâmpada queimada",
                "area_code": "BAR",
                "location": "Balcão",
                "category": "electrical",
                "priority": "medium",
                "description": "A lâmpada não acende.",
                "occurred_on": "2026-08-05",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn("Chamado MAN-", page)
        self.assertNotIn("Seu usuário não possui acesso a essa funcionalidade.", page)
        with app.app_context():
            request_row = get_db().execute(
                "SELECT id,created_by,status FROM maintenance_requests WHERE created_by=? ORDER BY id DESC LIMIT 1",
                (self.user_id,),
            ).fetchone()
            self.assertIsNotNone(request_row)
            self.assertEqual(request_row["status"], "open")
            history_row = get_db().execute(
                "SELECT status,observation FROM maintenance_request_history WHERE request_id=(SELECT id FROM maintenance_requests WHERE created_by=? ORDER BY id DESC LIMIT 1)",
                (self.user_id,),
            ).fetchone()
            self.assertEqual((history_row["status"], history_row["observation"]), ("open", ""))

        mine = self.client.get("/infra/maintenance/mine")
        self.assertEqual(mine.status_code, 200)
        mine_page = mine.get_data(as_text=True)
        self.assertIn("Meus chamados", mine_page)
        self.assertIn("Lâmpada queimada", mine_page)
        self.assertIn("Aberto", mine_page)
        self.assertIn("Detalhes", mine_page)
        self.assertIn("Chamados por status", mine_page)
        self.assertIn('data-status-count="open">1</strong>', mine_page)
        self.assertIn('data-status-count="completed">0</strong>', mine_page)
        self.assertIn("maintenance-card-priority-medium", mine_page)
        self.assertIn("Tempo aberto:", mine_page)
        self.assertIn("maintenance-age-green", mine_page)
        # A listagem é resumida; os dados completos ficam na tela de detalhes.
        self.assertNotIn("A lâmpada não acende.", mine_page)
        self.assertNotIn("/edit", mine_page)

        request_id = request_row["id"]
        detail = self.client.get(f"/infra/maintenance/mine/{request_id}")
        self.assertEqual(detail.status_code, 200)
        detail_page = detail.get_data(as_text=True)
        self.assertIn("Lâmpada queimada", detail_page)
        self.assertIn("A lâmpada não acende.", detail_page)
        self.assertIn("← Meus chamados", detail_page)
        self.assertNotIn("Editar", detail_page)

        # A tela detalhada também deve respeitar o vínculo com o peladeiro.
        denied_detail = self.client.get("/infra/maintenance/999999")
        self.assertEqual(denied_detail.status_code, 302)

        # Chamados antigos podem ter sido gravados por outro registro de
        # usuário, mas ainda pertencem ao mesmo peladeiro. Eles devem aparecer
        # em "Meus chamados" pelo vínculo users.player_id.
        with app.app_context():
            db = get_db()
            legacy_user = db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("peladeiro-legado", "Peladeiro legado", "hash", self.player_id),
            )
            legacy_user_id = legacy_user.lastrowid
            db.execute(
                """INSERT INTO maintenance_requests
                   (code,title,area_code,location,category,priority,description,
                    occurred_on,created_by)
                   VALUES(?,?,?,?,?,?,?,?,?)""",
                (
                    "MAN-LEGACY-0001",
                    "Chamado antigo",
                    "BAR",
                    "Balcão",
                    "electrical",
                    "medium",
                    "Registro criado antes da conta atual.",
                    "2026-08-01",
                    legacy_user_id,
                ),
            )
            db.commit()

        mine_after_legacy = self.client.get("/infra/maintenance/mine")
        self.assertEqual(mine_after_legacy.status_code, 200)
        self.assertIn("Chamado antigo", mine_after_legacy.get_data(as_text=True))

        duplicate = self.client.post(
            "/infra/maintenance/new",
            data={
                "title": "  lâmpada QUEIMADA ",
                "area_code": "BAR",
                "location": "Outro local",
                "category": "electrical",
                "priority": "high",
                "description": "Outro relato para o mesmo problema.",
                "occurred_on": "2026-08-05",
            },
            follow_redirects=True,
        )
        self.assertEqual(duplicate.status_code, 200)
        duplicate_page = duplicate.get_data(as_text=True)
        self.assertIn("Já existe o chamado MAN-", duplicate_page)
        self.assertIn("aberto por Peladeiro", duplicate_page)
        with app.app_context():
            self.assertEqual(
                get_db().execute(
                    "SELECT COUNT(*) total FROM maintenance_requests WHERE created_by=?", (self.user_id,)
                ).fetchone()["total"],
                1,
            )

    def test_photo_processor_preserves_dimensions_formats_metadata_and_transparency(self):
        def upload(raw, filename):
            return FileStorage(stream=BytesIO(raw), filename=filename)

        def encoded_image(data_url):
            self.assertTrue(data_url.startswith("data:image/jpeg;base64,"))
            raw = base64.b64decode(data_url.split(",", 1)[1], validate=True)
            image = Image.open(BytesIO(raw))
            image.load()
            return raw, image

        large = Image.effect_noise((1600, 800), 45).convert("RGB")
        large_buffer = BytesIO()
        large.save(large_buffer, format="JPEG", quality=95)
        original_bytes = large_buffer.getvalue()
        photo, thumbnail = process_material_photo(upload(original_bytes, "grande.jpg"))
        final_bytes, final = encoded_image(photo)
        thumbnail_bytes, thumb = encoded_image(thumbnail)
        self.assertEqual(final.size, (1200, 600))
        self.assertEqual(thumb.size, (180, 90))
        self.assertLess(len(final_bytes), len(original_bytes))
        self.assertGreater(len(thumbnail_bytes), 0)

        small_buffer = BytesIO()
        Image.new("RGB", (320, 180), "navy").save(small_buffer, format="JPEG")
        small_photo, _ = process_material_photo(upload(small_buffer.getvalue(), "pequena.jpg"))
        _, small = encoded_image(small_photo)
        self.assertEqual(small.size, (320, 180))

        transparent = Image.new("RGBA", (40, 30), (0, 0, 0, 0))
        transparent.paste((220, 20, 20, 255), (12, 8, 28, 22))
        png_buffer = BytesIO()
        transparent.save(png_buffer, format="PNG")
        png_photo, _ = process_material_photo(upload(png_buffer.getvalue(), "alpha.png"))
        _, flattened = encoded_image(png_photo)
        self.assertTrue(all(channel >= 245 for channel in flattened.getpixel((2, 2))))
        self.assertGreater(flattened.getpixel((20, 15))[0], 180)

        webp_buffer = BytesIO()
        Image.new("RGB", (90, 60), "green").save(webp_buffer, format="WEBP")
        webp_photo, _ = process_material_photo(upload(webp_buffer.getvalue(), "entrada.webp"))
        _, webp_result = encoded_image(webp_photo)
        self.assertEqual((webp_result.format, webp_result.size), ("JPEG", (90, 60)))

        exif = Image.Exif()
        exif[274] = 6
        exif[270] = "metadata must not persist"
        oriented_buffer = BytesIO()
        Image.new("RGB", (60, 30), "orange").save(oriented_buffer, format="JPEG", exif=exif)
        oriented_photo, _ = process_material_photo(upload(oriented_buffer.getvalue(), "orientada.jpg"))
        _, oriented = encoded_image(oriented_photo)
        self.assertEqual(oriented.size, (30, 60))
        self.assertEqual(len(oriented.getexif()), 0)

    def test_photo_processor_rejects_invalid_corrupt_oversized_and_excessive_pixels(self):
        def upload(raw, filename):
            return FileStorage(stream=BytesIO(raw), filename=filename)

        for raw, filename in ((b"not-an-image", "fake.jpg"), (b"\xff\xd8corrupt", "corrupt.jpg")):
            with self.subTest(filename=filename), self.assertRaisesRegex(ValueError, "inv\u00e1lida|corrompida"):
                process_material_photo(upload(raw, filename))
        with self.assertRaisesRegex(ValueError, "4 MB"):
            process_material_photo(upload(b"x" * (MAX_UPLOAD_BYTES + 1), "large.jpg"))
        pixels = BytesIO()
        Image.new("RGB", (5000, 4001), "white").save(pixels, format="PNG")
        with self.assertRaisesRegex(ValueError, "resolu\u00e7\u00e3o"):
            process_material_photo(upload(pixels.getvalue(), "pixels.png"))

    def test_load_entry_photo_failure_is_atomic_and_large_request_is_friendly(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            material_id = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Carga at\u00f4mica", "FCG-ATOM"),
            ).lastrowid
            db.commit()

        valid = BytesIO()
        Image.new("RGB", (640, 480), "blue").save(valid, format="JPEG")
        invalid = self.client.post(
            "/infra/load-relation/new",
            data={
                "material_id": str(material_id),
                "area_code": "BAR",
                "photos": [(BytesIO(valid.getvalue()), "ok.jpg"), (BytesIO(b"invalid"), "bad.jpg")],
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(invalid.status_code, 200)
        self.assertIn("foto enviada \u00e9 inv\u00e1lida", invalid.get_data(as_text=True))
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) FROM load_entries").fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM load_entry_photos").fetchone()[0], 0)

        oversized = self.client.post(
            "/infra/load-relation/new",
            data={
                "material_id": str(material_id),
                "area_code": "BAR",
                "photos": (BytesIO(b"x" * (4 * 1024 * 1024)), "too-large.jpg"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        self.assertEqual(oversized.status_code, 200)
        self.assertIn("no m\u00e1ximo 4 MB no total", oversized.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(get_db().execute("SELECT COUNT(*) FROM load_entries").fetchone()[0], 0)

    def test_material_crud_with_optimized_photo(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        invalid = self.client.post("/infra/materials/new", data={"description": ""})
        self.assertEqual(invalid.status_code, 200)
        self.assertIn("descrição é obrigatória", invalid.get_data(as_text=True))
        invalid_photo = self.client.post(
            "/infra/materials/new",
            data={"description": "Teste", "photo": (BytesIO(b"nao-e-imagem"), "foto.png")},
            content_type="multipart/form-data",
        )
        self.assertEqual(invalid_photo.status_code, 200)
        self.assertIn("foto enviada é inválida", invalid_photo.get_data(as_text=True))

        photo = BytesIO()
        Image.new("RGB", (1400, 900), color=(20, 110, 180)).save(photo, format="PNG")
        photo.seek(0)
        created = self.client.post(
            "/infra/materials/new",
            data={
                "description": "Analisador de espectro",
                "load_sheet": "FCG-1877",
                "notes": "Material em bom estado.",
                "photo": (photo, "analisador.png"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            material = get_db().execute("SELECT * FROM materials").fetchone()
            material_id = material["id"]
            original_photo = material["photo_data"]
            self.assertTrue(original_photo.startswith("data:image/jpeg;base64,"))
            self.assertTrue(material["thumbnail_data"].startswith("data:image/jpeg;base64,"))

        listing = self.client.get("/infra/materials?q=espectro").get_data(as_text=True)
        self.assertIn("Analisador de espectro", listing)
        self.assertIn("FCG-1877", listing)
        detail = self.client.get(f"/infra/materials/{material_id}").get_data(as_text=True)
        self.assertIn("Material em bom estado.", detail)
        self.assertIn("FCG - Código de controle patrimonial", detail)

        edited = self.client.post(
            f"/infra/materials/{material_id}/edit",
            data={"description": "Analisador atualizado", "load_sheet": "FCG-2000", "notes": "Revisado."},
        )
        self.assertEqual(edited.status_code, 302)
        with app.app_context():
            material = get_db().execute("SELECT * FROM materials WHERE id=?", (material_id,)).fetchone()
            self.assertEqual((material["description"], material["photo_data"]), ("Analisador atualizado", original_photo))

        removed = self.client.post(
            f"/infra/materials/{material_id}/edit",
            data={"description": "Analisador atualizado", "load_sheet": "", "notes": "", "remove_photo": "1"},
        )
        self.assertEqual(removed.status_code, 302)
        with app.app_context():
            material = get_db().execute("SELECT * FROM materials WHERE id=?", (material_id,)).fetchone()
            self.assertEqual((material["photo_data"], material["thumbnail_data"]), ("", ""))

        deleted = self.client.post(f"/infra/materials/{material_id}/delete")
        self.assertEqual(deleted.status_code, 302)
        with app.app_context():
            self.assertEqual(get_db().execute("SELECT COUNT(*) FROM materials").fetchone()[0], 0)

        self.assertEqual(self.client.get("/infra/load-relation").status_code, 200)

    def test_load_relation_crud_generates_bmp_photos_and_pdf(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            cursor = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Cadeira giratória", "FCG-1317918"),
            )
            material_id = cursor.lastrowid
            db.commit()

        missing_material = self.client.post("/infra/load-relation/new", data={"material_id": ""})
        self.assertEqual(missing_material.status_code, 200)
        self.assertIn("Selecione um material", missing_material.get_data(as_text=True))

        photos = []
        for index, color in enumerate(((25, 90, 150), (180, 110, 30)), start=1):
            photo = BytesIO()
            Image.new("RGB", (800, 600), color=color).save(photo, format="JPEG")
            photo.seek(0)
            photos.append((photo, f"foto-{index}.jpg"))
        created = self.client.post(
            "/infra/load-relation/new",
            data={
                "material_id": str(material_id),
                "area_code": "COZ",
                "serial_number": "SERIE-001",
                "location": "Sala G-7",
                "notes": "Carga em bom estado.",
                "photos": photos,
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            entry = db.execute("SELECT * FROM load_entries").fetchone()
            entry_id = entry["id"]
            self.assertEqual((entry["bmp"], entry["area_code"]), (f"BMP-{entry_id:06d} | COZ", "COZ"))
            stored_photos = db.execute(
                "SELECT * FROM load_entry_photos WHERE load_entry_id=? ORDER BY id", (entry_id,)
            ).fetchall()
            self.assertEqual(len(stored_photos), 2)
            self.assertTrue(stored_photos[0]["thumbnail_data"].startswith("data:image/jpeg;base64,"))
            first_photo_id = stored_photos[0]["id"]

        listing = self.client.get("/infra/load-relation?q=cadeira").get_data(as_text=True)
        self.assertIn("Cadeira giratória", listing)
        self.assertIn(f"BMP-{entry_id:06d}", listing)
        self.assertIn("| COZ", listing)
        filtered_listing = self.client.get("/infra/load-relation?area=BAR").get_data(as_text=True)
        self.assertNotIn("Cadeira giratória", filtered_listing)
        detail = self.client.get(f"/infra/load-relation/{entry_id}").get_data(as_text=True)
        self.assertIn("Carga em bom estado.", detail)
        self.assertIn("SERIE-001", detail)
        self.assertIn("Pendente", detail)
        check_page = self.client.get("/infra/load-relation/check")
        self.assertEqual(check_page.status_code, 200)
        self.assertIn(f"BMP-{entry_id:06d} | COZ", check_page.get_data(as_text=True))
        checked = self.client.post(f"/infra/load-relation/{entry_id}/check")
        self.assertEqual(checked.status_code, 302)
        with app.app_context():
            checked_entry = get_db().execute("SELECT * FROM load_entries WHERE id=?", (entry_id,)).fetchone()
            self.assertIsNotNone(checked_entry["last_checked_at"])
            self.assertEqual(checked_entry["last_checked_by"], self.user_id)
            self.assertIsNotNone(checked_entry["next_check_due_at"])
        checked_detail = self.client.get(f"/infra/load-relation/{entry_id}").get_data(as_text=True)
        self.assertIn("Válida até", checked_detail)
        with app.app_context():
            db = get_db()
            db.execute(
                """INSERT INTO load_entry_photos
                   (load_entry_id,photo_data,thumbnail_data,photo_kind,captured_by)
                   VALUES(?,?,?,?,?)""",
                (entry_id, "data:image/jpeg;base64,AAAA", "data:image/jpeg;base64,BBBB", "conference", self.user_id),
            )
            db.commit()
        timeline = self.client.get(f"/infra/load-relation/{entry_id}").get_data(as_text=True)
        self.assertIn("Linha do tempo das conferências", timeline)
        self.assertIn("Foto anterior", timeline)
        self.assertIn("Foto da conferência", timeline)

        qr_page = self.client.get(f"/infra/load-relation/{entry_id}/qr-code")
        self.assertEqual(qr_page.status_code, 200)
        self.assertIn("data:image/png;base64,", qr_page.get_data(as_text=True))
        self.assertIn(f"/infra/load-relation/{entry_id}", qr_page.get_data(as_text=True))

        qr_selection = self.client.get("/infra/load-relation/qr-codes?area=COZ")
        self.assertEqual(qr_selection.status_code, 200)
        self.assertIn(f"BMP-{entry_id:06d} | COZ", qr_selection.get_data(as_text=True))
        labels = self.client.post(
            "/infra/load-relation/qr-codes.pdf",
            data={"entry_ids": str(entry_id), "size": "standard", "area_code": "COZ"},
        )
        self.assertEqual(labels.status_code, 200)
        self.assertEqual(labels.mimetype, "application/pdf")
        self.assertTrue(labels.data.startswith(b"%PDF-"))

        blocked_material_delete = self.client.post(f"/infra/materials/{material_id}/delete")
        self.assertEqual(blocked_material_delete.status_code, 302)
        with app.app_context():
            self.assertIsNotNone(
                get_db().execute("SELECT id FROM materials WHERE id=?", (material_id,)).fetchone()
            )

        edited = self.client.post(
            f"/infra/load-relation/{entry_id}/edit",
            data={
                "material_id": str(material_id),
                "area_code": "SAL",
                "serial_number": "SERIE-002",
                "location": "Armário H-14",
                "notes": "Inventariado.",
                "remove_photo_ids": str(first_photo_id),
            },
        )
        self.assertEqual(edited.status_code, 302)
        with app.app_context():
            db = get_db()
            entry = db.execute("SELECT * FROM load_entries WHERE id=?", (entry_id,)).fetchone()
            photo_count = db.execute(
                "SELECT COUNT(*) FROM load_entry_photos WHERE load_entry_id=?", (entry_id,)
            ).fetchone()[0]
            self.assertEqual(
                (entry["bmp"], entry["area_code"], entry["serial_number"], entry["location"], photo_count),
                (f"BMP-{entry_id:06d} | SAL", "SAL", "SERIE-002", "Armário H-14", 2),
            )

        report = self.client.get("/infra/load-relation/report.pdf?q=cadeira")
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report.mimetype, "application/pdf")
        self.assertTrue(report.data.startswith(b"%PDF-"))
        self.assertIn("attachment", report.headers["Content-Disposition"])

        discharged = self.client.post(f"/infra/load-relation/{entry_id}/discharge")
        self.assertEqual(discharged.status_code, 302)
        with app.app_context():
            entry = get_db().execute("SELECT * FROM load_entries WHERE id=?", (entry_id,)).fetchone()
            self.assertEqual((entry["status"], entry["discharged_by"]), ("discharged", self.user_id))
            self.assertIsNotNone(entry["discharged_at"])
        listing = self.client.get("/infra/load-relation").get_data(as_text=True)
        self.assertIn("Descarregado", listing)
        self.assertNotIn(f'action="/infra/load-relation/{entry_id}/discharge"', listing)

        deleted = self.client.post(f"/infra/load-relation/{entry_id}/delete")
        self.assertEqual(deleted.status_code, 302)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) FROM load_entries").fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM load_entry_photos").fetchone()[0], 0)

    def test_load_entry_accepts_internal_headquarters_area(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            material_id = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Material interno", "FCG-INT"),
            ).lastrowid
            db.commit()

        form = self.client.get("/infra/load-relation/new")
        self.assertIn("INT - Interno sede", form.get_data(as_text=True))
        created = self.client.post(
            "/infra/load-relation/new",
            data={"material_id": str(material_id), "area_code": "INT", "status": "active"},
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            entry = get_db().execute(
                "SELECT bmp,area_code FROM load_entries WHERE material_id=?", (material_id,)
            ).fetchone()
            self.assertEqual(entry["area_code"], "INT")
            self.assertTrue(entry["bmp"].endswith(" | INT"))

    def test_qr_code_selection_filters_by_conference_status_and_combines_filters(self):
        self.login_manager()
        with app.app_context():
            db = get_db()
            chair = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES('Cadeira QR','FCG-QR-1')"
            ).lastrowid
            table = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES('Mesa QR','FCG-QR-2')"
            ).lastrowid
            entries = [
                (chair, "BMP-QR-PENDING-BAR", "BAR", None),
                (chair, "BMP-QR-CHECKED-BAR", "BAR", "2026-08-01 10:00:00"),
                (table, "BMP-QR-PENDING-COZ", "COZ", None),
                (table, "BMP-QR-CHECKED-COZ", "COZ", "2026-08-02 10:00:00"),
            ]
            for entry in entries:
                db.execute(
                    """INSERT INTO load_entries(material_id,bmp,area_code,last_checked_at)
                       VALUES(?,?,?,?)""",
                    entry,
                )
            db.commit()

        def page(query=""):
            response = self.client.get(f"/infra/load-relation/qr-codes{query}")
            self.assertEqual(response.status_code, 200)
            return response.get_data(as_text=True)

        all_entries = page()
        for bmp in ("BMP-QR-PENDING-BAR", "BMP-QR-CHECKED-BAR",
                    "BMP-QR-PENDING-COZ", "BMP-QR-CHECKED-COZ"):
            self.assertIn(bmp, all_entries)

        missing = page("?conference_status=missing")
        self.assertIn("BMP-QR-PENDING-BAR", missing)
        self.assertIn("BMP-QR-PENDING-COZ", missing)
        self.assertNotIn("BMP-QR-CHECKED-BAR", missing)
        self.assertNotIn("BMP-QR-CHECKED-COZ", missing)

        checked = page("?conference_status=checked")
        self.assertIn("BMP-QR-CHECKED-BAR", checked)
        self.assertIn("BMP-QR-CHECKED-COZ", checked)
        self.assertNotIn("BMP-QR-PENDING-BAR", checked)
        self.assertNotIn("BMP-QR-PENDING-COZ", checked)

        area_missing = page("?area=BAR&conference_status=missing")
        self.assertIn("BMP-QR-PENDING-BAR", area_missing)
        self.assertNotIn("BMP-QR-PENDING-COZ", area_missing)

        material_missing = page(f"?material_id={table}&conference_status=missing")
        self.assertIn("BMP-QR-PENDING-COZ", material_missing)
        self.assertNotIn("BMP-QR-PENDING-BAR", material_missing)

        combined = page(f"?area=COZ&material_id={table}&conference_status=missing")
        self.assertIn("BMP-QR-PENDING-COZ", combined)
        self.assertNotIn("BMP-QR-PENDING-BAR", combined)

        empty = page("?area=HIS&conference_status=missing")
        self.assertIn("Nenhum BMP", empty)

    def test_qr_load_check_requires_and_atomically_stores_photo_evidence(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            material_id = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Carga para conferência", "FCG-QR"),
            ).lastrowid
            entry_id = db.execute(
                """INSERT INTO load_entries(material_id,bmp,area_code,status)
                   VALUES(?,?,'BAR','active')""",
                (material_id, "BMP-QR | BAR"),
            ).lastrowid
            db.commit()

        missing = self.client.post(f"/infra/load-relation/{entry_id}/check-auto")
        self.assertEqual(missing.status_code, 400)
        self.assertIn("foto", missing.get_json()["error"].lower())
        with app.app_context():
            db = get_db()
            entry = db.execute("SELECT * FROM load_entries WHERE id=?", (entry_id,)).fetchone()
            self.assertIsNone(entry["last_checked_at"])
            self.assertEqual(
                db.execute("SELECT COUNT(*) FROM load_entry_photos WHERE load_entry_id=?", (entry_id,)).fetchone()[0],
                0,
            )

        photo = BytesIO()
        Image.new("RGB", (900, 700), color=(55, 125, 75)).save(photo, format="JPEG")
        photo.seek(0)
        checked = self.client.post(
            f"/infra/load-relation/{entry_id}/check-auto",
            data={"photo": (photo, "conferencia.jpg")},
            content_type="multipart/form-data",
        )
        self.assertEqual(checked.status_code, 200, checked.get_json())
        self.assertTrue(checked.get_json()["ok"])
        with app.app_context():
            db = get_db()
            entry = db.execute("SELECT * FROM load_entries WHERE id=?", (entry_id,)).fetchone()
            evidence = db.execute(
                "SELECT * FROM load_entry_photos WHERE load_entry_id=?", (entry_id,)
            ).fetchone()
            self.assertIsNotNone(entry["last_checked_at"])
            self.assertEqual(entry["last_checked_by"], self.user_id)
            self.assertEqual(evidence["photo_kind"], "reference")
            self.assertEqual(evidence["captured_by"], self.user_id)
            self.assertIsNotNone(evidence["captured_at"])
            self.assertTrue(evidence["photo_data"].startswith("data:image/jpeg;base64,"))

        detail = self.client.get(f"/infra/load-relation/{entry_id}").get_data(as_text=True)
        self.assertIn("Referência inicial", detail)
        self.assertIn("utilizada como referência", detail)
        self.assertIn("Por Teste", detail)

    def test_load_check_counter_counts_distinct_successes_in_current_tab_operation(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            material_id = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Cargas da operação", "FCG-OP"),
            ).lastrowid
            entry_ids = []
            for index in range(1, 10):
                entry_ids.append(db.execute(
                    """INSERT INTO load_entries(material_id,bmp,area_code,status)
                       VALUES(?,?,'BAR','active')""",
                    (material_id, f"BMP-OP-{index} | BAR"),
                ).lastrowid)
            other_user_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'staff')",
                ("conferente.outro", "Outro conferente", "hash"),
            ).lastrowid
            db.commit()

        def conference(entry_id, operation_ids=()):
            photo = BytesIO()
            Image.new("RGB", (80, 60), color=(30, 100, 160)).save(photo, format="JPEG")
            photo.seek(0)
            return self.client.post(
                f"/infra/load-relation/{entry_id}/check-auto",
                data={
                    "operation_entry_ids": [str(value) for value in operation_ids],
                    "photo": (photo, f"carga-{entry_id}.jpg"),
                },
                content_type="multipart/form-data",
            )

        # A successful conference from before this screen operation must not
        # enter the new tab's counter.
        previous = conference(entry_ids[0])
        self.assertEqual(previous.get_json()["checked_count"], 1)
        page = self.client.get("/infra/load-relation/check").get_data(as_text=True)
        self.assertIn('<strong id="checked-count">0</strong>', page)
        self.assertIn("body.append('operation_entry_ids', seenId)", page)
        self.assertIn("data.checked_count", page)

        current_operation = []
        first = conference(entry_ids[1], current_operation)
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.get_json()["checked_count"], 1)
        current_operation.append(entry_ids[1])

        duplicate = conference(entry_ids[1], current_operation)
        self.assertEqual(duplicate.get_json()["checked_count"], 1)

        for expected_count, entry_id in enumerate(entry_ids[2:6], start=2):
            response = conference(entry_id, current_operation)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.get_json()["checked_count"], expected_count)
            current_operation.append(entry_id)
        self.assertEqual(len(current_operation), 5)

        # A second tab starts with an empty operation list and therefore has
        # an independent counter without resetting the first tab.
        second_tab = conference(entry_ids[6])
        self.assertEqual(second_tab.get_json()["checked_count"], 1)
        first_tab_again = conference(entry_ids[5], current_operation)
        self.assertEqual(first_tab_again.get_json()["checked_count"], 5)

        # Persisted evidence belonging to another user is not included even
        # if its ID is submitted as part of this operation.
        with self.client.session_transaction() as session:
            session["user_id"] = other_user_id
        other_user = conference(entry_ids[7])
        self.assertEqual(other_user.get_json()["checked_count"], 1)
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        restricted = conference(entry_ids[5], [*current_operation, entry_ids[7]])
        self.assertEqual(restricted.get_json()["checked_count"], 5)

        with app.app_context():
            db = get_db()
            never_checked = db.execute(
                "SELECT last_checked_at FROM load_entries WHERE id=?", (entry_ids[8],)
            ).fetchone()
            self.assertIsNone(never_checked["last_checked_at"])

    def test_load_conference_filters_separate_valid_missing_expired_and_keep_pending(self):
        today = local_today()
        with app.app_context():
            db = get_db()
            material_id = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Cargas para filtros", "FCG-FILTRO"),
            ).lastrowid
            rows = (
                ("BMP-FILTRO-VALIDA | BAR", "active", today.isoformat(), (today + timedelta(days=10)).isoformat()),
                ("BMP-FILTRO-SEM | BAR", "active", None, None),
                ("BMP-FILTRO-VENCIDA | BAR", "active", today.isoformat(), (today - timedelta(days=1)).isoformat()),
                ("BMP-FILTRO-BAIXADA | BAR", "discharged", today.isoformat(), (today + timedelta(days=10)).isoformat()),
            )
            for bmp, status, last_checked_at, next_check_due_at in rows:
                db.execute(
                    """INSERT INTO load_entries(
                           material_id,bmp,area_code,status,last_checked_at,last_checked_by,next_check_due_at
                       ) VALUES(?,?,'BAR',?,?,?,?)""",
                    (material_id, bmp, status, last_checked_at, self.user_id if last_checked_at else None, next_check_due_at),
                )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        relation = self.client.get("/infra/load-relation").get_data(as_text=True)
        self.assertIn('<option value="">Todas</option>', relation)
        self.assertIn('value="valid"', relation)
        self.assertIn("Conferência validada", relation)
        self.assertIn("Sem conferência", relation)
        self.assertIn("Conferência vencida", relation)

        valid = self.client.get("/infra/load-relation?due=valid").get_data(as_text=True)
        self.assertIn("BMP-FILTRO-VALIDA", valid)
        self.assertNotIn("BMP-FILTRO-SEM | BAR</strong>", valid)
        self.assertNotIn("BMP-FILTRO-VENCIDA | BAR</strong>", valid)
        self.assertNotIn("BMP-FILTRO-BAIXADA | BAR</strong>", valid)

        missing = self.client.get("/infra/load-relation?due=missing").get_data(as_text=True)
        self.assertIn("BMP-FILTRO-SEM", missing)
        self.assertNotIn("BMP-FILTRO-VALIDA | BAR</strong>", missing)
        self.assertNotIn("BMP-FILTRO-VENCIDA | BAR</strong>", missing)

        expired = self.client.get("/infra/load-relation?due=expired").get_data(as_text=True)
        self.assertIn("BMP-FILTRO-VENCIDA", expired)
        self.assertNotIn("BMP-FILTRO-VALIDA | BAR</strong>", expired)
        self.assertNotIn("BMP-FILTRO-SEM | BAR</strong>", expired)

        legacy_pending = self.client.get("/infra/load-relation?due=pending").get_data(as_text=True)
        self.assertIn("BMP-FILTRO-SEM", legacy_pending)
        self.assertIn("BMP-FILTRO-VENCIDA", legacy_pending)
        self.assertNotIn("BMP-FILTRO-VALIDA | BAR</strong>", legacy_pending)

        report = self.client.get("/infra/load-relation/report?due=valid").get_data(as_text=True)
        self.assertIn("Conferência validada", report)
        self.assertIn("BMP-FILTRO-VALIDA", report)
        self.assertNotIn("BMP-FILTRO-VENCIDA</td>", report)

    def test_load_batch_movement_status_and_report_filters(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            db.execute("INSERT INTO materials(description,load_sheet) VALUES(?,?)", ("Banqueta", "FCG-9"))
            material_id = db.execute("SELECT id FROM materials WHERE description=?", ("Banqueta",)).fetchone()["id"]
            db.commit()
        response = self.client.post("/infra/load-relation/batch", data={
            "material_id": str(material_id), "quantity": "3", "area_code": "SAL",
            "serial_prefix": "BAN-", "location": "Salão", "responsible": "Equipe A",
            "status": "active", "notes": "Lote de teste",
        })
        self.assertEqual(response.status_code, 200)
        with app.app_context():
            db = get_db()
            entries = db.execute("SELECT * FROM load_entries ORDER BY id").fetchall()
            self.assertEqual(len(entries), 3)
            self.assertEqual([entry["bmp"] for entry in entries], [f"BMP-{entry['id']:06d} | SAL" for entry in entries])
            entry_id = entries[0]["id"]
        moved = self.client.post(f"/infra/load-relation/{entry_id}/move", data={
            "location": "Sala 2", "responsible": "Equipe B", "reason": "Remanejamento patrimonial"
        })
        self.assertEqual(moved.status_code, 302)
        changed = self.client.post(f"/infra/load-relation/{entry_id}/status", data={"status": "maintenance"})
        self.assertEqual(changed.status_code, 302)
        detail = self.client.get(f"/infra/load-relation/{entry_id}")
        self.assertEqual(detail.status_code, 200)
        body = detail.get_data(as_text=True)
        self.assertIn("Remanejamento patrimonial", body)
        self.assertIn("Em manutenção", body)
        report = self.client.get("/infra/load-relation/report?q=banqueta&status=maintenance")
        self.assertEqual(report.status_code, 200)
        self.assertIn("Equipe B", report.get_data(as_text=True))

    def test_load_loans_batch_partial_return_and_availability(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            material_id = db.execute(
                "INSERT INTO materials(description,load_sheet) VALUES(?,?)",
                ("Cadeira plástica", "FCG-CADEIRA"),
            ).lastrowid
            for index in range(6):
                cursor = db.execute(
                    """INSERT INTO load_entries(material_id,bmp,area_code,location,status)
                       VALUES(?,?,'INT','Sede','active')""",
                    (material_id, f"BMP-LOAN-{index}"),
                )
            db.commit()

        departure_photo = BytesIO()
        Image.new("RGB", (800, 600), color=(80, 110, 150)).save(departure_photo, format="JPEG")
        departure_photo.seek(0)
        created = self.client.post(
            "/infra/loans",
            data={
                "borrower_name": "Fernando Silva",
                "borrower_phone": "11999999999",
                "borrower_document": "DOC-123",
                "checkout_on": local_today().isoformat(),
                "due_on": (local_today() + timedelta(days=3)).isoformat(),
                "material_id": [str(material_id)],
                "quantity": ["4"],
                "notes": "Festa familiar",
                "departure_photo": (departure_photo, "retirada.jpg"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            loan = db.execute("SELECT * FROM load_loans").fetchone()
            loan_id = loan["id"]
            item = db.execute("SELECT * FROM load_loan_items WHERE loan_id=?", (loan_id,)).fetchone()
            self.assertEqual((loan["status"], item["quantity"], item["returned_quantity"]), ("open", 4, 0))
            self.assertTrue(loan["departure_photo_data"].startswith("data:image/jpeg;base64,"))

        blocked = self.client.post(
            "/infra/loans",
            data={
                "borrower_name": "Outro responsável",
                "checkout_on": local_today().isoformat(),
                "due_on": (local_today() + timedelta(days=2)).isoformat(),
                "material_id": [str(material_id)],
                "quantity": ["3"],
            },
        )
        self.assertEqual(blocked.status_code, 200)
        self.assertIn("somente 2 unidade(s) disponível(is)", blocked.get_data(as_text=True))

        partial = self.client.post(f"/infra/loans/{loan_id}/return", data={f"return_{item['id']}": "2"})
        self.assertEqual(partial.status_code, 302)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT status FROM load_loans WHERE id=?", (loan_id,)).fetchone()[0], "partial")
            self.assertEqual(db.execute("SELECT returned_quantity FROM load_loan_items WHERE id=?", (item["id"],)).fetchone()[0], 2)

        completed = self.client.post(f"/infra/loans/{loan_id}/return", data={f"return_{item['id']}": "2"})
        self.assertEqual(completed.status_code, 302)
        detail = self.client.get(f"/infra/loans/{loan_id}").get_data(as_text=True)
        self.assertIn("Devolvido", detail)
        self.assertIn("Fernando Silva", detail)
        with app.app_context():
            db = get_db()
            loan = db.execute("SELECT * FROM load_loans WHERE id=?", (loan_id,)).fetchone()
            self.assertEqual(loan["status"], "returned")
            self.assertIsNotNone(loan["returned_at"])
            self.assertEqual(db.execute("SELECT COUNT(*) FROM load_loan_history WHERE loan_id=?", (loan_id,)).fetchone()[0], 3)

    def test_maintenance_list_colors_priority_and_open_age(self):
        today = local_today()
        with app.app_context():
            db = get_db()
            for index, (priority, age) in enumerate((("low", 12), ("medium", 20), ("high", 30), ("urgent", 5)), start=1):
                db.execute(
                    """INSERT INTO maintenance_requests
                       (code,title,area_code,category,priority,description,status,occurred_on,created_at,created_by)
                       VALUES(?,?,?,'electrical',?,?,'open',?,?,?)""",
                    (
                        f"MAN-AGE-{index}", f"Chamado idade {age}", "BAR", priority,
                        "Teste de indicador de tempo.", today.isoformat(),
                        (today - timedelta(days=age)).isoformat(), self.user_id,
                    ),
                )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        page = self.client.get("/infra/maintenance").get_data(as_text=True)
        for priority in ("low", "medium", "high", "urgent"):
            self.assertIn(f"maintenance-row-priority-{priority}", page)
        self.assertIn('class="maintenance-age-green">12</strong>', page)
        self.assertIn('class="maintenance-age-orange">20</strong>', page)
        self.assertIn('class="maintenance-age-red">30</strong>', page)
        self.assertIn("Tempo aberto", page)
        self.assertIn("16 a 25 dias", page)
        self.assertIn('<option value="open" selected>Aberto</option>', page)

    def test_maintenance_crud_dashboard_photos_and_report(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        invalid = self.client.post("/infra/maintenance/new", data={"title": ""})
        self.assertEqual(invalid.status_code, 200)
        self.assertIn("título do problema é obrigatório", invalid.get_data(as_text=True))

        problem_photo = BytesIO()
        Image.new("RGB", (900, 700), color=(180, 60, 40)).save(problem_photo, format="JPEG")
        problem_photo.seek(0)
        created = self.client.post(
            "/infra/maintenance/new",
            data={
                "title": "Vazamento no banheiro",
                "area_code": "BAN",
                "location": "Banheiro masculino",
                "category": "plumbing",
                "priority": "urgent",
                "description": "Vazamento próximo ao lavatório.",
                "responsible": "Equipe hidráulica",
                "status": "open",
                "occurred_on": "2026-07-14",
                "due_on": "2026-07-15",
                "cost": "0,00",
                "problem_photos": (problem_photo, "problema.jpg"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            maintenance = db.execute("SELECT * FROM maintenance_requests").fetchone()
            request_id = maintenance["id"]
            self.assertEqual((maintenance["code"], maintenance["area_code"]), (f"MAN-{request_id:06d}", "BAN"))
            self.assertEqual(db.execute("SELECT COUNT(*) FROM maintenance_photos").fetchone()[0], 1)

        listing = self.client.get("/infra/maintenance?area=BAN&priority=urgent")
        self.assertEqual(listing.status_code, 200)
        self.assertIn("Vazamento no banheiro", listing.get_data(as_text=True))
        dashboard = self.client.get("/infra/maintenance/dashboard")
        self.assertEqual(dashboard.status_code, 200)
        self.assertIn("Painel de manutenção", dashboard.get_data(as_text=True))
        detail = self.client.get(f"/infra/maintenance/{request_id}")
        self.assertIn("Vazamento próximo", detail.get_data(as_text=True))

        resolution_photo = BytesIO()
        Image.new("RGB", (900, 700), color=(40, 150, 80)).save(resolution_photo, format="JPEG")
        resolution_photo.seek(0)
        completed = self.client.post(
            f"/infra/maintenance/{request_id}/edit",
            data={
                "title": "Vazamento no banheiro",
                "area_code": "BAN",
                "location": "Banheiro masculino",
                "category": "plumbing",
                "priority": "urgent",
                "description": "Vazamento próximo ao lavatório.",
                "responsible": "Equipe hidráulica",
                "status": "completed",
                "occurred_on": "2026-07-14",
                "due_on": "2026-07-15",
                "completed_on": "2026-07-14",
                "resolution": "Sifão substituído e instalação testada.",
                "cost": "125,50",
                "notes": "Serviço conferido.",
                "resolution_photos": (resolution_photo, "resolucao.jpg"),
            },
            content_type="multipart/form-data",
        )
        self.assertEqual(completed.status_code, 302)
        with app.app_context():
            db = get_db()
            maintenance = db.execute("SELECT * FROM maintenance_requests WHERE id=?", (request_id,)).fetchone()
            self.assertEqual((maintenance["status"], maintenance["cost_cents"]), ("completed", 12550))
            self.assertEqual(db.execute("SELECT COUNT(*) FROM maintenance_photos").fetchone()[0], 2)

        report = self.client.get("/infra/maintenance/report.pdf?area=BAN")
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report.mimetype, "application/pdf")
        self.assertTrue(report.data.startswith(b"%PDF-"))

        deleted = self.client.post(f"/infra/maintenance/{request_id}/delete")
        self.assertEqual(deleted.status_code, 302)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) FROM maintenance_requests").fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM maintenance_photos").fetchone()[0], 0)

    def test_login_shows_centered_logo_without_navigation_bar_and_copyright(self):
        page = self.client.get("/login").get_data(as_text=True)
        self.assertIn('class="club-card-logo-button mb-3"', page)
        self.assertIn('class="login-logo"', page)
        self.assertNotIn('class="navbar ', page)
        self.assertIn("PELADEIROS GPCTA", page)
        self.assertNotIn("BAR PELADEIROS GPCTA", page)
        self.assertIn("Copyright © 2026 | Grupo de Peladas do CTA - GPCTA", page)
        self.assertNotIn(">Sair</button>", page)

    def test_pwa_assets_are_public_installable_and_do_not_cache_private_pages(self):
        login = self.client.get("/login").get_data(as_text=True)
        self.assertIn('rel="manifest" href="/static/manifest.webmanifest"', login)
        self.assertIn('rel="apple-touch-icon"', login)
        self.assertIn('id="pwa-install"', login)
        self.assertIn('src="/static/pwa.js"', login)

        manifest_response = self.client.get("/static/manifest.webmanifest")
        self.assertEqual(manifest_response.status_code, 200)
        manifest = manifest_response.get_json()
        manifest_response.close()
        self.assertEqual(manifest["name"], "Peladeiros GPCTA")
        self.assertEqual(manifest["display"], "standalone")
        self.assertEqual({icon["sizes"] for icon in manifest["icons"]}, {"192x192", "512x512"})

        worker_response = self.client.get("/service-worker.js")
        worker = worker_response.get_data(as_text=True)
        worker_response.close()
        self.assertEqual(worker_response.status_code, 200)
        self.assertEqual(worker_response.headers["Service-Worker-Allowed"], "/")
        self.assertIn('const OFFLINE_URL = "/offline"', worker)
        self.assertIn('data.web_push === 8030', worker)
        self.assertIn('notification.navigate || notification.url', worker)
        self.assertNotIn('request.mode === "navigate"', worker)
        self.assertNotIn('"/sale"', worker)
        self.assertNotIn('"/finance"', worker)
        self.assertEqual(self.client.get("/offline").status_code, 200)

    def test_transient_database_failure_preserves_authenticated_session(self):
        class TransientError(RuntimeError):
            pgcode = "08006"

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        with patch("src.db.get_db", side_effect=TransientError("falha temporária simulada")):
            response = self.client.get("/players")

        self.assertEqual(response.status_code, 503)
        self.assertEqual(response.headers["Retry-After"], "3")
        self.assertIn("Sua sessão foi preservada", response.get_data(as_text=True))
        with self.client.session_transaction() as session:
            self.assertEqual(session.get("user_id"), self.user_id)

        with patch("app.get_db", side_effect=TransientError("falha temporária simulada")):
            static_response = self.client.get("/static/pwa.js")
        self.assertEqual(static_response.status_code, 200)
        static_response.close()

    def test_two_simultaneous_session_reads_are_read_only(self):
        """Concurrent page/unread-count requests must not race on the user row."""
        with app.app_context():
            db = get_db()
            db.execute("INSERT INTO players(name,war_name) VALUES(?,?)", ("Leitor", "leitor"))
            player_id = db.execute("SELECT id FROM players WHERE war_name=?", ("leitor",)).fetchone()["id"]
            db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("leitor", "Leitor", "hash", player_id),
            )
            db.commit()
            client_user_id = db.execute("SELECT id FROM users WHERE username=?", ("leitor",)).fetchone()["id"]

        clients = [app.test_client(), app.test_client()]
        for client in clients:
            with client.session_transaction() as session:
                session["user_id"] = client_user_id

        def fetch_unread(client):
            return client.get("/notifications/push/unread-count", headers={"Accept": "application/json"})

        # The schema is already prepared by setUp; prevent this test's two
        # independent connections from needlessly rerunning SQLite setup.
        with patch("src.db.init_sqlite"):
            with ThreadPoolExecutor(max_workers=2) as pool:
                responses = list(pool.map(fetch_unread, clients))

        self.assertEqual([response.status_code for response in responses], [200, 200])
        self.assertEqual([response.get_json()["count"] for response in responses], [0, 0])
        self.assertTrue(all(set(response.get_json()) == {"count"} for response in responses))

    def test_common_http_requests_never_execute_schema_sql(self):
        """Regression barrier for DDL/schema repair inside normal HTTP requests."""
        with app.app_context():
            db = get_db()
            db.execute("INSERT INTO players(name,war_name) VALUES(?,?)", ("Guarda DDL", "guardaddl"))
            player_id = db.execute("SELECT id FROM players WHERE war_name=?", ("guardaddl",)).fetchone()["id"]
            db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("guardaddl", "Guarda DDL", "hash", player_id),
            )
            db.commit()
            client_user_id = db.execute("SELECT id FROM users WHERE username=?", ("guardaddl",)).fetchone()["id"]

        original_execute = DbWrapper.execute
        statements = []

        def reject_schema_sql(wrapper, statement, params=()):
            normalized = " ".join(str(statement).split())
            if has_request_context():
                statements.append(normalized)
                if self.SCHEMA_SQL.match(normalized):
                    raise AssertionError(f"DDL executado durante request HTTP: {normalized[:120]}")
                if re.match(r"^\s*(?:INSERT\s+INTO|UPDATE|DELETE\s+FROM)\s+users\b", normalized, re.IGNORECASE):
                    raise AssertionError(f"Carregamento de sessao escreveu em users: {normalized[:120]}")
            return original_execute(wrapper, statement, params)

        with patch.object(DbWrapper, "execute", new=reject_schema_sql), patch("src.db.init_sqlite"):
            self.client.get("/login")

            with self.client.session_transaction() as session:
                session["user_id"] = self.user_id
            for endpoint in ("/", "/players", "/futebol", "/health"):
                response = self.client.get(endpoint)
                self.assertNotEqual(response.status_code, 500, endpoint)

            with self.client.session_transaction() as session:
                session["user_id"] = client_user_id
            response = self.client.get(
                "/notifications/push/unread-count", headers={"Accept": "application/json"}
            )
            self.assertEqual(response.status_code, 200)

        self.assertTrue(statements)
        self.assertFalse(any(self.SCHEMA_SQL.match(statement) for statement in statements))

    def test_priority_player_queries_do_not_load_original_photos(self):
        """Lists, authentication and match sheets must never fetch photo_data."""
        statements = []
        original_execute = DbWrapper.execute

        def record_player_queries(wrapper, statement, params=()):
            normalized = " ".join(str(statement).split())
            if re.search(r"\b(?:FROM|JOIN)\s+players\b", normalized, re.IGNORECASE):
                statements.append(normalized)
            return original_execute(wrapper, statement, params)

        with app.app_context():
            db = get_db()
            cursor = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,created_by) VALUES(?,?,?)",
                ("2026-08-12", "QUARTA", self.user_id),
            )
            sumula_id = cursor.lastrowid
            db.commit()
            with patch.object(DbWrapper, "execute", new=record_player_queries):
                self.assertIsNotNone(_sumula(db, sumula_id))

        with patch.object(DbWrapper, "execute", new=record_player_queries), patch("src.db.init_sqlite"):
            self.client.post("/login", data={"username": "usuario-inexistente", "password": "invalida"})
            with self.client.session_transaction() as session:
                session["user_id"] = self.user_id
            for endpoint in ("/players", "/sale", "/finance", "/futebol"):
                response = self.client.get(endpoint)
                self.assertNotEqual(response.status_code, 500, endpoint)

        self.assertTrue(statements)
        offending = [statement for statement in statements if re.search(r"\bphoto_data\b", statement, re.IGNORECASE)]
        self.assertEqual(offending, [])
        self.assertTrue(any("FROM football_participants" in statement for statement in statements))

    def test_password_hash_is_compatible_with_local_python(self):
        password_hash = make_password_hash("senha-segura-123")
        self.assertTrue(password_hash.startswith("pbkdf2:sha256:"))
        self.assertTrue(check_password_hash(password_hash, "senha-segura-123"))

    def test_manager_can_edit_user_display_name_and_username(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        response = self.client.post(
            f"/users/{self.user_id}/edit",
            data={"name": "Ana", "username": "ana.staff"},
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            user = get_db().execute("SELECT * FROM users WHERE id=?", (self.user_id,)).fetchone()
            self.assertEqual((user["name"], user["username"], user["role"]), ("Ana", "ana.staff", "manager"))

    def test_manager_can_assign_profile_after_user_creation(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        created = self.client.post(
            "/users",
            data={
                "name": "Equipe",
                "username": "equipe.teste",
                "role": "staff",
                "password": "senha-staff-123",
            },
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            target = get_db().execute("SELECT * FROM users WHERE username=?", ("equipe.teste",)).fetchone()
        changed = self.client.post(
            f"/users/{target['id']}/edit",
            data={"name": "Equipe", "username": "equipe.teste", "role": "infra"},
        )
        self.assertEqual(changed.status_code, 302)
        with app.app_context():
            target = get_db().execute("SELECT role,password_required FROM users WHERE id=?", (target["id"],)).fetchone()
            self.assertEqual((target["role"], target["password_required"]), ("infra", 1))

    def test_infra_user_sees_and_accesses_only_infra(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        created = self.client.post(
            "/users",
            data={
                "name": "Equipe Infra",
                "username": "infra.teste",
                "role": "infra",
                "password": "senha-infra-123",
            },
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            infra_user = get_db().execute(
                "SELECT * FROM users WHERE username=?", ("infra.teste",)
            ).fetchone()
            self.assertEqual(infra_user["role"], "infra")

        self.client.post("/logout")
        login = self.client.post(
            "/login", data={"username": "infra.teste", "password": "senha-infra-123"}
        )
        self.assertEqual(login.status_code, 303)
        self.assertTrue(login.headers["Location"].endswith("/infra/maintenance"))

        page = self.client.get("/infra/load-relation")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("<span>Infraestrutura</span>", html)
        self.assertIn(">Manutenção</a>", html)
        self.assertIn('class="sidebar-module sidebar-direct urgent ', html)
        self.assertIn("<span>Urgente</span>", html)
        self.assertIn('id="pwa-install"', html)
        for hidden_module in ("Bar", "Financeiro", "Administração"):
            self.assertNotIn(f"<span>{hidden_module}</span>", html)
        for hidden_link in ("Caixa", "Conferir Pix", "Estoque", "Produtos", "Pedidos", "Peladeiros", "Relatórios", "Usuários", "Venda rápida", "Compra rápida"):
            self.assertNotIn(f">{hidden_link}</a>", html)
        self.assertEqual(self.client.get("/infra/materials").status_code, 200)
        self.assertEqual(self.client.get("/infra/maintenance").status_code, 200)
        self.assertEqual(self.client.get("/urgent").status_code, 200)

        for forbidden_path in ("/", "/sale", "/stock", "/cash", "/players", "/users"):
            denied = self.client.get(forbidden_path)
            self.assertEqual(denied.status_code, 302)
            self.assertTrue(denied.headers["Location"].endswith("/infra/maintenance"))

        self.client.post("/logout")
        protected = self.client.get("/infra/materials")
        self.assertEqual(protected.status_code, 302)
        self.assertIn("next=/infra/materials", protected.headers["Location"])
        resumed = self.client.post(
            "/login",
            data={
                "username": "infra.teste", "password": "senha-infra-123",
                "next": "/infra/materials",
            },
        )
        self.assertTrue(resumed.headers["Location"].endswith("/infra/materials"))

    def test_reminders_calculate_debt_render_and_prevent_duplicate_email(self):
        sent_messages = []

        def fake_send(sender, password, recipient, subject, body):
            sent_messages.append((sender, recipient, subject, body))

        with app.app_context():
            db = get_db()
            settings = get_reminder_settings(db)
            debtors = outstanding_players(db, date(2026, 7, 5))
            self.assertEqual(debtors[0]["amount_cents"], 10500)
            first = dispatch_reminders(
                db, settings, "diretoriagpcta@gmail.com", "test", date(2026, 7, 5), fake_send
            )
            second = dispatch_reminders(
                db, settings, "diretoriagpcta@gmail.com", "test", date(2026, 7, 5), fake_send
            )
            self.assertEqual(first, {"sent": 1, "failed": 0, "skipped": 0, "without_email": 0})
            self.assertEqual(second, {"sent": 0, "failed": 0, "skipped": 1, "without_email": 0})
            self.assertEqual(len(sent_messages), 1)
            self.assertIn("Peladeiro", sent_messages[0][3])
            self.assertIn("R$ 105,00", sent_messages[0][3])

        with patch("src.services.email_reminders.smtplib.SMTP_SSL") as smtp_ssl:
            send_gmail("diretoriagpcta@gmail.com", "test", "teste@example.com", "Pendência", "Olá **Peladeiro**")
        message = smtp_ssl.return_value.__enter__.return_value.send_message.call_args.args[0]
        html_part = message.get_payload()[1].get_payload(decode=True).decode("utf-8")
        self.assertIn("PELADEIROS GPCTA", html_part)
        self.assertIn("Lembrete de pendência financeira", html_part)

    def test_manager_edits_reminder_and_cron_requires_secret(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/finance/reminders")
        self.assertEqual(page.status_code, 200)
        with patch("src.routes.finance.send_gmail") as send_test:
            test_email = self.client.post(
                "/finance/reminders/send-test", data={"test_email": "teste@example.com"}
            )
        self.assertEqual(test_email.status_code, 302)
        send_test.assert_called_once()
        self.assertEqual(send_test.call_args.args[2], "teste@example.com")
        self.assertIn("Peladeiro de teste", send_test.call_args.args[4])
        with patch("src.routes.finance.send_gmail") as invalid_send:
            invalid_email = self.client.post(
                "/finance/reminders/send-test", data={"test_email": "endereco-invalido"}
            )
        self.assertEqual(invalid_email.status_code, 302)
        invalid_send.assert_not_called()
        response = self.client.post(
            "/finance/reminders/settings",
            data={
                "enabled": "1",
                # O agendamento mensal aceita dias de 1 a 28 para funcionar
                # inclusive em fevereiro; o teste deve continuar válido nos
                # dias 29, 30 e 31 do mês corrente.
                "schedule_day": str(min(local_today().day, 28)),
                "subject": "Cobrança para {{ nome }}",
                "body": "Total: {{ total }}",
            },
        )
        self.assertEqual(response.status_code, 302)
        unauthorized = self.client.get("/cron/payment-reminders")
        self.assertEqual(unauthorized.status_code, 401)
        with patch("src.routes.finance.dispatch_reminders", return_value={
            "sent": 1, "failed": 0, "skipped": 0, "without_email": 0,
        }) as dispatch_mock:
            cron_day = local_today().replace(day=min(local_today().day, 28))
            with patch("src.routes.finance.local_today", return_value=cron_day):
                authorized = self.client.get(
                    "/cron/payment-reminders", headers={"Authorization": "Bearer cron-secret-test"}
                )
        self.assertEqual(authorized.status_code, 200)
        dispatch_mock.assert_called_once()

    def test_manager_downloads_debtors_pdf(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        response = self.client.get("/finance/reminders/debtors.pdf")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertTrue(response.data.startswith(b"%PDF-"))
        self.assertIn("attachment", response.headers["Content-Disposition"])

    def test_weekly_tribute_cron_requires_secret_and_runs_on_scheduled_days(self):
        unauthorized = self.client.get("/cron/weekly-tribute")
        self.assertEqual(unauthorized.status_code, 401)

        with patch("src.routes.finance.datetime") as clock, patch(
            "src.routes.finance.send_weekly_tribute_notifications", return_value=3
        ) as send_mock:
            clock.now.return_value = datetime(2026, 8, 5, 17, 0)
            response = self.client.get(
                "/cron/weekly-tribute",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sent"], 3)
        send_mock.assert_called_once()

        with patch("src.routes.finance.datetime") as clock, patch(
            "src.routes.finance.send_weekly_tribute_notifications", return_value=4
        ) as send_mock:
            clock.now.return_value = datetime(2026, 8, 8, 15, 7)
            response = self.client.get(
                "/cron/weekly-tribute",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sent"], 4)
        send_mock.assert_called_once()

        with patch("src.routes.finance.datetime") as clock, patch(
            "src.routes.finance.send_weekly_tribute_notifications"
        ) as send_mock:
            clock.now.return_value = datetime(2026, 8, 8, 14, 7)
            response = self.client.get(
                "/cron/weekly-tribute",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sent"], 0)
        send_mock.assert_not_called()

        with patch("src.routes.finance.datetime") as clock, patch(
            "src.routes.finance.send_weekly_tribute_notifications"
        ) as send_mock:
            clock.now.return_value = datetime(2026, 8, 3, 17, 0)
            response = self.client.get(
                "/cron/weekly-tribute",
                headers={"Authorization": "Bearer cron-secret-test"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sent"], 0)
        send_mock.assert_not_called()

    def test_weekly_tribute_workflow_uses_sao_paulo_schedule(self):
        workflow = (Path(__file__).parents[1] / ".github/workflows/weekly-tribute.yml").read_text()
        self.assertIn('cron: "7 17 * * 3"', workflow)
        self.assertIn('cron: "7 15 * * 6"', workflow)
        self.assertEqual(workflow.count('timezone: "America/Sao_Paulo"'), 2)
        self.assertNotIn('cron: "0 * * * *"', workflow)

    def test_weekly_tribute_accepts_dedicated_hashed_secret(self):
        dedicated_secret = "segredo-exclusivo-da-homenagem"
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO app_settings(key,value) VALUES(?,?)",
                ("weekly_tribute_cron_secret_hash", hashlib.sha256(dedicated_secret.encode()).hexdigest()),
            )
            db.commit()
        with patch("src.routes.finance.datetime") as clock, patch(
            "src.routes.finance.send_weekly_tribute_notifications", return_value=2
        ) as send_mock:
            clock.now.return_value = datetime(2026, 8, 5, 17, 7)
            response = self.client.get(
                "/cron/weekly-tribute",
                headers={"Authorization": f"Bearer {dedicated_secret}"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sent"], 2)
        send_mock.assert_called_once()

    def test_manager_sends_tribute_test_to_only_one_active_player(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with patch("src.routes.football.send_player_push", return_value={"sent": 1, "skipped": 0}) as send_mock:
            response = self.client.post(
                "/futebol/notificacoes/testar-homenagem",
                data={"player_id": str(self.player_id)},
            )
        self.assertEqual(response.status_code, 302)
        send_mock.assert_called_once_with(
            unittest.mock.ANY,
            self.player_id,
            "PELADEIROS GPCTA",
            "🗣️ VEEENHAAAMMM...",
            "/notificacoes",
            "/futebol/notificacoes/homenagem/imagem",
            True,
            True,
            "🗣️ VEEENHAAAMMM...",
        )
        with app.app_context():
            saved = get_db().execute(
                "SELECT audience,sent_count FROM push_announcements ORDER BY id DESC LIMIT 1"
            ).fetchone()
            self.assertEqual(saved["audience"], f"player:{self.player_id}")
            self.assertEqual(saved["sent_count"], 1)

    def test_tribute_test_rejects_invalid_player(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with patch("src.routes.football.send_player_push") as send_mock:
            response = self.client.post(
                "/futebol/notificacoes/testar-homenagem",
                data={"player_id": "999999"},
            )
        self.assertEqual(response.status_code, 302)
        send_mock.assert_not_called()

    def test_manager_configures_tribute_schedule_and_sanitizes_rich_text(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        response = self.client.post(
            "/futebol/notificacoes/homenagem/configuracao",
            data={
                "tribute_enabled": "1", "tribute_title": "Convocação",
                "tribute_body_html": "<strong>Venham</strong><script>alert(1)</script><big>agora</big>",
                "day_1": "1", "hour_1": "19",
            },
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            db = get_db(); settings = db.execute("SELECT * FROM tribute_settings WHERE id=1").fetchone()
            schedule = db.execute("SELECT * FROM tribute_schedules WHERE weekday=1").fetchone()
            self.assertEqual(settings["title"], "Convocação")
            self.assertEqual(settings["body"], "Venhamagora")
            self.assertNotIn("script", settings["body_html"])
            self.assertEqual((schedule["enabled"], schedule["hour"]), (1, 19))

    def test_tribute_image_is_in_push_payload_and_inbox(self):
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO push_subscriptions(player_id,endpoint,p256dh,auth) VALUES(?,?,?,?)",
                (self.player_id, "https://push.example/test", "key", "auth"),
            )
            db.commit()
            webpush_mock = unittest.mock.Mock()
            pywebpush_module = ModuleType("pywebpush")
            pywebpush_module.webpush = webpush_mock
            with patch.dict(os.environ, {"VAPID_PRIVATE_KEY": "test-key"}), patch.dict(
                sys.modules, {"pywebpush": pywebpush_module}
            ):
                result = send_player_push(
                    db,
                    self.player_id,
                    "Teste da homenagem",
                    "🗣️ VEEENHAAAMMM...",
                    "/notificacoes",
                    "/static/images/veeenhaaammm.png",
                    True,
                    True,
                )
            self.assertEqual(result["sent"], 1)
            payload = json.loads(webpush_mock.call_args.kwargs["data"])
            self.assertEqual(payload["web_push"], 8030)
            self.assertEqual(payload["notification"]["title"], "Teste da homenagem")
            self.assertEqual(payload["notification"]["navigate"], "/notificacoes")
            self.assertIs(payload["notification"]["silent"], False)
            self.assertEqual(
                payload["notification"]["image"],
                "/static/images/veeenhaaammm.png",
            )
            inbox = db.execute(
                "SELECT image_url FROM push_inbox WHERE player_id=? ORDER BY id DESC LIMIT 1",
                (self.player_id,),
            ).fetchone()
            self.assertEqual(inbox["image_url"], "/static/images/veeenhaaammm.png")
            subscription = db.execute(
                "SELECT last_push_status,last_push_at FROM push_subscriptions WHERE player_id=?",
                (self.player_id,),
            ).fetchone()
            self.assertEqual(subscription["last_push_status"], "accepted")
            self.assertIsNotNone(subscription["last_push_at"])

    def test_stock_page_groups_intelligent_alerts(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            db.execute(
                "UPDATE products SET stock=1,min_stock=2,expiry_date=? WHERE id=?",
                ((local_today() + timedelta(days=5)).isoformat(), self.product_id),
            )
            db.commit()
        page = self.client.get("/stock").get_data(as_text=True)
        self.assertIn("Alertas inteligentes", page)
        self.assertIn("Estoque baixo", page)
        self.assertIn("Próximos do vencimento", page)
        self.assertIn("5 dia(s)", page)

    def test_manager_downloads_monthly_sales_accountability_pdf(self):
        month = local_today().strftime("%Y-%m")
        with app.app_context():
            db = get_db()
            for method, total, paid_time, quantity in (
                ("Dinheiro", 600, f"{month}-10 15:00:00", 2),
                ("Pix", 300, f"{month}-11 15:00:00", 1),
                ("Pix", 300, f"{month}-12 15:00:00", 1),
                ("Cortesia", 300, f"{month}-13 15:00:00", 1),
            ):
                sale = db.execute(
                    """INSERT INTO sales(player_id,payment_method,total_cents,paid,paid_at)
                    VALUES(?,?,?,?,?)""",
                    (self.player_id, method, total, 1, paid_time),
                )
                db.execute(
                    """INSERT INTO sale_items
                    (sale_id,product_id,quantity,unit_price_cents,unit_cost_cents)
                    VALUES(?,?,?,?,?)""",
                    (sale.lastrowid, self.product_id, quantity, 300, 100),
                )
            db.commit()
            data = monthly_sales_data(db, month)
            self.assertEqual(
                (data["summary"]["revenue"], data["summary"]["sales_count"], data["summary"]["items_sold"], data["summary"]["profit"]),
                (1200, 3, 4, 800),
            )
            self.assertEqual((data["most_used_payment"], data["summary"]["courtesy_items"]), ("Pix", 1))
            self.assertEqual(
                (data["consumers"][0]["name"], data["consumers"][0]["purchases"], data["consumers"][0]["items"], data["consumers"][0]["total"]),
                ("Peladeiro", 3, 4, 1200),
            )

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get(f"/reports?month={month}").get_data(as_text=True)
        self.assertIn("PDF de vendas mensais", page)
        report = self.client.get(f"/reports/monthly-sales.pdf?month={month}")
        self.assertEqual(report.status_code, 200)
        self.assertEqual(report.mimetype, "application/pdf")
        self.assertTrue(report.data.startswith(b"%PDF-"))
        self.assertIn(f"vendas-mensais-{month}.pdf", report.headers["Content-Disposition"])

    def test_legacy_pix_remains_available_until_credentials_are_configured(self):
        app.config.update(
            MERCADOPAGO_ACCESS_TOKEN=None,
            MERCADOPAGO_POS_ID=None,
            MERCADOPAGO_WEBHOOK_SECRET=None,
        )
        response = self.client.get(
            "/pix/qrcode?amount_cents=300",
            headers={"Accept": "application/json", "X-Pix-Token": self.token},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["image"].startswith("data:image/png;base64,"))

    def test_webhook_approves_order_and_api_failure_restores_stock(self):
        sale_id = self.create_order("ORD-WEBHOOK", 1)
        data_id = "ORD-WEBHOOK"
        request_id = "request-webhook"
        timestamp = "1742505638683"
        template = f"id:{data_id.lower()};request-id:{request_id};ts:{timestamp};"
        signature = hmac.new(b"webhook-secret", template.encode(), hashlib.sha256).hexdigest()
        approved = {
            "id": data_id,
            "external_reference": None,
            "status": "processed",
            "status_detail": "accredited",
            "total_paid_amount": "3.00",
            "transactions": {"payments": [{"id": "PAY-WEBHOOK"}]},
        }
        with patch("src.routes.sales.get_order") as get_order_mock:
            response = self.client.post(
                f"/webhooks/mercadopago?data.id={data_id}&type=order",
                headers={"X-Request-Id": request_id, "X-Signature": f"ts={timestamp},v1={signature}"},
                json={"type": "order", "data": approved},
            )
        self.assertEqual(response.status_code, 200)
        get_order_mock.assert_not_called()
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertEqual((sale["paid"], sale["payment_status"]), (1, "approved"))

        with patch("src.routes.sales.create_pix_order", side_effect=MercadoPagoError("falha simulada")):
            failed = self.client.post(
                "/pix/mercadopago/orders",
                headers=self.headers(),
                json={
                    "player_id": self.player_id,
                    "items": [{"product_id": self.product_id, "quantity": 1}],
                },
            )
        self.assertEqual(failed.status_code, 502)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 4)

    def test_partial_credit_pix_charges_remainder_and_consumes_once_on_approval(self):
        from src.routes.sales import apply_mercadopago_status

        with app.app_context():
            db = get_db()
            db.execute("UPDATE products SET price_cents=1500 WHERE id=?", (self.product_id,))
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,400)",
                (self.player_id,),
            )
            db.commit()
        order_response = {
            "id": "ORD-PARTIAL-CREDIT",
            "status": "action_required",
            "transactions": {"payments": [{
                "id": "PAY-PARTIAL-CREDIT",
                "payment_method": {"qr_code": "000201PARTIAL"},
            }]},
        }
        with patch("src.routes.sales.create_pix_order", return_value=order_response) as create_order:
            created = self.client.post(
                "/pix/mercadopago/orders", headers=self.headers(),
                json={
                    "player_id": self.player_id,
                    "use_bar_credit": True,
                    "items": [{"product_id": self.product_id, "quantity": 1}],
                },
            )
        self.assertEqual(created.status_code, 201, created.get_json())
        sale_id = created.get_json()["sale_id"]
        self.assertEqual(create_order.call_args.args[2], 1100)
        self.assertEqual(created.get_json()["amount"], "R$ 11,00")
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            reservation = db.execute(
                "SELECT * FROM bar_credit_reservations WHERE sale_id=?", (sale_id,)
            ).fetchone()
            self.assertEqual((sale["total_cents"], sale["paid"]), (1500, 0))
            self.assertEqual((reservation["amount_cents"], reservation["status"]), (400, "reserved"))
            self.assertEqual(db.execute(
                "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
            ).fetchone()[0], 400)
            approved_order = {
                "status": "processed", "status_detail": "accredited",
                "total_paid_amount": "11.00",
                "transactions": {"payments": [{"id": "PAY-APPROVED"}]},
            }
            self.assertEqual(apply_mercadopago_status(db, sale, approved_order), "approved")
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertEqual(apply_mercadopago_status(db, sale, approved_order), "approved")
            self.assertEqual(db.execute(
                "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
            ).fetchone()[0], 0)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM bar_credit_transactions WHERE sale_id=? AND type='CONSUMPTION'",
                (sale_id,),
            ).fetchone()[0], 1)
            self.assertEqual(db.execute(
                "SELECT status FROM bar_credit_reservations WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], "consumed")
            self.assertEqual((sale["paid"], sale["payment_status"], sale["ready_for_delivery"]), (1, "approved", 1))

    def test_partial_credit_pix_terminal_and_creation_failure_release_reservation(self):
        from src.routes.sales import apply_mercadopago_status

        with app.app_context():
            db = get_db()
            db.execute("UPDATE products SET price_cents=1500,stock=5 WHERE id=?", (self.product_id,))
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,400)",
                (self.player_id,),
            )
            db.commit()

        def create_pending(order_id):
            response_data = {
                "id": order_id,
                "transactions": {"payments": [{
                    "id": f"PAY-{order_id}", "payment_method": {"qr_code": "000201TEST"},
                }]},
            }
            with patch("src.routes.sales.create_pix_order", return_value=response_data):
                response = self.client.post(
                    "/pix/mercadopago/orders", headers=self.headers(),
                    json={"player_id": self.player_id, "use_bar_credit": True,
                          "items": [{"product_id": self.product_id, "quantity": 1}]},
                )
            return response.get_json()["sale_id"]

        for terminal_status in ("expired", "canceled"):
            sale_id = create_pending(f"ORD-{terminal_status}")
            with app.app_context():
                db = get_db()
                sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
                self.assertEqual(apply_mercadopago_status(
                    db, sale, {"status": terminal_status, "transactions": {"payments": []}}
                ), terminal_status)
                self.assertEqual(db.execute(
                    "SELECT status FROM bar_credit_reservations WHERE sale_id=?", (sale_id,)
                ).fetchone()[0], "released")
                self.assertEqual(db.execute(
                    "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
                ).fetchone()[0], 400)

        with patch("src.routes.sales.create_pix_order", side_effect=MercadoPagoError("falha simulada")):
            failed = self.client.post(
                "/pix/mercadopago/orders", headers=self.headers(),
                json={"player_id": self.player_id, "use_bar_credit": True,
                      "items": [{"product_id": self.product_id, "quantity": 1}]},
            )
        self.assertEqual(failed.status_code, 502)
        with app.app_context():
            db = get_db()
            failed_sale = db.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            self.assertEqual(db.execute(
                "SELECT status FROM bar_credit_reservations WHERE sale_id=?", (failed_sale["id"],)
            ).fetchone()[0], "released")

    def test_pix_selection_with_full_credit_skips_mercadopago(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE products SET price_cents=300 WHERE id=?", (self.product_id,))
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,1000)",
                (self.player_id,),
            )
            db.commit()
        with patch("src.routes.sales.create_pix_order") as create_order:
            response = self.client.post(
                "/pix/mercadopago/orders", headers=self.headers(),
                json={"player_id": self.player_id, "use_bar_credit": True,
                      "items": [{"product_id": self.product_id, "quantity": 1}]},
            )
        self.assertEqual(response.status_code, 201)
        self.assertTrue(response.get_json()["paid"])
        create_order.assert_not_called()
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            self.assertEqual((sale["payment_method"], sale["paid"], sale["payment_status"]), ("Créditos", 1, "approved"))
            self.assertIsNone(db.execute(
                "SELECT id FROM bar_credit_reservations WHERE sale_id=?", (sale["id"],)
            ).fetchone())

    def test_webhook_simulator_acknowledges_unknown_order(self):
        data_id = "123456"
        request_id = "request-simulator"
        timestamp = "1742505638683"
        template = f"id:{data_id};request-id:{request_id};ts:{timestamp};"
        signature = hmac.new(b"webhook-secret", template.encode(), hashlib.sha256).hexdigest()
        payload = {
            "action": "order.processed",
            "type": "order",
            "data": {
                "id": data_id,
                "external_reference": "ext_ref_1234",
                "status": "processed",
                "status_detail": "accredited",
                "total_paid_amount": 100000,
                "type": "point",
            },
        }
        with (
            patch("src.routes.sales.get_order") as get_order_mock,
            patch("src.routes.sales.get_db") as get_db_mock,
        ):
            response = self.client.post(
                f"/webhooks/mercadopago?data.id={data_id}&type=order",
                headers={"X-Request-Id": request_id, "X-Signature": f"ts={timestamp},v1={signature}"},
                json=payload,
            )
        self.assertEqual(response.status_code, 200)
        get_order_mock.assert_not_called()
        get_db_mock.assert_not_called()

    def test_paid_pix_enters_delivery_queue_and_staff_confirms_it(self):
        sale_id = self.create_order("ORD-DELIVERY", 2)
        data_id = "ORD-DELIVERY"
        request_id = "request-delivery"
        timestamp = "1742505638683"
        template = f"id:{data_id.lower()};request-id:{request_id};ts:{timestamp};"
        signature = hmac.new(b"webhook-secret", template.encode(), hashlib.sha256).hexdigest()
        approved = {
            "id": data_id,
            "type": "online",
            "status": "processed",
            "status_detail": "accredited",
            "total_paid_amount": "6.00",
            "transactions": {"payments": [{"id": "PAY-DELIVERY"}]},
        }
        response = self.client.post(
            f"/webhooks/mercadopago?data.id={data_id}&type=order",
            headers={"X-Request-Id": request_id, "X-Signature": f"ts={timestamp},v1={signature}"},
            json={"type": "order", "data": approved},
        )
        self.assertEqual(response.status_code, 200)

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        page = self.client.get("/orders")
        self.assertEqual(page.status_code, 200)
        self.assertIn("Pedidos para entregar", page.get_data(as_text=True))
        feed = self.client.get("/orders/feed", headers={"Accept": "application/json"})
        self.assertEqual(feed.status_code, 200)
        pending = feed.get_json()["pending"]
        self.assertEqual(len(pending), 1)
        self.assertEqual((pending[0]["id"], pending[0]["items"][0]["quantity"]), (sale_id, 2))

        delivered = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(delivered.status_code, 200)
        feed = self.client.get("/orders/feed", headers={"Accept": "application/json"}).get_json()
        self.assertEqual(feed["pending"], [])
        self.assertEqual(feed["delivered"][0]["delivered_by_name"], "Teste")

    def test_pending_order_queues_show_newest_orders_first(self):
        with app.app_context():
            db = get_db()
            sale_ids = []
            for created_at in (
                "2026-08-12 10:00:00",
                "2026-08-12 11:00:00",
                "2026-08-12 12:00:00",
            ):
                sale_id = db.execute(
                    """INSERT INTO sales(
                           player_id,payment_method,total_cents,paid,payment_status,
                           ready_for_delivery,created_at,paid_at
                       ) VALUES(?, 'Pix', 300, 1, 'approved', 1, ?, ?)""",
                    (self.player_id, created_at, created_at),
                ).lastrowid
                db.execute(
                    """INSERT INTO sale_items(
                           sale_id,product_id,quantity,unit_price_cents,unit_cost_cents
                       ) VALUES(?,?,?,?,?)""",
                    (sale_id, self.product_id, 1, 300, 100),
                )
                sale_ids.append(sale_id)
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        expected = list(reversed(sale_ids))

        orders_feed = self.client.get("/orders/feed").get_json()["pending"]
        self.assertEqual([order["id"] for order in orders_feed], expected)

        panel_feed = self.client.get("/painel/feed").get_json()["orders"]
        self.assertEqual([order["id"] for order in panel_feed], expected)

        pending_page = self.client.get("/orders/pending-delivery").get_data(as_text=True)
        positions = [pending_page.index(f"Pedido #{sale_id}") for sale_id in expected]
        self.assertEqual(positions, sorted(positions))

        with patch("src.routes.sales.build_pending_delivery_pdf", return_value=BytesIO(b"%PDF-1.4\n%%EOF")) as build_pdf:
            pdf = self.client.get("/orders/pending-delivery.pdf")
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual([order["id"] for order in build_pdf.call_args.args[0]], expected)

    def test_manager_restores_accidental_full_delivery_without_changing_payment_or_stock(self):
        sale_id = self.create_order("ORD-RESTORE-DELIVERY", 2)
        with app.app_context():
            db = get_db()
            db.execute(
                "UPDATE sales SET paid=1,payment_status='approved',ready_for_delivery=1,paid_at=CURRENT_TIMESTAMP WHERE id=?",
                (sale_id,),
            )
            db.commit()
            stock_before = db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"]

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        delivered = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(delivered.status_code, 200)

        with app.app_context():
            db = get_db()
            staff = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'staff')",
                ("atendente.restore", "Atendente Restore", "hash"),
            ).lastrowid
            db.execute(
                "CREATE TRIGGER fail_delivery_restore BEFORE UPDATE OF delivered_at ON sales "
                "WHEN OLD.id=%d BEGIN SELECT RAISE(ABORT, 'falha simulada'); END" % sale_id
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = staff
        denied = self.client.post(
            f"/orders/{sale_id}/restore-delivery",
            json={"reason": "Correção solicitada."},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(denied.status_code, 403)

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        failed = self.client.post(
            f"/orders/{sale_id}/restore-delivery",
            json={"reason": "Falha transacional simulada."},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(failed.status_code, 500)
        with app.app_context():
            db = get_db()
            self.assertIsNotNone(db.execute("SELECT delivered_at FROM sales WHERE id=?", (sale_id,)).fetchone()["delivered_at"])
            self.assertGreater(
                db.execute(
                    "SELECT COUNT(*) total FROM sale_item_deliveries sid JOIN sale_items si ON si.id=sid.sale_item_id WHERE si.sale_id=?",
                    (sale_id,),
                ).fetchone()["total"],
                0,
            )
            db.execute("DROP TRIGGER fail_delivery_restore")
            db.commit()

        invalid = self.client.post(
            f"/orders/{sale_id}/restore-delivery",
            json={"reason": "ops"},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(invalid.status_code, 400)

        restored = self.client.post(
            f"/orders/{sale_id}/restore-delivery",
            json={"reason": "Entrega total registrada por engano."},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(restored.status_code, 200)
        with app.app_context():
            db = get_db()
            sale = db.execute(
                "SELECT paid,payment_status,ready_for_delivery,delivered_at,delivered_by FROM sales WHERE id=?",
                (sale_id,),
            ).fetchone()
            deliveries = db.execute(
                "SELECT COUNT(*) total FROM sale_item_deliveries sid JOIN sale_items si ON si.id=sid.sale_item_id WHERE si.sale_id=?",
                (sale_id,),
            ).fetchone()["total"]
            stock_after = db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"]
            self.assertEqual((sale["paid"], sale["payment_status"], sale["ready_for_delivery"]), (1, "approved", 1))
            self.assertIsNone(sale["delivered_at"])
            self.assertIsNone(sale["delivered_by"])
            self.assertEqual(deliveries, 0)
            self.assertEqual(stock_after, stock_before)
            sale_item_id = db.execute(
                "SELECT id FROM sale_items WHERE sale_id=? ORDER BY id LIMIT 1",
                (sale_id,),
            ).fetchone()["id"]

        partial = self.client.post(
            f"/orders/{sale_id}/deliver",
            json={
                "sale_item_id": sale_item_id,
                "quantity": 1,
            },
            headers={"Accept": "application/json"},
        )
        self.assertEqual(partial.status_code, 200)
        self.assertTrue(partial.get_json()["partial"])
        self.assertEqual(partial.get_json()["remaining_items"][0]["quantity"], 1)

        history = self.client.get("/orders/delivered").get_data(as_text=True)
        self.assertIn("Corrigir entrega", history)

    def test_cash_order_waits_for_staff_payment_delivery_or_cancel(self):
        with app.app_context():
            db = get_db()
            db.execute("UPDATE products SET stock=20 WHERE id=?", (self.product_id,))
            cursor = db.execute(
                "INSERT INTO users(username,name,password_hash,role,password_required) VALUES(?,?,?,'client',0)",
                ("peladeiro.caixa", "Peladeiro Caixa", "hash"),
            )
            client_id = cursor.lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = client_id

        created = self.client.post(
            "/sale",
            data={
                "player_id": str(self.player_id),
                "product_id": [str(self.product_id)],
                "quantity": ["10"],
                "payment_method": "Dinheiro",
                "notes": "Precisa de troco.",
            },
        )
        self.assertEqual(created.status_code, 303)
        with app.app_context():
            db = get_db()
            cash_sale = db.execute(
                "SELECT * FROM sales WHERE payment_method='Dinheiro' ORDER BY id DESC LIMIT 1"
            ).fetchone()
            sale_id = cash_sale["id"]
            sale_item_id = db.execute(
                "SELECT id FROM sale_items WHERE sale_id=?", (sale_id,)
            ).fetchone()["id"]
            self.assertEqual(
                (cash_sale["paid"], cash_sale["payment_status"], cash_sale["ready_for_delivery"]),
                (0, "pending_cash", 1),
            )
            self.assertEqual(
                db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"],
                10,
            )

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        orders_page = self.client.get("/orders").get_data(as_text=True)
        self.assertIn("Confirmar pagamento", orders_page)
        self.assertNotIn("Confirmar pagamento e entregar", orders_page)
        self.assertIn("Cancelar", orders_page)
        feed = self.client.get("/orders/feed", headers={"Accept": "application/json"}).get_json()
        self.assertEqual(len(feed["pending"]), 1)
        self.assertEqual(
            (feed["pending"][0]["id"], feed["pending"][0]["waiting_cash"], feed["pending"][0]["notes"]),
            (sale_id, True, "Precisa de troco."),
        )

        blocked_delivery = self.client.post(
            f"/orders/{sale_id}/deliver",
            json={"sale_item_id": sale_item_id, "quantity": 1},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(blocked_delivery.status_code, 409)

        confirmed = self.client.post(
            f"/orders/{sale_id}/confirm-payment",
            json={"amount_received_cents": 3000, "convert_change_to_credit": False},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(confirmed.status_code, 200)
        self.assertFalse(confirmed.get_json()["already_paid"])
        repeated_confirmation = self.client.post(
            f"/orders/{sale_id}/confirm-payment",
            json={"amount_received_cents": 3000, "convert_change_to_credit": False},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(repeated_confirmation.status_code, 200)
        self.assertTrue(repeated_confirmation.get_json()["already_paid"])
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertEqual((sale["paid"], sale["payment_status"]), (1, "approved"))
            self.assertIsNotNone(sale["paid_at"])
            self.assertIsNone(sale["delivered_at"])
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM sale_item_deliveries WHERE sale_item_id=?", (sale_item_id,)
            ).fetchone()[0], 0)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM sale_delivery_operations WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], 0)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM notification_outbox WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], 0)

        paid_feed = self.client.get("/orders/feed", headers={"Accept": "application/json"}).get_json()
        paid_order = next(order for order in paid_feed["pending"] if order["id"] == sale_id)
        self.assertFalse(paid_order["waiting_cash"])

        partial = self.client.post(
            f"/orders/{sale_id}/deliver",
            json={"sale_item_id": sale_item_id, "quantity": 1},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(partial.status_code, 200)
        self.assertTrue(partial.get_json()["partial"])
        self.assertEqual(partial.get_json()["remaining_items"][0]["quantity"], 9)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT SUM(quantity) FROM sale_item_deliveries WHERE sale_item_id=?", (sale_item_id,)
            ).fetchone()[0], 1)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM sale_delivery_operations WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], 1)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM notification_outbox WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], 3)

        delivered = self.client.post(f"/orders/{sale_id}/deliver", headers={"Accept": "application/json"})
        self.assertEqual(delivered.status_code, 200)
        with app.app_context():
            sale = get_db().execute("SELECT * FROM sales WHERE id=?", (sale_id,)).fetchone()
            self.assertIsNotNone(sale["delivered_at"])

        with self.client.session_transaction() as session:
            session["user_id"] = client_id
        self.client.post(
            "/sale",
            data={
                "player_id": str(self.player_id),
                "product_id": [str(self.product_id)],
                "quantity": ["1"],
                "payment_method": "Dinheiro",
                "notes": "Pedido a cancelar.",
            },
        )
        with app.app_context():
            db = get_db()
            canceled_id = db.execute("SELECT MAX(id) FROM sales").fetchone()[0]
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 9)
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        canceled = self.client.post(f"/orders/{canceled_id}/cancel", headers={"Accept": "application/json"})
        self.assertEqual(canceled.status_code, 200)
        repeated = self.client.post(f"/orders/{canceled_id}/cancel", headers={"Accept": "application/json"})
        self.assertEqual(repeated.status_code, 409)
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales WHERE id=?", (canceled_id,)).fetchone()
            self.assertEqual((sale["paid"], sale["payment_status"], sale["ready_for_delivery"]), (0, "canceled", 0))
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 10)

    def test_partial_credit_cash_reserves_and_consumes_only_on_confirmation(self):
        self.login_manager()
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,400)",
                (self.player_id,),
            )
            db.commit()

        created = self.client.post(
            "/sale",
            data={
                "sale_type": "player", "player_id": str(self.player_id),
                "payment_method": "Dinheiro", "use_bar_credit": "1",
                "product_id": [str(self.product_id)], "quantity": ["5"],
            },
        )
        self.assertEqual(created.status_code, 303)
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            sale_id = sale["id"]
            reservation = db.execute(
                "SELECT * FROM bar_credit_reservations WHERE sale_id=?", (sale_id,)
            ).fetchone()
            self.assertEqual((sale["total_cents"], sale["payment_method"], sale["payment_status"]), (1500, "Dinheiro", "pending_cash"))
            self.assertEqual((reservation["amount_cents"], reservation["status"]), (400, "reserved"))
            self.assertEqual(db.execute(
                "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
            ).fetchone()[0], 400)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM bar_credit_transactions WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], 0)

        feed_order = next(
            order for order in self.client.get("/orders/feed").get_json()["pending"]
            if order["id"] == sale_id
        )
        self.assertEqual(
            (feed_order["total_cents"], feed_order["credit_reserved_cents"], feed_order["cash_due_cents"]),
            (1500, 400, 1100),
        )
        confirmed = self.client.post(
            f"/orders/{sale_id}/confirm-payment",
            json={"amount_received_cents": 1100, "convert_change_to_credit": False},
        )
        self.assertEqual(confirmed.status_code, 200)
        self.assertEqual(
            (confirmed.get_json()["cash_due_cents"], confirmed.get_json()["change_cents"], confirmed.get_json()["credit_consumed_cents"]),
            (1100, 0, 400),
        )
        repeated = self.client.post(
            f"/orders/{sale_id}/confirm-payment",
            json={"amount_received_cents": 1100, "convert_change_to_credit": False},
        )
        self.assertTrue(repeated.get_json()["already_paid"])
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
            ).fetchone()[0], 0)
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM bar_credit_transactions WHERE sale_id=? AND type='CONSUMPTION'", (sale_id,)
            ).fetchone()[0], 1)
            self.assertEqual(db.execute(
                "SELECT status FROM bar_credit_reservations WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], "consumed")
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM sale_item_deliveries sid JOIN sale_items si ON si.id=sid.sale_item_id WHERE si.sale_id=?",
                (sale_id,),
            ).fetchone()[0], 0)

    def test_partial_credit_ui_has_safe_catalog_initialization_order(self):
        with app.app_context():
            db = get_db()
            client_id = db.execute(
                """INSERT INTO users(username,name,password_hash,role,player_id)
                   VALUES(?,?,?,'client',?)""",
                ("credito.parcial", "Crédito Parcial", "hash", self.player_id),
            ).lastrowid
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,2000)",
                (self.player_id,),
            )
            db.execute("UPDATE products SET price_cents=2400 WHERE id=?", (self.product_id,))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = client_id

        page = self.client.get("/sale")
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn('id="partial-credit-box"', html)
        self.assertIn('id="use-bar-credit"', html)
        self.assertIn("method.value==='Dinheiro'", html)
        self.assertIn("catalogMode==='bar'", html)
        self.assertIn("creditBalance>0", html)
        self.assertLess(
            html.index("let catalogMode='bar'"),
            html.index("renderPlayerSummary();toggleSaleTarget();total();"),
        )

    def test_partial_credit_cash_change_and_cancel_are_atomic(self):
        self.login_manager()
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,400)",
                (self.player_id,),
            )
            db.commit()

        form = {
            "sale_type": "player", "player_id": str(self.player_id),
            "payment_method": "Dinheiro", "use_bar_credit": "1",
            "product_id": [str(self.product_id)], "quantity": ["5"],
        }
        self.client.post("/sale", data=form)
        with app.app_context():
            sale_id = get_db().execute("SELECT MAX(id) FROM sales").fetchone()[0]
        paid = self.client.post(
            f"/orders/{sale_id}/confirm-payment",
            json={"amount_received_cents": 2000, "convert_change_to_credit": True},
        )
        self.assertEqual(paid.status_code, 200)
        self.assertEqual((paid.get_json()["change_cents"], paid.get_json()["balance_cents"]), (900, 900))
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT SUM(amount_cents) FROM bar_credit_transactions WHERE sale_id=?", (sale_id,)
            ).fetchone()[0], 500)

            db.execute("UPDATE products SET stock=5 WHERE id=?", (self.product_id,))
            db.execute("UPDATE bar_credit_accounts SET balance_cents=400 WHERE player_id=?", (self.player_id,))
            db.commit()
        self.client.post("/sale", data=form)
        with app.app_context():
            canceled_id = get_db().execute("SELECT MAX(id) FROM sales").fetchone()[0]
        canceled = self.client.post(
            f"/orders/{canceled_id}/cancel", json={"reason": "Cliente desistiu"}
        )
        self.assertEqual(canceled.status_code, 200)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT status FROM bar_credit_reservations WHERE sale_id=?", (canceled_id,)
            ).fetchone()[0], "released")
            self.assertEqual(db.execute(
                "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
            ).fetchone()[0], 400)

    def test_partial_credit_backend_recalculates_and_full_credit_stays_immediate(self):
        self.login_manager()
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO bar_credit_accounts(player_id,balance_cents) VALUES(?,250)",
                (self.player_id,),
            )
            db.commit()
        self.client.post(
            "/sale", data={
                "sale_type": "player", "player_id": str(self.player_id),
                "payment_method": "Dinheiro", "use_bar_credit": "1",
                "product_id": [str(self.product_id)], "quantity": ["1"],
                "credit_amount": "999999",
            },
        )
        with app.app_context():
            db = get_db()
            sale = db.execute("SELECT * FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            self.assertEqual(db.execute(
                "SELECT amount_cents FROM bar_credit_reservations WHERE sale_id=?", (sale["id"],)
            ).fetchone()[0], 250)
            db.execute("UPDATE products SET stock=5 WHERE id=?", (self.product_id,))
            db.execute("UPDATE bar_credit_accounts SET balance_cents=1000 WHERE player_id=?", (self.player_id,))
            db.commit()
        self.client.post(
            "/sale", data={
                "sale_type": "player", "player_id": str(self.player_id),
                "payment_method": "Dinheiro", "use_bar_credit": "1",
                "product_id": [str(self.product_id)], "quantity": ["1"],
            },
        )
        with app.app_context():
            db = get_db()
            full_credit_sale = db.execute("SELECT * FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            self.assertEqual(
                (full_credit_sale["payment_method"], full_credit_sale["paid"], full_credit_sale["payment_status"]),
                ("Créditos", 1, "approved"),
            )
            self.assertIsNone(db.execute(
                "SELECT id FROM bar_credit_reservations WHERE sale_id=?", (full_credit_sale["id"],)
            ).fetchone())

    def test_cash_change_uses_credit_wallet_atomically_and_idempotently(self):
        def create_cash_sale(player_id=self.player_id, guest_name=None):
            with app.app_context():
                db = get_db()
                sale_id = db.execute(
                    """INSERT INTO sales
                       (player_id,guest_name,payment_method,total_cents,paid,payment_status,ready_for_delivery)
                       VALUES(?,?,'Dinheiro',600,0,'pending_cash',1)""",
                    (player_id, guest_name or ""),
                ).lastrowid
                db.commit()
                return sale_id

        with app.app_context():
            db = get_db()
            db.execute("DELETE FROM bar_credit_audit WHERE player_id=?", (self.player_id,))
            db.execute("DELETE FROM bar_credit_transactions WHERE player_id=?", (self.player_id,))
            db.execute("DELETE FROM bar_credit_accounts WHERE player_id=?", (self.player_id,))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        returned_sale_id = create_cash_sale()
        returned = self.client.post(
            f"/orders/{returned_sale_id}/confirm-payment",
            json={"amount_received_cents": 1000, "convert_change_to_credit": False},
        )
        self.assertEqual(returned.status_code, 200)
        self.assertEqual(returned.get_json()["change_cents"], 400)
        self.assertFalse(returned.get_json()["credited"])
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute(
                "SELECT COALESCE(SUM(amount_cents),0) FROM bar_credit_transactions WHERE player_id=?",
                (self.player_id,),
            ).fetchone()[0], 0)

        converted_sale_id = create_cash_sale()
        converted = self.client.post(
            f"/orders/{converted_sale_id}/confirm-payment",
            json={"amount_received_cents": 1000, "convert_change_to_credit": True},
        )
        self.assertEqual(converted.status_code, 200)
        self.assertEqual(
            (converted.get_json()["change_cents"], converted.get_json()["credited"], converted.get_json()["balance_cents"]),
            (400, True, 400),
        )
        repeated = self.client.post(
            f"/orders/{converted_sale_id}/confirm-payment",
            json={"amount_received_cents": 1000, "convert_change_to_credit": True},
        )
        self.assertEqual(repeated.status_code, 200)
        self.assertTrue(repeated.get_json()["already_paid"])
        with app.app_context():
            db = get_db()
            account = db.execute(
                "SELECT balance_cents FROM bar_credit_accounts WHERE player_id=?", (self.player_id,)
            ).fetchone()
            transactions = db.execute(
                """SELECT type,amount_cents,balance_after_cents,description,sale_id,created_by
                   FROM bar_credit_transactions WHERE player_id=?""",
                (self.player_id,),
            ).fetchall()
            audit = db.execute(
                """SELECT action,amount_cents,actor_user_id,reason
                   FROM bar_credit_audit WHERE player_id=?""",
                (self.player_id,),
            ).fetchone()
            self.assertEqual(account["balance_cents"], 400)
            self.assertEqual(len(transactions), 1)
            self.assertEqual(
                tuple(transactions[0]),
                ("ADJUSTMENT", 400, 400, "Troco convertido em crédito", converted_sale_id, self.user_id),
            )
            self.assertEqual(
                tuple(audit), ("TROCO_CONVERTIDO", 400, self.user_id, f"Pedido #{converted_sale_id}")
            )

        exact_sale_id = create_cash_sale()
        exact = self.client.post(
            f"/orders/{exact_sale_id}/confirm-payment",
            json={"amount_received_cents": 600, "convert_change_to_credit": True},
        )
        self.assertEqual(exact.status_code, 200)
        self.assertEqual((exact.get_json()["change_cents"], exact.get_json()["credited"]), (0, False))

        insufficient_sale_id = create_cash_sale()
        missing_amount = self.client.post(
            f"/orders/{insufficient_sale_id}/confirm-payment", json={}
        )
        self.assertEqual(missing_amount.status_code, 400)
        insufficient = self.client.post(
            f"/orders/{insufficient_sale_id}/confirm-payment",
            json={"amount_received_cents": 599, "convert_change_to_credit": False},
        )
        self.assertEqual(insufficient.status_code, 400)
        with app.app_context():
            sale = get_db().execute("SELECT paid,payment_status FROM sales WHERE id=?", (insufficient_sale_id,)).fetchone()
            self.assertEqual((sale["paid"], sale["payment_status"]), (0, "pending_cash"))

        guest_sale_id = create_cash_sale(player_id=None, guest_name="Convidado Evento")
        guest_conversion = self.client.post(
            f"/orders/{guest_sale_id}/confirm-payment",
            json={"amount_received_cents": 1000, "convert_change_to_credit": True},
        )
        self.assertEqual(guest_conversion.status_code, 400)
        guest_returned = self.client.post(
            f"/orders/{guest_sale_id}/confirm-payment",
            json={"amount_received_cents": 1000, "convert_change_to_credit": False},
        )
        self.assertEqual(guest_returned.status_code, 200)

        failed_sale_id = create_cash_sale()
        with patch("src.routes.sales.credit_cash_change", side_effect=RuntimeError("wallet unavailable")):
            failed = self.client.post(
                f"/orders/{failed_sale_id}/confirm-payment",
                json={"amount_received_cents": 1000, "convert_change_to_credit": True},
            )
        self.assertEqual(failed.status_code, 500)
        with app.app_context():
            db = get_db()
            failed_sale = db.execute(
                "SELECT paid,payment_status,paid_at FROM sales WHERE id=?", (failed_sale_id,)
            ).fetchone()
            self.assertEqual((failed_sale["paid"], failed_sale["payment_status"], failed_sale["paid_at"]), (0, "pending_cash", None))
            self.assertEqual(db.execute(
                "SELECT COUNT(*) FROM bar_credit_transactions WHERE sale_id=?", (failed_sale_id,)
            ).fetchone()[0], 0)

    def test_cash_register_reconciles_sales_movements_reversal_and_closing(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        opened = self.client.post(
            "/cash/open", data={"opening_cash": "100,00", "opening_bank": "500,00"}
        )
        self.assertEqual(opened.status_code, 303)
        today = local_today().isoformat()
        with app.app_context():
            db = get_db()
            db.execute(
                """INSERT INTO sales(player_id,payment_method,total_cents,paid,paid_at)
                VALUES(?, 'Dinheiro',300,1,?)""",
                (self.player_id, f"{today} 15:00:00"),
            )
            db.execute(
                """INSERT INTO sales(player_id,payment_method,total_cents,paid,paid_at)
                VALUES(?, 'Pix',500,1,?)""",
                (self.player_id, f"{today} 15:01:00"),
            )
            db.execute(
                """INSERT INTO sales(player_id,payment_method,total_cents,paid,paid_at)
                VALUES(?, 'Cortesia',900,1,?)""",
                (self.player_id, f"{today} 15:02:00"),
            )
            db.commit()

        movement_response = self.client.post(
            "/cash/movements",
            data={
                "account": "cash",
                "direction": "out",
                "category": "expense",
                "amount": "2,00",
                "description": "Compra de gelo",
            },
        )
        self.assertEqual(movement_response.status_code, 303)
        with app.app_context():
            db = get_db()
            cash_session = get_session(db)
            summary = session_summary(db, cash_session)
            self.assertEqual(
                (summary["cash_sales"], summary["bank_sales"], summary["expected_cash"], summary["expected_bank"]),
                (300, 500, 10100, 50500),
            )
            movement_id = db.execute(
                "SELECT id FROM cash_movements WHERE description='Compra de gelo'"
            ).fetchone()["id"]

        reversed_response = self.client.post(f"/cash/movements/{movement_id}/reverse")
        self.assertEqual(reversed_response.status_code, 303)
        with app.app_context():
            db = get_db()
            cash_session = get_session(db)
            self.assertEqual(session_summary(db, cash_session)["expected_cash"], 10300)

        page = self.client.get("/cash").get_data(as_text=True)
        self.assertIn("Dinheiro físico esperado", page)
        self.assertIn("R$ 103,00", page)
        self.assertIn("Estornado", page)

        closed = self.client.post("/cash/close", data={"counted_cash": "103,00", "counted_bank": "504,00", "closing_notes": "Conferido."})
        self.assertEqual(closed.status_code, 303)
        with app.app_context():
            db = get_db()
            cash_session = get_session(db)
            self.assertEqual(cash_session["status"], "closed")
            self.assertEqual((cash_session["expected_cash_cents"], cash_session["expected_bank_cents"], cash_session["cash_difference_cents"], cash_session["bank_difference_cents"]), (10300, 50500, 0, -100))
            movement_count = db.execute("SELECT COUNT(*) total FROM cash_movements").fetchone()["total"]
        rejected = self.client.post("/cash/movements", data={"account": "cash", "direction": "in", "category": "other", "amount": "1,00", "description": "Tardio"})
        self.assertEqual(rejected.status_code, 303)
        with app.app_context():
            self.assertEqual(get_db().execute("SELECT COUNT(*) total FROM cash_movements").fetchone()["total"], movement_count)

    def test_cash_sales_are_paginated_by_ten(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        self.client.post("/cash/open", data={"opening_cash": "0,00", "opening_bank": "0,00"})
        with app.app_context():
            db = get_db()
            today = local_today().isoformat()
            for amount in range(1, 13):
                db.execute(
                    """INSERT INTO sales(player_id,payment_method,total_cents,paid,paid_at)
                       VALUES(?, 'Dinheiro', ?, 1, ?)""",
                    (self.player_id, amount * 100, f"{today} 10:{amount:02d}:00"),
                )
            db.commit()
        first = self.client.get("/cash").get_data(as_text=True)
        second = self.client.get("/cash?sales_page=2").get_data(as_text=True)
        self.assertIn("12 venda(s)", first)
        self.assertIn('aria-label="Paginação das vendas"', first)
        self.assertIn("#12", first)
        self.assertIn("#2", second)

    def test_staff_operates_cash_register_without_receiving_financial_balances(self):
        yesterday = (local_today() - timedelta(days=1)).isoformat()
        with app.app_context():
            db = get_db()
            staff = db.execute(
                """INSERT INTO users(username,name,password_hash,role)
                VALUES('atendente-caixa','Atendente Caixa','hash','staff')"""
            )
            staff_id = staff.lastrowid
            db.execute(
                """INSERT INTO cash_sessions
                (business_date,opening_cash_cents,opening_bank_cents,status,opened_by,
                 counted_cash_cents,counted_bank_cents,expected_cash_cents,expected_bank_cents,
                 cash_difference_cents,bank_difference_cents,closed_by,closed_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP)""",
                (yesterday, 10000, 60000, "closed", self.user_id, 12345, 67890,
                 12345, 67890, 0, 0, self.user_id),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = staff_id

        opening_page = self.client.get("/cash").get_data(as_text=True)
        self.assertIn("Abrir caixa de hoje", opening_page)
        for private_text in (
            "Dinheiro físico inicial", "Saldo inicial em conta", "Histórico avançado",
            "PDF deste dia", "R$ 123,45", "R$ 678,90",
        ):
            self.assertNotIn(private_text, opening_page)

        opened = self.client.post(
            "/cash/open", data={"opening_cash": "9999,99", "opening_bank": "9999,99"}
        )
        self.assertEqual(opened.status_code, 303)
        with app.app_context():
            cash_session = get_session(get_db())
            self.assertEqual(
                (cash_session["opening_cash_cents"], cash_session["opening_bank_cents"]),
                (12345, 67890),
            )

        open_page = self.client.get("/cash").get_data(as_text=True)
        self.assertIn("Encerrar caixa", open_page)
        for private_text in (
            "Dinheiro físico esperado", "Conta bancária / Pix esperada",
            "Nova entrada ou saída", "Transferir entre contas", "Vendas contabilizadas",
            "R$ 123,45", "R$ 678,90",
        ):
            self.assertNotIn(private_text, open_page)

        self.assertEqual(self.client.get("/cash/history").status_code, 302)
        self.assertEqual(self.client.get("/cash/history.pdf").status_code, 302)
        self.assertEqual(
            self.client.post(
                "/cash/movements",
                data={"account": "cash", "direction": "in", "category": "other",
                      "amount": "1,00", "description": "Tentativa"},
            ).status_code,
            302,
        )
        self.assertEqual(
            self.client.post(
                "/cash/transfers",
                data={"from_account": "cash", "to_account": "bank", "amount": "1,00"},
            ).status_code,
            302,
        )

        closed = self.client.post(
            "/cash/close", data={"counted_cash": "9999,99", "counted_bank": "9999,99"}
        )
        self.assertEqual(closed.status_code, 303)
        with app.app_context():
            cash_session = get_session(get_db())
            session_id = cash_session["id"]
            self.assertEqual(cash_session["status"], "closed")
            self.assertIsNone(cash_session["counted_cash_cents"])
            self.assertIsNone(cash_session["counted_bank_cents"])
            self.assertEqual(
                (cash_session["expected_cash_cents"], cash_session["expected_bank_cents"]),
                (12345, 67890),
            )

        closed_page = self.client.get("/cash").get_data(as_text=True)
        self.assertIn("Aguardando conferência", closed_page)
        self.assertNotIn("R$ 123,45", closed_page)
        self.assertNotIn("R$ 678,90", closed_page)

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        manager_page = self.client.get("/cash").get_data(as_text=True)
        self.assertIn("Concluir conferência financeira", manager_page)
        self.assertIn("R$ 123,45", manager_page)
        self.assertIn("R$ 678,90", manager_page)
        self.assertIn("Pendente", self.client.get("/cash/history").get_data(as_text=True))
        pending_pdf = self.client.get("/cash/history.pdf")
        self.assertEqual(pending_pdf.status_code, 200)
        self.assertTrue(pending_pdf.data.startswith(b"%PDF-"))
        reconciled = self.client.post(
            f"/cash/{session_id}/reconcile",
            data={"counted_cash": "120,00", "counted_bank": "680,00", "closing_notes": "Conferido pelo gerente."},
        )
        self.assertEqual(reconciled.status_code, 303)
        with app.app_context():
            cash_session = get_session(get_db())
            self.assertEqual(
                (cash_session["counted_cash_cents"], cash_session["counted_bank_cents"],
                 cash_session["cash_difference_cents"], cash_session["bank_difference_cents"]),
                (12000, 68000, -345, 110),
            )

    def test_finance_and_bar_keep_separate_balances_with_audited_transfers(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        setup_page = self.client.get("/finance/ledger").get_data(as_text=True)
        self.assertIn("Informe o patrimônio atual do Financeiro", setup_page)
        initialized = self.client.post(
            "/finance/ledger/initialize",
            data={"opening_cash": "200,00", "opening_bank": "1.000,00"},
        )
        self.assertEqual(initialized.status_code, 303)
        self.client.post("/cash/open", data={"opening_cash": "100,00", "opening_bank": "500,00"})

        fundraising = self.client.post(
            "/finance/ledger/movements",
            data={"account": "bank", "direction": "in", "category": "fundraising",
                  "amount": "300,00", "description": "Arrecadação do evento"},
        )
        expense = self.client.post(
            "/finance/ledger/movements",
            data={"account": "cash", "direction": "out", "category": "expense",
                  "amount": "50,00", "description": "Material administrativo"},
        )
        self.assertEqual((fundraising.status_code, expense.status_code), (303, 303))

        membership = self.client.post(
            "/finance",
            data={"player_id": self.player_id, "start_month": local_today().strftime("%Y-%m"),
                  "months_count": "1", "payment_method": "Pix", "notes": "Recebido"},
        )
        self.assertEqual(membership.status_code, 302)
        with app.app_context():
            db = get_db()
            payment_id = db.execute("SELECT id FROM membership_payments").fetchone()["id"]
            membership_entry = db.execute(
                "SELECT * FROM finance_movements WHERE source='membership' AND source_id=?",
                (payment_id,),
            ).fetchone()
            self.assertEqual(
                (membership_entry["account"], membership_entry["direction"], membership_entry["amount_cents"]),
                ("bank", "in", 1500),
            )

        to_bar = self.client.post(
            "/finance/ledger/transfers",
            data={"direction": "finance_to_bar", "amount": "250,00", "description": "Aporte para estoque"},
        )
        to_finance = self.client.post(
            "/finance/ledger/transfers",
            data={"direction": "bar_to_finance", "amount": "100,00", "description": "Devolução de aporte"},
        )
        self.assertEqual((to_bar.status_code, to_finance.status_code), (303, 303))

        with app.app_context():
            db = get_db()
            from src.services.finance_accounts import finance_summary, latest_bar_balances
            finance_balances = finance_summary(db)
            bar_balances = latest_bar_balances(db)
            self.assertEqual(
                (finance_balances["cash"], finance_balances["bank"], bar_balances["bank"]),
                (15000, 116500, 65000),
            )
            self.assertEqual(db.execute("SELECT COUNT(*) FROM interaccount_transfers").fetchone()[0], 2)
            self.assertEqual(
                db.execute("SELECT COUNT(*) FROM cash_movements WHERE source='finance_transfer'").fetchone()[0],
                2,
            )

        accounts_page = self.client.get("/finance/ledger").get_data(as_text=True)
        self.assertIn("Banco do Financeiro", accounts_page)
        self.assertIn("R$ 1.165,00", accounts_page)
        self.assertIn("R$ 1.815,00", accounts_page)
        self.assertIn("Aporte para estoque", accounts_page)
        self.assertIn("Histórico do Livro-caixa", accounts_page)
        filtered = self.client.get("/finance/ledger?account=bank&category=fundraising&q=evento")
        self.assertEqual(filtered.status_code, 200)
        self.assertIn("Arrecadação do evento", filtered.get_data(as_text=True))
        ledger_pdf = self.client.get("/finance/ledger.pdf?account=bank")
        self.assertEqual(ledger_pdf.status_code, 200)
        self.assertTrue(ledger_pdf.data.startswith(b"%PDF-"))
        self.assertIn("livro-caixa-financeiro-", ledger_pdf.headers["Content-Disposition"])

        deleted = self.client.post(f"/finance/{payment_id}/delete")
        self.assertEqual(deleted.status_code, 302)
        with app.app_context():
            db = get_db()
            from src.services.finance_accounts import finance_summary
            self.assertEqual(finance_summary(db)["bank"], 115000)
            reversal = db.execute(
                "SELECT * FROM finance_movements WHERE source='membership_reversal' AND source_id=?",
                (payment_id,),
            ).fetchone()
            self.assertEqual((reversal["direction"], reversal["amount_cents"]), ("out", 1500))

        transfer_reversal = self.client.post(
            "/finance/ledger/transfers/1/reverse",
            data={"reason": "Lançamento duplicado no aporte"},
        )
        self.assertEqual(transfer_reversal.status_code, 303)
        with app.app_context():
            db = get_db()
            from src.services.finance_accounts import finance_summary, latest_bar_balances
            self.assertEqual(finance_summary(db)["bank"], 140000)
            self.assertEqual(latest_bar_balances(db)["bank"], 40000)
            transfer = db.execute("SELECT * FROM interaccount_transfers WHERE id=1").fetchone()
            self.assertIsNotNone(transfer["reversed_at"])
            finance_reversal = db.execute(
                "SELECT * FROM finance_movements WHERE source='interaccount_transfer_reversal' AND source_id=1"
            ).fetchone()
            self.assertEqual((finance_reversal["direction"], finance_reversal["amount_cents"]), ("in", 25000))
            cash_reversal = db.execute(
                "SELECT * FROM cash_movements WHERE source='finance_transfer_reversal' AND source_id=1"
            ).fetchone()
            self.assertEqual((cash_reversal["direction"], cash_reversal["amount_cents"]), ("out", 25000))
        audit_page = self.client.get("/finance/ledger").get_data(as_text=True)
        self.assertIn("Lançamento duplicado no aporte", audit_page)
        self.assertIn("Estornado", audit_page)

        with app.app_context():
            db = get_db()
            staff = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES('staff-finance','Staff','hash','staff')"
            )
            db.commit()
            staff_id = staff.lastrowid
        with self.client.session_transaction() as session:
            session["user_id"] = staff_id
        self.assertEqual(self.client.get("/finance/ledger").status_code, 302)
        self.assertEqual(
            self.client.post(
                "/finance/ledger/movements",
                data={"account": "bank", "direction": "in", "category": "other",
                      "amount": "999,00", "description": "Tentativa"},
            ).status_code,
            302,
        )

    def _create_bar_restock_request(self, quantity=10, units_per_case=0):
        with app.app_context():
            db = get_db()
            db.execute(
                "UPDATE products SET units_per_case=? WHERE id=?",
                (units_per_case, self.product_id),
            )
            staff_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'staff')",
                ("staff.restock", "Atendente Reposição", "hash"),
            ).lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = staff_id
        response = self.client.post(
            "/stock/restock-request",
            data={f"quantity_{self.product_id}": str(quantity)},
        )
        self.assertEqual(response.status_code, 303)
        with app.app_context():
            db = get_db()
            restock_request = db.execute(
                "SELECT * FROM bar_restock_requests ORDER BY id DESC LIMIT 1"
            ).fetchone()
            item = db.execute(
                "SELECT * FROM bar_restock_request_items WHERE request_id=?",
                (restock_request["id"],),
            ).fetchone()
        return restock_request["id"], item["id"], staff_id

    def _purchase_bar_restock_request(self, request_id, amount="90.00"):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        self.client.post("/cash/open", data={"opening_cash": "0", "opening_bank": "10000"})
        for status in ("VISTA", "EM_PROCESSO"):
            response = self.client.post(
                "/stock/restock-requests",
                data={"request_id": request_id, "status": status},
            )
            self.assertEqual(response.status_code, 303)
        response = self.client.post(
            "/stock/restock-requests",
            data={
                "request_id": request_id,
                "status": "COMPRA_EFETUADA",
                "supplier": "Fornecedor Teste",
                "purchase_amount": amount,
                "payment_account": "bank",
            },
        )
        self.assertEqual(response.status_code, 303)

    def test_restock_request_approval_preserves_requested_and_controls_purchase(self):
        request_id, item_id, staff_id = self._create_bar_restock_request(
            quantity=10, units_per_case=2
        )
        with app.app_context():
            columns = {
                row[1] for row in get_db().execute(
                    "PRAGMA table_info(bar_restock_request_items)"
                ).fetchall()
            }
            self.assertIn("approved_quantity", columns)

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        manager_page = self.client.get("/stock/restock-requests").get_data(as_text=True)
        self.assertIn("Quantidade solicitada", manager_page)
        self.assertIn("Quantidade aprovada", manager_page)
        self.assertIn("Editar aprovação", manager_page)
        self.assertIn(f'name="item_id" value="{item_id}"', manager_page)
        approved = self.client.post(
            f"/stock/restock-requests/{request_id}/approval",
            data={
                "item_id": item_id,
                "approved_quantity": 6,
                "reason": "Limite disponível no caixa",
            },
        )
        self.assertEqual(approved.status_code, 303)
        with app.app_context():
            db = get_db()
            item = db.execute(
                "SELECT * FROM bar_restock_request_items WHERE id=?", (item_id,)
            ).fetchone()
            self.assertEqual((item["quantity"], item["approved_quantity"]), (10, 6))
            history = db.execute(
                "SELECT * FROM bar_restock_request_history WHERE request_id=? ORDER BY id DESC LIMIT 1",
                (request_id,),
            ).fetchone()
            self.assertIn("não definida (solicitado: 10) → 6 caixas", history["notes"])
            self.assertEqual(history["changed_by"], self.user_id)

        with self.client.session_transaction() as session:
            session["user_id"] = staff_id
        staff_page = self.client.get("/stock/restock-request").get_data(as_text=True)
        self.assertIn("Solicitado: 10 caixas", staff_page)
        self.assertIn("Aprovado: 6 caixas", staff_page)

        self._purchase_bar_restock_request(request_id)
        with app.app_context():
            db = get_db()
            self.assertEqual(
                db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"],
                17,
            )
            self.assertEqual(
                db.execute("SELECT quantity FROM restocks ORDER BY id DESC LIMIT 1").fetchone()["quantity"],
                12,
            )
        blocked = self.client.post(
            f"/stock/restock-requests/{request_id}/approval",
            data={"item_id": item_id, "approved_quantity": 5, "reason": "Tentativa tardia"},
            follow_redirects=True,
        )
        self.assertIn("não pode ser alterada após", blocked.get_data(as_text=True))
        with app.app_context():
            self.assertEqual(
                get_db().execute(
                    "SELECT approved_quantity FROM bar_restock_request_items WHERE id=?", (item_id,)
                ).fetchone()["approved_quantity"],
                6,
            )

    def test_restock_request_null_approval_uses_requested_quantity(self):
        request_id, item_id, _staff_id = self._create_bar_restock_request(quantity=3)
        self._purchase_bar_restock_request(request_id, amount="30.00")
        with app.app_context():
            db = get_db()
            item = db.execute(
                "SELECT * FROM bar_restock_request_items WHERE id=?", (item_id,)
            ).fetchone()
            self.assertIsNone(item["approved_quantity"])
            self.assertEqual(
                db.execute("SELECT quantity FROM restocks ORDER BY id DESC LIMIT 1").fetchone()["quantity"],
                3,
            )
            self.assertEqual(
                db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"],
                8,
            )

    def test_restock_request_approval_validates_quantity_and_manager_role(self):
        request_id, item_id, staff_id = self._create_bar_restock_request(quantity=10)
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        for invalid in ("0", "-1", "11", "invalida"):
            response = self.client.post(
                f"/stock/restock-requests/{request_id}/approval",
                data={"item_id": item_id, "approved_quantity": invalid, "reason": "Valor inválido"},
            )
            self.assertEqual(response.status_code, 303)
        with app.app_context():
            self.assertIsNone(
                get_db().execute(
                    "SELECT approved_quantity FROM bar_restock_request_items WHERE id=?", (item_id,)
                ).fetchone()["approved_quantity"]
            )

        with self.client.session_transaction() as session:
            session["user_id"] = staff_id
        denied_staff = self.client.post(
            f"/stock/restock-requests/{request_id}/approval",
            headers={"Accept": "application/json"},
            data={"item_id": item_id, "approved_quantity": 6, "reason": "Sem permissão"},
        )
        self.assertEqual(denied_staff.status_code, 403)
        denied_staff_value = self.client.post(
            f"/stock/restock-requests/{request_id}/value-correction",
            headers={"Accept": "application/json"},
            data={"purchase_amount": "90.00", "reason": "Sem permissão"},
        )
        self.assertEqual(denied_staff_value.status_code, 403)

        with app.app_context():
            db = get_db()
            client_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'client')",
                ("cliente.restock", "Cliente Reposição", "hash"),
            ).lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = client_id
        denied_client = self.client.post(
            f"/stock/restock-requests/{request_id}/approval",
            headers={"Accept": "application/json"},
            data={"item_id": item_id, "approved_quantity": 6, "reason": "Sem permissão"},
        )
        self.assertEqual(denied_client.status_code, 403)
        denied_client_value = self.client.post(
            f"/stock/restock-requests/{request_id}/value-correction",
            headers={"Accept": "application/json"},
            data={"purchase_amount": "90.00", "reason": "Sem permissão"},
        )
        self.assertEqual(denied_client_value.status_code, 403)

    def test_restock_request_value_correction_without_cash_is_audited_and_atomic(self):
        request_id, _item_id, _staff_id = self._create_bar_restock_request(quantity=10)
        with app.app_context():
            db = get_db()
            db.execute(
                "UPDATE bar_restock_requests SET purchase_amount_cents=900000 WHERE id=?",
                (request_id,),
            )
            original_product = db.execute(
                "SELECT stock,cost_cents FROM products WHERE id=?", (self.product_id,)
            ).fetchone()
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        missing_reason = self.client.post(
            f"/stock/restock-requests/{request_id}/value-correction",
            data={"purchase_amount": "90.00", "reason": ""},
        )
        self.assertEqual(missing_reason.status_code, 303)
        corrected = self.client.post(
            f"/stock/restock-requests/{request_id}/value-correction",
            data={"purchase_amount": "90.00", "reason": "Valor digitado incorretamente"},
        )
        self.assertEqual(corrected.status_code, 303)
        with app.app_context():
            db = get_db()
            request_row = db.execute(
                "SELECT * FROM bar_restock_requests WHERE id=?", (request_id,)
            ).fetchone()
            product = db.execute(
                "SELECT stock,cost_cents FROM products WHERE id=?", (self.product_id,)
            ).fetchone()
            self.assertEqual(request_row["purchase_amount_cents"], 9000)
            self.assertEqual(tuple(product), tuple(original_product))
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM restocks").fetchone()["total"], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM cash_movements").fetchone()["total"], 0)
            history = db.execute(
                "SELECT * FROM bar_restock_request_history WHERE request_id=? ORDER BY id DESC LIMIT 1",
                (request_id,),
            ).fetchone()
            self.assertIn("R$ 9.000,00 → R$ 90,00", history["notes"])
            self.assertIn("Valor digitado incorretamente", history["notes"])

        with app.app_context():
            db = get_db()
            db.execute(
                """CREATE TRIGGER fail_restock_value_history BEFORE INSERT ON bar_restock_request_history
                   WHEN NEW.notes LIKE 'Valor da compra corrigido:%'
                   BEGIN SELECT RAISE(ABORT, 'falha simulada'); END"""
            )
            db.commit()
        failed = self.client.post(
            f"/stock/restock-requests/{request_id}/value-correction",
            data={"purchase_amount": "80.00", "reason": "Falha transacional simulada"},
        )
        self.assertEqual(failed.status_code, 303)
        with app.app_context():
            self.assertEqual(
                get_db().execute(
                    "SELECT purchase_amount_cents FROM bar_restock_requests WHERE id=?", (request_id,)
                ).fetchone()["purchase_amount_cents"],
                9000,
            )

    def test_restock_request_value_correction_after_reversal_creates_no_movement(self):
        request_id, _item_id, _staff_id = self._create_bar_restock_request(quantity=10)
        self._purchase_bar_restock_request(request_id, amount="9000.00")
        with app.app_context():
            db = get_db()
            movement_id = db.execute(
                "SELECT id FROM cash_movements WHERE source='bar_restock_request' AND source_id=?",
                (request_id,),
            ).fetchone()["id"]
            product_before = db.execute(
                "SELECT stock,cost_cents FROM products WHERE id=?", (self.product_id,)
            ).fetchone()
        reversed_response = self.client.post(f"/cash/movements/{movement_id}/reverse")
        self.assertEqual(reversed_response.status_code, 303)
        corrected = self.client.post(
            f"/stock/restock-requests/{request_id}/value-correction",
            data={"purchase_amount": "90.00", "reason": "Compra real foi noventa reais"},
        )
        self.assertEqual(corrected.status_code, 303)
        with app.app_context():
            db = get_db()
            self.assertEqual(
                db.execute(
                    "SELECT purchase_amount_cents FROM bar_restock_requests WHERE id=?", (request_id,)
                ).fetchone()["purchase_amount_cents"],
                9000,
            )
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM cash_movements").fetchone()["total"], 2)
            product_after = db.execute(
                "SELECT stock,cost_cents FROM products WHERE id=?", (self.product_id,)
            ).fetchone()
            self.assertEqual(tuple(product_after), tuple(product_before))

    def test_restock_request_value_correction_blocks_active_movement_even_when_closed(self):
        request_id, _item_id, _staff_id = self._create_bar_restock_request(quantity=10)
        self._purchase_bar_restock_request(request_id, amount="9000.00")
        correction_url = f"/stock/restock-requests/{request_id}/value-correction"
        payload = {"purchase_amount": "90.00", "reason": "Valor digitado incorretamente"}
        active = self.client.post(correction_url, data=payload, follow_redirects=True)
        self.assertIn("movimentação financeira ativa", active.get_data(as_text=True))
        with app.app_context():
            db = get_db()
            db.execute("UPDATE cash_sessions SET status='closed'")
            db.commit()
        closed = self.client.post(correction_url, data=payload, follow_redirects=True)
        self.assertIn("movimentação financeira ativa", closed.get_data(as_text=True))
        with app.app_context():
            db = get_db()
            self.assertEqual(
                db.execute(
                    "SELECT purchase_amount_cents FROM bar_restock_requests WHERE id=?", (request_id,)
                ).fetchone()["purchase_amount_cents"],
                900000,
            )
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM cash_movements").fetchone()["total"], 1)

    def test_restock_adjustment_requests_execute_no_schema_sql(self):
        request_id, item_id, _staff_id = self._create_bar_restock_request(quantity=10)
        with app.app_context():
            db = get_db()
            db.execute(
                "UPDATE bar_restock_requests SET purchase_amount_cents=900000 WHERE id=?",
                (request_id,),
            )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        original_execute = DbWrapper.execute
        statements = []

        def reject_schema_sql(wrapper, statement, params=()):
            normalized = " ".join(str(statement).split())
            if has_request_context():
                statements.append(normalized)
                if self.SCHEMA_SQL.match(normalized):
                    raise AssertionError(f"DDL executado durante request HTTP: {normalized[:120]}")
            return original_execute(wrapper, statement, params)

        with patch.object(DbWrapper, "execute", new=reject_schema_sql), patch("src.db.init_sqlite"):
            approval = self.client.post(
                f"/stock/restock-requests/{request_id}/approval",
                data={"item_id": item_id, "approved_quantity": 6, "reason": "Compra parcial"},
            )
            correction = self.client.post(
                f"/stock/restock-requests/{request_id}/value-correction",
                data={"purchase_amount": "90.00", "reason": "Correção do valor"},
            )
        self.assertEqual((approval.status_code, correction.status_code), (303, 303))
        self.assertTrue(statements)
        self.assertFalse(any(self.SCHEMA_SQL.match(statement) for statement in statements))

    def test_manager_corrects_restock_with_audit_trail(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        created = self.client.post(
            "/stock",
            data={
                "product_id": self.product_id,
                "quantity": 10,
                "cases": 0,
                "unit_cost": "2,00",
                "notes": "Entrada digitada errada",
            },
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            restock_id = db.execute("SELECT MAX(id) id FROM restocks").fetchone()["id"]
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 15)

        corrected = self.client.post(
            f"/stock/restocks/{restock_id}/correct",
            data={"quantity": 6, "unit_cost": "1,50", "reason": "Quantidade e custo digitados errados"},
        )
        self.assertEqual(corrected.status_code, 303)
        with app.app_context():
            db = get_db()
            product = db.execute("SELECT * FROM products WHERE id=?", (self.product_id,)).fetchone()
            original = db.execute("SELECT * FROM restocks WHERE id=?", (restock_id,)).fetchone()
            correction = db.execute("SELECT * FROM restock_corrections WHERE restock_id=?", (restock_id,)).fetchone()
            self.assertEqual((product["stock"], product["cost_cents"]), (11, 150))
            self.assertEqual((original["quantity"], original["unit_cost_cents"]), (10, 200))
            self.assertEqual(
                (correction["previous_quantity"], correction["corrected_quantity"], correction["previous_unit_cost_cents"], correction["corrected_unit_cost_cents"]),
                (10, 6, 200, 150),
            )

        corrected_again = self.client.post(
            f"/stock/restocks/{restock_id}/correct",
            data={"quantity": 7, "unit_cost": "1,75", "reason": "Recontagem feita pelo gerente"},
        )
        self.assertEqual(corrected_again.status_code, 303)
        with app.app_context():
            db = get_db()
            product = db.execute("SELECT * FROM products WHERE id=?", (self.product_id,)).fetchone()
            latest = db.execute("SELECT * FROM restock_corrections ORDER BY id DESC LIMIT 1").fetchone()
            self.assertEqual((product["stock"], product["cost_cents"]), (12, 175))
            self.assertEqual((latest["previous_quantity"], latest["corrected_quantity"]), (6, 7))
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM restock_corrections").fetchone()["total"], 2)

        page = self.client.get("/stock").get_data(as_text=True)
        self.assertIn("Corrigida", page)
        self.assertIn("Original: 10 un.", page)
        self.assertIn("Recontagem feita pelo gerente", page)

        with app.app_context():
            db = get_db()
            staff = db.execute(
                "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'staff')",
                ("staff.estoque", "Staff Estoque", "hash"),
            ).lastrowid
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = staff
        denied = self.client.get(f"/stock/restocks/{restock_id}/correct")
        self.assertEqual(denied.status_code, 302)

    def test_paid_stock_purchase_creates_atomic_cash_outflow(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        rejected = self.client.post(
            "/stock",
            data={
                "product_id": self.product_id,
                "quantity": 3,
                "cases": 0,
                "unit_cost": "2,00",
                "payment_account": "bank",
                "notes": "Sem caixa",
            },
        )
        self.assertEqual(rejected.status_code, 302)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 5)
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM restocks").fetchone()["total"], 0)

        self.client.post("/cash/open", data={"opening_cash": "0,00", "opening_bank": "100,00"})
        accepted = self.client.post(
            "/stock",
            data={
                "product_id": self.product_id,
                "quantity": 3,
                "cases": 0,
                "unit_cost": "2,00",
                "payment_account": "bank",
                "notes": "Compra paga por Pix",
            },
        )
        self.assertEqual(accepted.status_code, 302)
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 8)
            movement = db.execute("SELECT * FROM cash_movements WHERE source='restock'").fetchone()
            self.assertEqual(
                (movement["account"], movement["direction"], movement["category"], movement["amount_cents"]),
                ("bank", "out", "purchase", 600),
            )
            summary = session_summary(db, get_session(db))
            self.assertEqual(summary["expected_bank"], 9400)

    def test_stock_report_pdf_lists_current_stock_entries_and_exits(self):
        from src.services.stock_report_pdf import stock_report_data

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            db.execute(
                "INSERT INTO restocks(product_id,quantity,unit_cost_cents,notes,created_at) VALUES(?,?,?,?,?)",
                (self.product_id, 4, 100, "Entrada de teste", "2026-07-10 12:00:00"),
            )
            sale = db.execute(
                "INSERT INTO sales(player_id,payment_method,total_cents,paid,payment_status,paid_at,created_at) VALUES(?,?,?,?,?,?,?)",
                (self.player_id, "Dinheiro", 600, 1, "approved", "2026-07-11 12:00:00", "2026-07-11 12:00:00"),
            )
            db.execute("INSERT INTO sale_items(sale_id,product_id,quantity,unit_price_cents,unit_cost_cents) VALUES(?,?,?,?,?)",
                       (sale.lastrowid, self.product_id, 2, 300, 100))
            db.commit()
            rows = stock_report_data(db, "2026-07-01", "2026-07-31")
            agua = next(row for row in rows if row["name"] == "Água")
            self.assertEqual((agua["stock"], agua["entries"], agua["exits"], agua["net"]), (5, 4, 2, 2))

        response = self.client.get("/stock/report.pdf?start=2026-07-01&end=2026-07-31")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("relatorio-estoque-", response.headers["Content-Disposition"])
        self.assertTrue(response.data.startswith(b"%PDF-"))

    def test_monthly_stock_conference_requires_reason_and_preserves_expected_stock(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        invalid = self.client.post(
            "/stock/conference",
            data={
                "conference_month": "2026-08",
                f"physical_{self.product_id}": "3",
                f"reason_{self.product_id}": "",
            },
        )
        self.assertEqual(invalid.status_code, 200)
        self.assertIn("Informe o motivo da diferença", invalid.get_data(as_text=True))
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) FROM stock_conferences").fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 5)

        created = self.client.post(
            "/stock/conference",
            data={
                "conference_month": "2026-08",
                "notes": "Fechamento do bar",
                f"physical_{self.product_id}": "3",
                f"reason_{self.product_id}": "Perda registrada",
            },
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            db = get_db()
            conference = db.execute("SELECT * FROM stock_conferences WHERE conference_month='2026-08'").fetchone()
            item = db.execute("SELECT * FROM stock_conference_items WHERE conference_id=?", (conference["id"],)).fetchone()
            audit = db.execute("SELECT action FROM stock_conference_audit WHERE conference_month='2026-08'").fetchone()
            self.assertEqual((item["expected_stock"], item["physical_stock"], item["difference"], item["reason"]), (5, 3, -2, "Perda registrada"))
            self.assertEqual(audit["action"], "REGISTRADA")
            self.assertEqual(db.execute("SELECT stock FROM products WHERE id=?", (self.product_id,)).fetchone()["stock"], 5)

        duplicate = self.client.post(
            "/stock/conference",
            data={"conference_month": "2026-08", f"physical_{self.product_id}": "5"},
            follow_redirects=True,
        )
        self.assertEqual(duplicate.status_code, 200)
        self.assertIn("Já existe uma conferência", duplicate.get_data(as_text=True))

    def test_manager_can_remove_conference_with_audit_and_staff_cannot(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        created = self.client.post(
            "/stock/conference",
            data={"conference_month": "2026-09", f"physical_{self.product_id}": "5"},
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            conference_id = get_db().execute("SELECT id FROM stock_conferences WHERE conference_month='2026-09'").fetchone()["id"]
            db = get_db()
            db.execute("UPDATE users SET role='staff' WHERE id=?", (self.user_id,))
            db.commit()
        denied = self.client.post(f"/stock/conference/{conference_id}/delete")
        self.assertIn(denied.status_code, (302, 403))
        with app.app_context():
            db = get_db()
            self.assertIsNotNone(db.execute("SELECT id FROM stock_conferences WHERE id=?", (conference_id,)).fetchone())
            db.execute("UPDATE users SET role='manager' WHERE id=?", (self.user_id,))
            db.commit()
        deleted = self.client.post(f"/stock/conference/{conference_id}/delete")
        self.assertEqual(deleted.status_code, 302)
        with app.app_context():
            db = get_db()
            self.assertIsNone(db.execute("SELECT id FROM stock_conferences WHERE id=?", (conference_id,)).fetchone())
            self.assertEqual(db.execute("SELECT action FROM stock_conference_audit WHERE conference_month='2026-09' ORDER BY id DESC").fetchone()["action"], "EXCLUIDA")

    def test_low_stock_alert_consolidates_products_for_supplier_once(self):
        with app.app_context(), patch.dict("os.environ", {
            "GMAIL_SMTP_USER": "bar@example.com",
            "GMAIL_APP_PASSWORD": "app-password",
            "STOCK_ALERT_ATTENDANT_EMAIL": "atendente@example.com",
            "STOCK_ALERT_MANAGER_EMAIL": "gerente@example.com",
        }):
            db = get_db()
            db.execute("UPDATE products SET stock=2,min_stock=2,supplier_email='fornecedor@example.com' WHERE id=?", (self.product_id,))
            sent = []
            result = notify_low_stock(db, [self.product_id], send_func=lambda *args: sent.append(args[2]))
            self.assertEqual(result["sent"], 1)
            self.assertEqual(sent, ["fornecedor@example.com"])
            again = notify_low_stock(db, [self.product_id], send_func=lambda *args: sent.append(args[2]))
            self.assertEqual(again["skipped"], 1)
            self.assertEqual(len(sent), 1)

    def test_low_stock_alert_resets_after_replenishment(self):
        with app.app_context(), patch.dict("os.environ", {
            "GMAIL_SMTP_USER": "bar@example.com",
            "GMAIL_APP_PASSWORD": "app-password",
            "STOCK_ALERT_ATTENDANT_EMAIL": "atendente@example.com",
        }):
            db = get_db()
            db.execute("UPDATE products SET stock=1,min_stock=2,supplier_email='fornecedor@example.com' WHERE id=?", (self.product_id,))
            sent = []
            notify_low_stock(db, [self.product_id], send_func=lambda *args: sent.append(args[2]))
            db.execute("UPDATE products SET stock=5 WHERE id=?", (self.product_id,))
            notify_low_stock(db, [self.product_id], send_func=lambda *args: sent.append(args[2]))
            db.execute("UPDATE products SET stock=2 WHERE id=?", (self.product_id,))
            result = notify_low_stock(db, [self.product_id], send_func=lambda *args: sent.append(args[2]))
            self.assertEqual(result["sent"], 1)
            self.assertEqual(len(sent), 2)

    def test_staff_and_manager_can_generate_low_stock_pdf(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with app.app_context():
            db = get_db()
            db.execute("UPDATE products SET stock=1,min_stock=5,supplier_email='fornecedor@example.com' WHERE id=?", (self.product_id,))
            db.commit()
        response = self.client.get("/stock/low-report.pdf")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("estoque-baixo-", response.headers["Content-Disposition"])
        self.assertTrue(response.data.startswith(b"%PDF-"))
    def test_cash_transfer_history_filters_and_pdf(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        self.client.post("/cash/open", data={"opening_cash": "100,00", "opening_bank": "50,00"})

        transferred = self.client.post(
            "/cash/transfers",
            data={
                "from_account": "cash",
                "to_account": "bank",
                "amount": "30,00",
                "description": "Depósito do dinheiro das vendas",
            },
        )
        self.assertEqual(transferred.status_code, 303)
        with app.app_context():
            db = get_db()
            cash_session = get_session(db)
            summary = session_summary(db, cash_session)
            self.assertEqual((summary["expected_cash"], summary["expected_bank"]), (7000, 8000))
            transfer = db.execute("SELECT * FROM cash_transfers").fetchone()
            transfer_id = transfer["id"]
            legs = db.execute(
                "SELECT account,direction,amount_cents FROM cash_movements ORDER BY id"
            ).fetchall()
            self.assertEqual(
                [(row["account"], row["direction"], row["amount_cents"]) for row in legs],
                [("cash", "out", 3000), ("bank", "in", 3000)],
            )

        history = self.client.get(
            "/cash/history?account=cash&category=transfer&q=dep%C3%B3sito"
        )
        self.assertEqual(history.status_code, 200)
        history_html = history.get_data(as_text=True)
        self.assertIn("Histórico avançado do Caixa", history_html)
        self.assertIn("Depósito do dinheiro das vendas", history_html)
        self.assertIn("Movimentações e transferências (1)", history_html)

        pdf = self.client.get("/cash/history.pdf?category=transfer")
        self.assertEqual(pdf.status_code, 200)
        self.assertTrue(pdf.data.startswith(b"%PDF-"))
        self.assertIn("caixa-", pdf.headers["Content-Disposition"])

        reversed_transfer = self.client.post(f"/cash/transfers/{transfer_id}/reverse")
        self.assertEqual(reversed_transfer.status_code, 303)
        with app.app_context():
            db = get_db()
            summary = session_summary(db, get_session(db))
            self.assertEqual((summary["expected_cash"], summary["expected_bank"]), (10000, 5000))
            transfer = db.execute("SELECT * FROM cash_transfers WHERE id=?", (transfer_id,)).fetchone()
            self.assertIsNotNone(transfer["reversed_at"])
            self.assertEqual(db.execute("SELECT COUNT(*) total FROM cash_movements").fetchone()["total"], 4)

    def test_goal_types_control_assists_and_own_goal_statistics(self):
        with app.app_context():
            db = get_db()
            assist_player_id = db.execute(
                "INSERT INTO players(name,war_name) VALUES(?,?)",
                ("Assistente", "Garçom"),
            ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','EM_ANDAMENTO',?)",
                ("2026-08-08", self.user_id),
            ).lastrowid
            match_id = db.execute(
                "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,1,3,0,'ENCERRADA')",
                (sumula_id,),
            ).lastrowid
            for player_id in (self.player_id, assist_player_id):
                db.execute(
                    "INSERT INTO football_participants(sumula_id,player_id,status) VALUES(?,?,'CONFIRMADO')",
                    (sumula_id, player_id),
                )
            client_user_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("artilheiro", "Artilheiro", "hash", self.player_id),
            ).lastrowid
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        detail_html = self.client.get(f"/futebol/sumulas/{sumula_id}").get_data(as_text=True)
        self.assertNotIn("Gol normal", detail_html)
        self.assertNotIn('name="minute"', detail_html)
        self.assertNotIn('placeholder="Min."', detail_html)
        self.assertIn("['NORMAL','Gol']", detail_html)
        self.assertIn("['REBOTE','Rebote']", detail_html)
        self.assertIn("['FALTA','Falta']", detail_html)
        self.assertIn("['PENALTY','Penalty']", detail_html)
        self.assertIn("['ROUBADA','Roubada']", detail_html)
        self.assertIn("['CONTRA','Gol Contra']", detail_html)

        common_goal = {
            "action": "goal",
            "match_id": str(match_id),
            "author_player_id": str(self.player_id),
            "assist_player_id": str(assist_player_id),
            "benefited_team": "AZUL",
            "minute": "1",
        }
        for goal_type in ("NORMAL", "REBOTE", "CONTRA"):
            response = self.client.post(
                f"/futebol/sumulas/{sumula_id}",
                data={**common_goal, "goal_type": goal_type},
            )
            self.assertEqual(response.status_code, 302)

        with app.app_context():
            db = get_db()
            goals = db.execute(
                "SELECT goal_type,own_goal,assist_player_id FROM football_goals ORDER BY id"
            ).fetchall()
            self.assertEqual(
                [(row["goal_type"], row["own_goal"], row["assist_player_id"]) for row in goals],
                [("NORMAL", 0, assist_player_id), ("REBOTE", 0, None), ("CONTRA", 1, None)],
            )
            self.assertEqual(
                db.execute("SELECT COUNT(*) FROM football_goals WHERE minute IS NOT NULL").fetchone()[0],
                0,
            )
            db.execute(
                "UPDATE football_sumulas SET situacao='FINALIZADA',finalized_at=CURRENT_TIMESTAMP WHERE id=?",
                (sumula_id,),
            )
            db.commit()

        statistics_html = self.client.get("/futebol/estatisticas").get_data(as_text=True)
        self.assertIn("Gols contra", statistics_html)
        self.assertRegex(statistics_html, r"Assistências registradas</small><h2>1</h2>")
        self.assertIn('<h2 class="text-danger">−1</h2>', statistics_html)
        self.assertIn('<td class="text-danger fw-semibold">−1</td>', statistics_html)
        self.assertRegex(
            statistics_html,
            r"Peladeiro</td><td>100\.0%</td><td>0</td><td>0</td><td>0</td><td>0</td><td>1</td><td class=\"text-danger fw-semibold\">−1</td><td>0</td>",
        )
        self.assertRegex(
            statistics_html,
            r"Garçom</td><td>100\.0%</td><td>0</td><td>0</td><td>0</td><td>0</td><td>0</td><td class=\"text-muted\">0</td><td>1</td>",
        )

        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id
        client_html = self.client.get("/futebol/minha-pelada").get_data(as_text=True)
        self.assertIn(
            '<strong class="fs-3">1</strong><small class="d-block text-muted">Gols</small>',
            client_html,
        )
        self.assertIn(
            '<strong class="fs-3">0</strong><small class="d-block text-muted">Assistências</small>',
            client_html,
        )

    def test_football_statistics_paginates_ranking_and_recent_results_independently(self):
        with app.app_context():
            db = get_db()
            for number in range(1, 17):
                player_id = db.execute(
                    "INSERT INTO players(name,war_name) VALUES(?,?)",
                    (f"Jogador paginado {number:02d}", f"Ranking {number:02d}"),
                ).lastrowid
                db.execute(
                    "INSERT INTO football_historical_stats(player_id,stat_date,goals,created_by) VALUES(?,'2026-08-08',1,?)",
                    (player_id, self.user_id),
                )
            for day in range(1, 13):
                sumula_id = db.execute(
                    "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','FINALIZADA',?)",
                    (f"2026-08-{day:02d}", self.user_id),
                ).lastrowid
                db.execute(
                    "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,1,2,1,'ENCERRADA')",
                    (sumula_id,),
                )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        first_page = self.client.get("/futebol/estatisticas?year=2026&month=8").get_data(as_text=True)
        self.assertEqual(first_page.count('data-ranking-row="'), 15)
        self.assertEqual(first_page.count('data-result-date="'), 10)
        self.assertIn('data-ranking-page="1"', first_page)
        self.assertIn('data-results-page="1"', first_page)
        self.assertIn('data-result-date="2026-08-12"', first_page)
        self.assertNotIn('data-result-date="2026-08-02"', first_page)

        second_page = self.client.get(
            "/futebol/estatisticas?year=2026&month=8&ranking_page=2&results_page=2"
        ).get_data(as_text=True)
        self.assertEqual(second_page.count('data-ranking-row="'), 1)
        self.assertEqual(second_page.count('data-result-date="'), 2)
        self.assertIn('data-ranking-page="2"', second_page)
        self.assertIn('data-results-page="2"', second_page)
        self.assertIn('data-result-date="2026-08-02"', second_page)
        self.assertIn('data-result-date="2026-08-01"', second_page)
        self.assertNotIn('data-result-date="2026-08-03"', second_page)

    def test_football_statistics_ranks_players_by_first_and_second_match(self):
        with app.app_context():
            db = get_db()
            rival_id = db.execute(
                "INSERT INTO players(name,war_name) VALUES(?,?)", ("Rival da partida", "Rival")
            ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES('2026-08-08','SABADO','FINALIZADA',?)",
                (self.user_id,),
            ).lastrowid
            match_1 = db.execute(
                "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,1,1,2,'ENCERRADA')",
                (sumula_id,),
            ).lastrowid
            match_2 = db.execute(
                "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,2,3,1,'ENCERRADA')",
                (sumula_id,),
            ).lastrowid
            for match_id in (match_1, match_2):
                db.execute(
                    "INSERT INTO football_lineups(match_id,player_id,team,position) VALUES(?,?,'AZUL','ATACANTE')",
                    (match_id, self.player_id),
                )
                db.execute(
                    "INSERT INTO football_lineups(match_id,player_id,team,position) VALUES(?,?,'BRANCO','DEFENSOR')",
                    (match_id, rival_id),
                )
            db.execute(
                "INSERT INTO football_goals(match_id,author_player_id,benefited_team,assist_player_id,goal_type,own_goal) VALUES(?,?,'AZUL',?,'NORMAL',0)",
                (match_1, self.player_id, rival_id),
            )
            db.execute(
                "INSERT INTO football_goals(match_id,author_player_id,benefited_team,goal_type,own_goal) VALUES(?,?,'BRANCO','CONTRA',1)",
                (match_1, self.player_id),
            )
            db.execute(
                "INSERT INTO football_goals(match_id,author_player_id,benefited_team,goal_type,own_goal) VALUES(?,?,'AZUL','NORMAL',0)",
                (match_2, self.player_id),
            )
            db.execute(
                "INSERT INTO football_goals(match_id,author_player_id,benefited_team,goal_type,own_goal) VALUES(?,?,'AZUL','CONTRA',1)",
                (match_2, rival_id),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        html = self.client.get("/futebol/estatisticas?year=2026&month=8").get_data(as_text=True)
        self.assertIn("Ranking por partida", html)
        self.assertIn('data-match-ranking="1"', html)
        self.assertIn('data-match-ranking="2"', html)
        self.assertIn(
            f'<tr data-match-ranking-row="1-{self.player_id}"><td>Peladeiro</td><td>1</td><td>1</td><td class="">0</td><td class="text-danger fw-semibold">−1</td><td>0</td></tr>',
            html,
        )
        self.assertIn(
            f'<tr data-match-ranking-row="1-{rival_id}"><td>Rival</td><td>1</td><td>0</td><td class="">0</td><td class="text-muted">0</td><td>1</td></tr>',
            html,
        )
        self.assertIn(
            f'<tr data-match-ranking-row="2-{self.player_id}"><td>Peladeiro</td><td>1</td><td>0</td><td class="">1</td><td class="text-muted">0</td><td>0</td></tr>',
            html,
        )
        self.assertIn(
            f'<tr data-match-ranking-row="2-{rival_id}"><td>Rival</td><td>1</td><td>1</td><td class="text-danger fw-semibold">-1</td><td class="text-danger fw-semibold">−1</td><td>0</td></tr>',
            html,
        )

    def test_football_statistics_paginates_each_match_ranking_independently(self):
        with app.app_context():
            db = get_db()
            player_ids = []
            for number in range(1, 13):
                player_ids.append(
                    db.execute(
                        "INSERT INTO players(name,war_name) VALUES(?,?)",
                        (f"Jogador por partida {number:02d}", f"Partida {number:02d}"),
                    ).lastrowid
                )
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES('2026-08-08','SABADO','FINALIZADA',?)",
                (self.user_id,),
            ).lastrowid
            for match_number in (1, 2):
                match_id = db.execute(
                    "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,?,2,1,'ENCERRADA')",
                    (sumula_id, match_number),
                ).lastrowid
                for player_id in player_ids:
                    db.execute(
                        "INSERT INTO football_lineups(match_id,player_id,team,position) VALUES(?,?,'AZUL','ATACANTE')",
                        (match_id, player_id),
                    )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        first_page = self.client.get(
            "/futebol/estatisticas?year=2026&month=8"
        ).get_data(as_text=True)
        self.assertEqual(first_page.count('data-match-ranking-row="1-'), 10)
        self.assertEqual(first_page.count('data-match-ranking-row="2-'), 10)
        self.assertIn('data-match-ranking-page="1-1"', first_page)
        self.assertIn('data-match-ranking-page="2-1"', first_page)

        independent_page = self.client.get(
            "/futebol/estatisticas?year=2026&month=8&match1_page=2"
        ).get_data(as_text=True)
        self.assertEqual(independent_page.count('data-match-ranking-row="1-'), 2)
        self.assertEqual(independent_page.count('data-match-ranking-row="2-'), 10)
        self.assertIn('data-match-ranking-page="1-2"', independent_page)
        self.assertIn('data-match-ranking-page="2-1"', independent_page)

        both_second_pages = self.client.get(
            "/futebol/estatisticas?year=2026&month=8&match1_page=2&match2_page=2"
        ).get_data(as_text=True)
        self.assertEqual(both_second_pages.count('data-match-ranking-row="1-'), 2)
        self.assertEqual(both_second_pages.count('data-match-ranking-row="2-'), 2)
        self.assertIn('data-match-ranking-page="1-2"', both_second_pages)
        self.assertIn('data-match-ranking-page="2-2"', both_second_pages)

    def test_transfer_window_lists_only_players_eligible_for_a_new_position(self):
        with app.app_context():
            db = get_db()
            eligible_id = db.execute(
                "INSERT INTO players(name,war_name,football_position,football_join_date) VALUES(?,?,?,?)",
                ("Apto transferência", "Apto", "DEFESA", "2025-01-01"),
            ).lastrowid
            db.execute(
                "INSERT INTO players(name,war_name,football_position,football_join_date) VALUES(?,?,?,?)",
                ("Sem tempo mínimo", "Recente", "DEFESA", local_today().isoformat()),
            )
            db.execute(
                "INSERT INTO players(name,war_name,football_position,football_join_date) VALUES(?,?,?,?)",
                ("Sem frequência", "Ausente", "DEFESA", "2025-01-01"),
            )
            db.execute(
                "INSERT INTO players(name,football_position) VALUES(?,?)",
                ("Defensor de apoio", "DEFESA"),
            )
            for index in range(3):
                db.execute(
                    "INSERT INTO players(name,football_position) VALUES(?,?)",
                    (f"Meio de apoio {index}", "MEIO"),
                )
                db.execute(
                    "INSERT INTO players(name,football_position) VALUES(?,?)",
                    (f"Atacante de apoio {index}", "ATAQUE"),
                )
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','FINALIZADA',?)",
                (local_today().isoformat(), self.user_id),
            ).lastrowid
            db.execute(
                "INSERT INTO football_participants(sumula_id,player_id,status) VALUES(?,?,'CONFIRMADO')",
                (sumula_id, eligible_id),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        html = self.client.get("/futebol/transferencia").get_data(as_text=True)
        self.assertIn("Peladeiros aptos à transferência", html)
        self.assertIn(f'data-eligible-transfer-player="{eligible_id}"', html)
        self.assertIn("Apto", html)
        self.assertIn("Disponível para solicitar", html)
        self.assertIn("Meio", html)
        self.assertIn("Ataque", html)
        self.assertNotIn("Recente", html)
        self.assertNotIn("Ausente", html)

    def test_client_mathematician_menu_shows_mathematical_and_zebra_percentages(self):
        with app.app_context():
            db = get_db()
            client_user_id = db.execute(
                "INSERT INTO users(username,name,password_hash,role,player_id) VALUES(?,?,?,'client',?)",
                ("matematico", "Cliente matemático", "hash", self.player_id),
            ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES('2026-08-08','SABADO','FINALIZADA',?)",
                (self.user_id,),
            ).lastrowid
            for number, blue_score, white_score in ((1, 3, 2), (2, 4, 2), (3, 5, 2)):
                db.execute(
                    "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,?,?,?,'ENCERRADA')",
                    (sumula_id, number, blue_score, white_score),
                )
            ignored_sumula = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES('2026-08-09','SABADO','RASCUNHO',?)",
                (self.user_id,),
            ).lastrowid
            db.execute(
                "INSERT INTO football_matches(sumula_id,number,blue_score,white_score,status) VALUES(?,1,9,0,'ENCERRADA')",
                (ignored_sumula,),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = client_user_id
        response = self.client.get("/futebol/e-matematico")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("É matemático!!!!", html)
        self.assertIn('href="/futebol/e-matematico"', html)
        self.assertIn("66.7%", html)
        self.assertIn("33.3%", html)
        self.assertIn("2 de 3", html)
        self.assertIn("1 de 3", html)
        self.assertEqual(html.count('data-mathematician-result="'), 3)
        self.assertIn("Matemática", html)
        self.assertIn("Zebra", html)

    def test_third_match_accepts_participant_orders_45_to_66(self):
        with app.app_context():
            db = get_db()
            occupied_player_ids = [self.player_id]
            for number in range(2, 45):
                occupied_player_ids.append(
                    db.execute(
                        "INSERT INTO players(name,war_name) VALUES(?,?)",
                        (f"Jogador {number}", f"J{number}"),
                    ).lastrowid
                )
            third_match_player_id = db.execute(
                "INSERT INTO players(name,war_name) VALUES(?,?)",
                ("Jogador da terceira", "Terceira"),
            ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','EM_ANDAMENTO',?)",
                ("2026-08-15", self.user_id),
            ).lastrowid
            for match_number in (1, 2, 3):
                db.execute(
                    "INSERT INTO football_matches(sumula_id,number) VALUES(?,?)",
                    (sumula_id, match_number),
                )
            for draw_order, player_id in enumerate(occupied_player_ids, start=1):
                db.execute(
                    "INSERT INTO football_participants(sumula_id,player_id,status,draw_order) VALUES(?,?,'CONFIRMADO',?)",
                    (sumula_id, player_id, draw_order),
                )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        page = self.client.get(f"/futebol/sumulas/{sumula_id}").get_data(as_text=True)
        self.assertIn('max="66"', page)
        self.assertIn('value="45" placeholder="Ordem do sorteio"', page)
        self.assertIn("Partida 3", page)
        self.assertIn("ordens 45 a 66", page)

        added = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "participant",
                "player_id": str(third_match_player_id),
                "status": "CONFIRMADO",
                "preferred_position": "ATACANTE",
            },
        )
        self.assertEqual(added.status_code, 302)

        with app.app_context():
            db = get_db()
            participant = db.execute(
                "SELECT id,draw_order FROM football_participants WHERE sumula_id=? AND player_id=?",
                (sumula_id, third_match_player_id),
            ).fetchone()
            self.assertEqual(participant["draw_order"], 45)
            participant_id = participant["id"]

        moved = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "update_participant_order",
                "participant_id": str(participant_id),
                "draw_order": "66",
            },
        )
        self.assertEqual(moved.status_code, 302)
        rejected = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "update_participant_order",
                "participant_id": str(participant_id),
                "draw_order": "67",
            },
        )
        self.assertEqual(rejected.status_code, 302)

        with app.app_context():
            draw_order = get_db().execute(
                "SELECT draw_order FROM football_participants WHERE id=?",
                (participant_id,),
            ).fetchone()[0]
            self.assertEqual(draw_order, 66)

        detail_html = self.client.get(f"/futebol/sumulas/{sumula_id}").get_data(as_text=True)
        self.assertIn("Terceira", detail_html)
        self.assertIn('value="66" title="Ordem do sorteio"', detail_html)
        print_html = self.client.get(f"/futebol/sumulas/{sumula_id}/imprimir").get_data(as_text=True)
        self.assertIn("3ª PARTIDA (ordens 45 a 66)", print_html)
        self.assertIn("Terceira", print_html)

    def test_participant_order_can_start_at_three_and_then_stays_sequential(self):
        with app.app_context():
            db = get_db()
            second_player_id = db.execute(
                "INSERT INTO players(name,war_name) VALUES(?,?)",
                ("Segundo da sequência", "Sequência 2"),
            ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','EM_ANDAMENTO',?)",
                ("2026-08-22", self.user_id),
            ).lastrowid
            db.execute("INSERT INTO football_matches(sumula_id,number) VALUES(?,1)", (sumula_id,))
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        first = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "participant",
                "player_id": str(self.player_id),
                "status": "CONFIRMADO",
                "preferred_position": "DEFENSOR",
                "draw_order": "3",
            },
        )
        self.assertEqual(first.status_code, 302)
        page = self.client.get(f"/futebol/sumulas/{sumula_id}").get_data(as_text=True)
        self.assertIn('value="4" placeholder="Ordem do sorteio"', page)
        self.assertNotIn("participantOrder.value=", page)

        second = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "participant",
                "player_id": str(second_player_id),
                "status": "CONFIRMADO",
                "preferred_position": "MEIO_CAMPO",
                "draw_order": "",
            },
        )
        self.assertEqual(second.status_code, 302)

        with app.app_context():
            rows = get_db().execute(
                "SELECT player_id,draw_order FROM football_participants WHERE sumula_id=? ORDER BY draw_order",
                (sumula_id,),
            ).fetchall()
            self.assertEqual(
                [(row["player_id"], row["draw_order"]) for row in rows],
                [(self.player_id, 3), (second_player_id, 4)],
            )

    def test_bulk_lineup_modal_saves_and_replaces_only_selected_match_period(self):
        with app.app_context():
            db = get_db()
            second_player_id = db.execute(
                "INSERT INTO players(name,war_name,football_position) VALUES(?,?,?)",
                ("Segundo escalado", "Segundo", "ATAQUE"),
            ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','EM_ANDAMENTO',?)",
                ("2026-08-23", self.user_id),
            ).lastrowid
            match_id = db.execute(
                "INSERT INTO football_matches(sumula_id,number) VALUES(?,1)", (sumula_id,)
            ).lastrowid
            for order, player_id, position in (
                (1, self.player_id, "DEFENSOR"),
                (2, second_player_id, "ATACANTE"),
            ):
                db.execute(
                    "INSERT INTO football_participants(sumula_id,player_id,status,preferred_position,draw_order) "
                    "VALUES(?,?,'CONFIRMADO',?,?)",
                    (sumula_id, player_id, position, order),
                )
            db.execute(
                "INSERT INTO football_lineups(match_id,player_id,team,position,period) VALUES(?,?,'BRANCO','DEFENSOR',2)",
                (match_id, self.player_id),
            )
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        page = self.client.get(f"/futebol/sumulas/{sumula_id}")
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn('id="bulk-lineup-modal"', html)
        self.assertIn("Montar escalação", html)
        self.assertIn(f'name="assignment_{self.player_id}"', html)
        self.assertIn(f'name="assignment_{second_player_id}"', html)

        saved = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "bulk_lineup", "match_id": str(match_id), "period": "1",
                f"assignment_{self.player_id}": "AZUL",
                f"position_{self.player_id}": "DEFENSOR",
                f"assignment_{second_player_id}": "BRANCO",
                f"position_{second_player_id}": "ATACANTE",
            },
        )
        self.assertEqual(saved.status_code, 302)
        self.assertTrue(saved.headers["Location"].endswith("#lineup-management"))

        replaced = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "bulk_lineup", "match_id": str(match_id), "period": "1",
                f"assignment_{self.player_id}": "BRANCO",
                f"position_{self.player_id}": "MEIO_CAMPO",
                f"assignment_{second_player_id}": "",
            },
        )
        self.assertEqual(replaced.status_code, 302)
        with app.app_context():
            rows = get_db().execute(
                "SELECT player_id,team,position,period FROM football_lineups WHERE match_id=? ORDER BY period,player_id",
                (match_id,),
            ).fetchall()
            self.assertEqual(
                [(row["player_id"], row["team"], row["position"], row["period"]) for row in rows],
                [
                    (self.player_id, "BRANCO", "MEIO_CAMPO", 1),
                    (self.player_id, "BRANCO", "DEFENSOR", 2),
                ],
            )

        rejected = self.client.post(
            f"/futebol/sumulas/{sumula_id}",
            data={
                "action": "bulk_lineup", "match_id": str(match_id), "period": "1",
                f"assignment_{second_player_id}": "VERDE",
                f"position_{second_player_id}": "ATACANTE",
            },
            follow_redirects=True,
        )
        self.assertIn("Time inválido na escalação", rejected.get_data(as_text=True))
        with app.app_context():
            preserved = get_db().execute(
                "SELECT player_id,team,position FROM football_lineups WHERE match_id=? AND period=1",
                (match_id,),
            ).fetchall()
            self.assertEqual(
                [(row["player_id"], row["team"], row["position"]) for row in preserved],
                [(self.player_id, "BRANCO", "MEIO_CAMPO")],
            )

    def test_simple_participant_text_import_creates_matches_and_draw_orders(self):
        with app.app_context():
            db = get_db()
            player_ids = {}
            for name in ("EDVAL", "DIEGO", "NEWTON", "WALTER", "REGIO", "BARBOZA", "LUCCA"):
                player_ids[name] = db.execute(
                    "INSERT INTO players(name,war_name) VALUES(?,?)", (name.title(), name)
                ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','RASCUNHO',?)",
                ("2026-08-29", self.user_id),
            ).lastrowid
            db.execute("INSERT INTO football_matches(sumula_id,number) VALUES(?,1)", (sumula_id,))
            db.commit()

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        detail = self.client.get(f"/futebol/sumulas/{sumula_id}")
        detail_html = detail.get_data(as_text=True)
        self.assertIn("Importar participantes", detail_html)
        self.assertIn("Relação dos participantes", detail_html)
        self.assertIn("Baixar Excel simples", detail_html)
        template = self.client.get("/futebol/sumulas/modelo-importacao.xlsx")
        self.assertEqual(template.status_code, 200)
        workbook = load_workbook(BytesIO(template.data))
        self.assertEqual(workbook.sheetnames, ["Instruções", "Participantes", "Peladeiros"])
        self.assertEqual(
            tuple(cell.value for cell in workbook["Participantes"][1]),
            ("Partida", "Posição", "ID do peladeiro", "Nome do peladeiro", "Status", "Observação"),
        )
        participant_text = """1ª PARTIDA
G1: EDVAL
D1: DIEGO
M1: NEWTON
A1: WALTER
2ª PARTIDA
D1: REGIO
3ª PARTIDA
M1: BARBOZA
A5: LUCCA"""
        imported = self.client.post(
            f"/futebol/sumulas/{sumula_id}/importar-participantes-texto",
            data={"participant_text": participant_text},
            follow_redirects=True,
        )
        html = imported.get_data(as_text=True)
        self.assertEqual(imported.status_code, 200)
        self.assertIn("7 participantes importados em 3 partida(s)", html)
        with app.app_context():
            db = get_db()
            rows = db.execute(
                """SELECT p.war_name,fp.draw_order,fp.preferred_position
                   FROM football_participants fp JOIN players p ON p.id=fp.player_id
                   WHERE fp.sumula_id=? ORDER BY fp.draw_order""",
                (sumula_id,),
            ).fetchall()
            self.assertEqual(
                [(row["war_name"], row["draw_order"], row["preferred_position"]) for row in rows],
                [
                    ("EDVAL", 1, "GOLEIRO"), ("DIEGO", 3, "DEFENSOR"),
                    ("NEWTON", 11, "MEIO_CAMPO"), ("WALTER", 17, "ATACANTE"),
                    ("REGIO", 25, "DEFENSOR"), ("BARBOZA", 55, "MEIO_CAMPO"),
                    ("LUCCA", 65, "ATACANTE"),
                ],
            )
            self.assertEqual(db.execute("SELECT COUNT(*) FROM football_matches WHERE sumula_id=?", (sumula_id,)).fetchone()[0], 3)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM football_lineups").fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM football_goals").fetchone()[0], 0)

    def test_invalid_participant_text_import_writes_nothing(self):
        with app.app_context():
            db = get_db()
            second_id = db.execute("INSERT INTO players(name,war_name) VALUES(?,?)", ("Segundo", "Segundo")).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','RASCUNHO',?)",
                ("2026-09-05", self.user_id),
            ).lastrowid
            db.execute("INSERT INTO football_matches(sumula_id,number) VALUES(?,1)", (sumula_id,))
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        rejected = self.client.post(
            f"/futebol/sumulas/{sumula_id}/importar-participantes-texto",
            data={"participant_text": "1ª PARTIDA\nD1: Peladeiro\nD1: Segundo"},
            follow_redirects=True,
        )
        self.assertIn("posição D1 repetida na partida 1", rejected.get_data(as_text=True))
        with app.app_context():
            db = get_db()
            self.assertEqual(db.execute("SELECT COUNT(*) FROM football_participants WHERE sumula_id=?", (sumula_id,)).fetchone()[0], 0)
            self.assertEqual(db.execute("SELECT COUNT(*) FROM football_goals").fetchone()[0], 0)

    def test_result_text_imports_scores_goals_assists_referee_and_cards(self):
        names = ("FERNANDES", "WALTER", "ANDRÉ", "CAZOLARE", "GERMANO", "JANDER", "KIJARA", "JONY", "REGIO", "NILSON")
        with app.app_context():
            db = get_db()
            player_ids = {}
            for name in names:
                player_ids[name] = db.execute(
                    "INSERT INTO players(name,war_name) VALUES(?,?)", (name.title(), name)
                ).lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','RASCUNHO',?)",
                ("2026-09-12", self.user_id),
            ).lastrowid
            match_ids = {}
            for number in (1, 2):
                match_ids[number] = db.execute(
                    "INSERT INTO football_matches(sumula_id,number) VALUES(?,?)", (sumula_id, number)
                ).lastrowid
            for order, player_id in enumerate(player_ids.values(), start=1):
                db.execute(
                    "INSERT INTO football_participants(sumula_id,player_id,status,draw_order) VALUES(?,?,'CONFIRMADO',?)",
                    (sumula_id, player_id, order),
                )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id

        page = self.client.get(f"/futebol/sumulas/{sumula_id}").get_data(as_text=True)
        self.assertIn("Importar resultados, gols e ocorrências", page)
        self.assertNotIn('id="participant-text"', page)
        result_text = """1ª PARTIDA (AZUL 1 x 1 BRANCO)

Gols e Assistências:
Time Azul: Gol de FERNANDES (Assistência: WALTER)
Time Branco: Gol de ANDRÉ (Assistência: CAZOLARE)
2ª PARTIDA (AZUL 0 x 2 BRANCO)

Gols e Assistências:
Time Branco:
1º Gol: GERMANO (Assistência: JANDER)
2º Gol: KIJARA (Assistência: JONY)
Juízes:
REGIO
Cartão Amarelo:
REGIO
NILSON"""
        imported = self.client.post(
            f"/futebol/sumulas/{sumula_id}/importar-resultados-texto",
            data={"result_text": result_text},
            follow_redirects=True,
        )
        html = imported.get_data(as_text=True)
        self.assertEqual(imported.status_code, 200)
        self.assertIn("Resultados importados: 2 partida(s), 4 gol(s), 1 juiz(es) e 2 cartão(ões)", html)
        with app.app_context():
            db = get_db()
            scores = db.execute(
                "SELECT number,blue_score,white_score,status FROM football_matches WHERE sumula_id=? ORDER BY number",
                (sumula_id,),
            ).fetchall()
            self.assertEqual(
                [(row["number"], row["blue_score"], row["white_score"], row["status"]) for row in scores],
                [(1, 1, 1, "ENCERRADA"), (2, 0, 2, "ENCERRADA")],
            )
            goals = db.execute(
                "SELECT author_player_id,assist_player_id,benefited_team FROM football_goals ORDER BY id"
            ).fetchall()
            self.assertEqual(
                [(row["author_player_id"], row["assist_player_id"], row["benefited_team"]) for row in goals],
                [
                    (player_ids["FERNANDES"], player_ids["WALTER"], "AZUL"),
                    (player_ids["ANDRÉ"], player_ids["CAZOLARE"], "BRANCO"),
                    (player_ids["GERMANO"], player_ids["JANDER"], "BRANCO"),
                    (player_ids["KIJARA"], player_ids["JONY"], "BRANCO"),
                ],
            )
            referee = db.execute("SELECT match_id,player_id,responsibility_type FROM football_responsibles").fetchone()
            self.assertEqual((referee["match_id"], referee["player_id"], referee["responsibility_type"]), (match_ids[2], player_ids["REGIO"], "ARBITRO_VOLUNTARIO"))
            cards = db.execute("SELECT match_id,player_id,card FROM football_incidents ORDER BY id").fetchall()
            self.assertEqual(
                [(row["match_id"], row["player_id"], row["card"]) for row in cards],
                [(match_ids[2], player_ids["REGIO"], "AMARELO"), (match_ids[2], player_ids["NILSON"], "AMARELO")],
            )

    def test_result_text_suggests_similar_name_and_is_atomic(self):
        with app.app_context():
            db = get_db()
            author_id = db.execute("INSERT INTO players(name,war_name) VALUES('André','ANDRÉ')").lastrowid
            assistant_id = db.execute("INSERT INTO players(name,war_name) VALUES('Cazolare','CAZOLARE')").lastrowid
            sumula_id = db.execute(
                "INSERT INTO football_sumulas(match_date,day_pelada,situacao,created_by) VALUES(?,'SABADO','RASCUNHO',?)",
                ("2026-09-19", self.user_id),
            ).lastrowid
            db.execute("INSERT INTO football_matches(sumula_id,number) VALUES(?,1)", (sumula_id,))
            for order, player_id in enumerate((author_id, assistant_id), start=1):
                db.execute(
                    "INSERT INTO football_participants(sumula_id,player_id,status,draw_order) VALUES(?,?,'CONFIRMADO',?)",
                    (sumula_id, player_id, order),
                )
            db.commit()
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        rejected = self.client.post(
            f"/futebol/sumulas/{sumula_id}/importar-resultados-texto",
            data={"result_text": "1ª PARTIDA (AZUL 0 x 1 BRANCO)\nGols e Assistências:\nTime Branco: Gol de ANDRÉ (Assistência: CAZOLARI)"},
            follow_redirects=True,
        )
        html = rejected.get_data(as_text=True)
        self.assertIn("Você quis dizer “CAZOLARE”?", html)
        with app.app_context():
            db = get_db()
            match = db.execute("SELECT blue_score,white_score,status FROM football_matches WHERE sumula_id=?", (sumula_id,)).fetchone()
            self.assertEqual((match["blue_score"], match["white_score"], match["status"]), (0, 0, "PLANEJADA"))
            self.assertEqual(db.execute("SELECT COUNT(*) FROM football_goals").fetchone()[0], 0)

    def test_stage3_session_load_is_read_only_explicit_and_one_connection(self):
        statements = []
        original_execute = DbWrapper.execute
        real_connect = connect_db

        def record(wrapper, statement, params=()):
            statements.append(" ".join(str(statement).split()))
            return original_execute(wrapper, statement, params)

        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with patch.object(DbWrapper, "execute", new=record), patch(
            "src.db.connect_db", wraps=real_connect
        ) as connect:
            response = self.client.get("/players")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(connect.call_count, 1)
        session_queries = [sql for sql in statements if "FROM users WHERE id=" in sql]
        self.assertEqual(len(session_queries), 1)
        self.assertNotIn("SELECT *", session_queries[0].upper())
        self.assertFalse(any(re.match(r"^(INSERT|UPDATE|DELETE)", sql, re.I) for sql in statements))
        self.assertFalse(any("SELECT 1 FROM users LIMIT 1" in sql for sql in statements))

    def test_stage3_retry_only_transient_and_auth_error_is_not_retried(self):
        class DbError(RuntimeError):
            def __init__(self, sqlstate=None, message="database failure"):
                super().__init__(message)
                self.pgcode = sqlstate

        class FakeDb:
            def __init__(self, errors):
                self.errors = list(errors)
                self.execute_count = 0
                self.rollback_count = 0

            def execute(self, _sql, _params):
                self.execute_count += 1
                if self.errors:
                    raise self.errors.pop(0)
                return self

            def fetchone(self):
                return {"id": 1}

            def rollback(self):
                self.rollback_count += 1

        retryable_errors = (
            DbError("40001"),
            DbError("40P01"),
            DbError(message="tuple concurrently updated"),
        )
        for error in retryable_errors:
            with self.subTest(error=str(error), sqlstate=error.pgcode):
                transient = FakeDb([error])
                with app.test_request_context("/players"), patch(
                    "src.db.get_db", return_value=transient
                ), patch("src.db.time.sleep") as sleep:
                    self.assertEqual(read_user_from_session(1)["id"], 1)
                self.assertEqual(transient.execute_count, 2)
                self.assertEqual(transient.rollback_count, 1)
                sleep.assert_called_once()

        permanent_errors = (
            DbError("28P01", "password authentication failed"),
            DbError("42P01", "relation does not exist"),
            DbError(None, "connection timeout"),
        )
        for error in permanent_errors:
            with self.subTest(no_retry=str(error), sqlstate=error.pgcode):
                database = FakeDb([error])
                with app.test_request_context("/players"), patch("src.db.get_db", return_value=database):
                    with self.assertRaises(DbError):
                        read_user_from_session(1)
                self.assertEqual(database.execute_count, 1)

    def test_stage3_health_and_safe_database_logs(self):
        response = self.client.get("/health")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"database": "ok", "status": "ok"})

        secret = "postgresql://admin:senha-secreta@example.test/db"
        class TemporaryConnectionError(RuntimeError):
            pgcode = "08006"

        with self.assertLogs(app.logger.name, level="ERROR") as captured, patch(
            "app.get_db", side_effect=TemporaryConnectionError(secret)
        ):
            failed = self.client.get("/health")
        self.assertEqual(failed.status_code, 503)
        logs = "\n".join(captured.output)
        self.assertIn("DB_HEALTHCHECK_ERROR", logs)
        self.assertNotIn(secret, logs)
        self.assertNotIn("senha-secreta", logs)

        class SchemaError(RuntimeError):
            pgcode = "42P01"

        with patch("app.get_db", side_effect=SchemaError("relation does not exist")):
            permanent = self.client.get("/health")
        self.assertEqual(permanent.status_code, 500)

    def test_stage3_teardown_rolls_back_after_error_and_closes(self):
        class FakeConnection:
            def __init__(self):
                self.rolled_back = False
                self.closed = False

            def rollback(self):
                self.rolled_back = True

            def close(self):
                self.closed = True

        connection = FakeConnection()
        with self.assertRaisesRegex(RuntimeError, "boom"):
            with app.app_context():
                from flask import g
                g.db = connection
                raise RuntimeError("boom")
        self.assertTrue(connection.rolled_back)
        self.assertTrue(connection.closed)

    def test_stage4_sqlite_connect_is_schema_free(self):
        database = str(Path(self.tempdir.name) / "connection-only.db")
        config = type("AppConfig", (), {"config": {"DATABASE_URL": None, "DATABASE": database}})()
        with patch.dict(os.environ, {"DATABASE_URL": "", "SUPABASE_DB_URL": ""}), patch(
            "src.db.init_sqlite"
        ) as schema_setup:
            connection = connect_db(config)
        try:
            tables = connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        finally:
            connection.close()
        schema_setup.assert_not_called()
        self.assertEqual(tables, [])

    def test_stage4_explicit_sqlite_setup_is_idempotent_and_preserves_data(self):
        database = str(Path(self.tempdir.name) / "explicit-setup.db")
        first_path = initialize_sqlite_database(database)
        first = connect_db(type("Config", (), {"config": {"DATABASE_URL": None, "DATABASE": database}})())
        first.execute(
            "INSERT INTO users(username,name,password_hash,role) VALUES(?,?,?,'manager')",
            ("preservado", "Preservado", "hash"),
        )
        first.commit()
        first.close()

        second_path = initialize_sqlite_database(database)
        second = connect_db(type("Config", (), {"config": {"DATABASE_URL": None, "DATABASE": database}})())
        try:
            row = second.execute("SELECT name FROM users WHERE username=?", ("preservado",)).fetchone()
        finally:
            second.close()
        self.assertEqual(first_path, second_path)
        self.assertEqual(row["name"], "Preservado")

    def test_stage4_setup_and_migrations_never_run_during_requests(self):
        with self.client.session_transaction() as session:
            session["user_id"] = self.user_id
        with patch("src.db.init_sqlite") as sqlite_setup, patch(
            "src.db.run_postgres_migrations"
        ) as postgres_setup:
            response = self.client.get("/players")
        self.assertEqual(response.status_code, 200)
        sqlite_setup.assert_not_called()
        postgres_setup.assert_not_called()

        with app.test_request_context("/health"):
            with self.assertRaisesRegex(RuntimeError, "requisição HTTP"):
                initialize_sqlite_database(str(Path(self.tempdir.name) / "forbidden.db"))
            with self.assertRaisesRegex(RuntimeError, "requests HTTP"):
                run_postgres_migrations("postgresql://redacted.invalid/db")

    def test_stage4_commands_require_explicit_execution(self):
        from scripts.init_local_db import main as init_local_main
        from scripts.migrate_postgres_schema import main as postgres_migration_main

        database = str(Path(self.tempdir.name) / "command.db")
        self.assertEqual(init_local_main(["--database", database]), 0)
        connection = connect_db(type("Config", (), {"config": {"DATABASE_URL": None, "DATABASE": database}})())
        try:
            self.assertIsNotNone(connection.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='users'"
            ).fetchone())
        finally:
            connection.close()

        with patch.dict(os.environ, {
            "DATABASE_URL": "postgresql://redacted.invalid/db",
            "SUPABASE_DB_URL": "",
            "APPLY_POSTGRES_MIGRATIONS": "",
        }), patch("scripts.migrate_postgres_schema.run_postgres_migrations") as runner:
            with self.assertRaises(SystemExit):
                postgres_migration_main()
            runner.assert_not_called()

        with patch.dict(os.environ, {
            "DATABASE_URL": "postgresql://redacted.invalid/db",
            "SUPABASE_DB_URL": "",
            "APPLY_POSTGRES_MIGRATIONS": "1",
        }), patch("scripts.migrate_postgres_schema.run_postgres_migrations", return_value=[]) as runner:
            postgres_migration_main()
            runner.assert_called_once_with("postgresql://redacted.invalid/db")

    def test_stage5_postgres_runner_orders_skips_records_and_closes_with_mocks(self):
        migration_dir = Path(self.tempdir.name) / "migrations"
        migration_dir.mkdir()
        (migration_dir / "001_applied.sql").write_text("SELECT 10;", encoding="utf-8")
        (migration_dir / "002_empty.sql").write_text("", encoding="utf-8")
        (migration_dir / "003_pending.sql").write_text("SELECT 30;", encoding="utf-8")

        class FakeConnection:
            def __init__(self):
                self.applied = {"001_applied.sql"}
                self.statements = []
                self.commits = 0
                self.rollbacks = 0
                self.closed = False
                self.fail_sql = None

            def cursor(self):
                connection = self

                class Cursor:
                    result = None

                    def __enter__(self):
                        return self

                    def __exit__(self, *_args):
                        return False

                    def execute(self, sql, params=None):
                        normalized = " ".join(str(sql).split())
                        connection.statements.append((normalized, params))
                        if normalized == connection.fail_sql:
                            raise RuntimeError("falha simulada")
                        if "to_regclass" in normalized:
                            self.result = ("users",)
                        elif normalized.startswith("SELECT 1 FROM schema_migrations"):
                            self.result = (1,) if params[0] in connection.applied else None
                        elif normalized.startswith("INSERT INTO schema_migrations"):
                            connection.applied.add(params[0])

                    def fetchone(self):
                        return self.result

                return Cursor()

            def commit(self):
                self.commits += 1

            def rollback(self):
                self.rollbacks += 1

            def close(self):
                self.closed = True

        connection = FakeConnection()
        with patch("psycopg2.connect", return_value=connection):
            applied = run_postgres_migrations("postgresql://redacted.invalid/db", migration_dir)

        self.assertEqual(applied, ["002_empty.sql", "003_pending.sql"])
        self.assertEqual(connection.commits, 3)
        self.assertEqual(connection.rollbacks, 0)
        self.assertTrue(connection.closed)
        self.assertIn("002_empty.sql", connection.applied)
        self.assertIn("003_pending.sql", connection.applied)
        self.assertFalse(any(sql == "SELECT 10;" for sql, _params in connection.statements))
        self.assertTrue(any(sql == "SELECT 30;" for sql, _params in connection.statements))
        self.assertTrue(any("pg_advisory_lock" in sql for sql, _params in connection.statements))
        self.assertTrue(any("pg_advisory_unlock" in sql for sql, _params in connection.statements))

        failed_connection = FakeConnection()
        failed_connection.applied.add("002_empty.sql")
        failed_connection.fail_sql = "SELECT 30;"
        with patch("psycopg2.connect", return_value=failed_connection):
            with self.assertRaisesRegex(RuntimeError, "falha simulada"):
                run_postgres_migrations("postgresql://redacted.invalid/db", migration_dir)
        self.assertEqual(failed_connection.rollbacks, 1)
        self.assertTrue(failed_connection.closed)
        self.assertNotIn("003_pending.sql", failed_connection.applied)


if __name__ == "__main__":
    unittest.main()
